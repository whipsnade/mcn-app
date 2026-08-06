"""Gate C 受控 Excel 模板生成：从用户来源模板清洗为 brand_report_v3.xlsx /
kol_selection_v3.xlsx / campaign_report_v2.xlsx。

模板只承载版式（标题锚点、表头、列宽、合并、打印设置）；样例数据与图表被
清除（图表由导出器现场重建），数据区由导出器按固定行号填充。固定模板版本
写入隐藏 metadata（定义名 ``TEMPLATE_VERSION`` 指向首表隐藏单元格）。

行号契约（与 ``exporters/*.py`` 一一对应，改布局必须同步改渲染器并更新测试）：
- brand 8 Sheet：
  综合概览 标题@1 + 周期@A2/B2 + 来源@A3/B3 + 搜索方式@A4/B4 + 指标表头@6
  （数据@7-10）；情感分析/日趋势/内容类型与达人/地域分布/热门帖子TOP/
  舆情洞察/方法论 标题@1 + 表头@3（渲染器重写），数据自 4 行起。
- kol 4 Sheet（首表重命名为「达人圈选总表」）：
  达人圈选总表 标题@1 + 元数据@2 + 表头@4，数据自 5 行起；
  达人详细画像 全表由渲染器自 1 行写 Top20 详情块；
  粉丝画像详情 表头@1，数据自 2 行起；
  评分方法论与数据来源 标题@1 + 章节@3 + 表头@4，数据自 5 行起。

用法：
cd backend
.venv/bin/python scripts/build_agent_artifact_templates.py \
  --brand-source /path/brand_report.xlsx \
  --kol-source '/path/KOL匹配度分析报告.xlsx' \
  [--output-dir app/agent_artifacts/templates]
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

BACKEND = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND))

TEMPLATE_VERSION = "gate-c-v1"
DEFAULT_OUT = BACKEND / "app" / "agent_artifacts" / "templates"

# 每个 brand Sheet 的数据清理起点：保留标题@1 与（如适用）固定表头行。
BRAND_CLEAR_FROM = {
    "综合概览": 5,  # 保留标题/周期/来源/搜索方式锚点，清 5 行起
    "情感分析": 2,
    "日趋势": 2,
    "内容类型与达人": 2,
    "地域分布": 2,
    "热门帖子TOP": 2,
    "舆情洞察": 2,
    "方法论": 2,
}

# kol Sheet 清理起点（首表重命名后处理）。
KOL_CLEAR_FROM = {
    "达人圈选总表": 3,  # 保留标题/元数据，清 3 行起（表头@4 由渲染器重写）
    "达人详细画像": 1,
    "粉丝画像详情": 1,  # 表头@1 由渲染器重写
    "评分方法论与数据来源": 5,  # 保留标题/章节/表头@4
}


def _strip_art(ws: Worksheet) -> None:
    ws._charts = []
    ws._images = []


def _clear_data(ws: Worksheet, start_row: int) -> None:
    from openpyxl.cell.cell import MergedCell

    for row in range(start_row, ws.max_row + 1):
        for column in range(1, ws.max_column + 1):
            cell = ws.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _write_template_metadata(wb, sheet_name: str) -> None:
    """隐藏 metadata：定义名 TEMPLATE_VERSION 指向首表 A1000。"""
    ws = wb[sheet_name]
    ws.cell(1000, 1).value = TEMPLATE_VERSION
    ws.cell(1000, 2).value = "template_version"
    from openpyxl.workbook.defined_name import DefinedName

    wb.defined_names["TEMPLATE_VERSION"] = DefinedName(
        "TEMPLATE_VERSION", attr_text=f"'{sheet_name}'!$A$1000"
    )


def clean_source_template(source: Path, sheet_clear: dict[str, int], *, sheet_renames: dict[str, str] | None = None) -> None:
    if not source.exists():
        raise SystemExit(f"source template missing: {source}")
    wb = load_workbook(source)
    for ws in list(wb.worksheets):
        _strip_art(ws)
    renames = sheet_renames or {}
    for ws in list(wb.worksheets):
        new_title = renames.get(ws.title)
        lookup = new_title or ws.title
        clear_from = sheet_clear.get(lookup)
        if clear_from is None:
            # 未知 Sheet 直接删除：受控模板只保留契约 Sheet。
            wb.remove(ws)
            continue
        _clear_data(ws, clear_from)
        if new_title:
            ws.title = new_title
    # 首表写模板版本 metadata。
    _write_template_metadata(wb, next(iter(sheet_clear)))
    return wb


def build_brand(source: Path, output_dir: Path) -> Path:
    wb = clean_source_template(source, BRAND_CLEAR_FROM)
    target = output_dir / "brand_report_v3.xlsx"
    wb.save(target)
    return target


def build_kol(source: Path, output_dir: Path) -> Path:
    renames = {"小红书KOL匹配度筛选": "达人圈选总表"}
    wb = clean_source_template(source, KOL_CLEAR_FROM, sheet_renames=renames)
    target = output_dir / "kol_selection_v3.xlsx"
    wb.save(target)
    return target


def build_campaign(output_dir: Path) -> Path:
    """程序化生成活动模板（9 基础 Sheet；ROI Sheet 由导出器按数据条件插入）。

    视觉规范：标题 #1F4E79、交替行 #D6E4F0、微软雅黑、统一打印区。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1F4E79")
    SECTION_FONT = Font(name="微软雅黑", bold=True, size=11, color="1F4E79")
    HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
    HEADER_FONT = Font(name="微软雅黑", bold=True, size=10)
    THIN = Side(style="thin", color="B0B0B0")
    HEADER_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ALTERNATE_FILL = PatternFill("solid", fgColor="D6E4F0")
    CENTER = Alignment(horizontal="center", vertical="center")

    wb = Workbook()
    wb.remove(wb.active)

    sheet_specs = [
        ("活动综合概览", 6),
        ("周期对比与趋势", 6),
        ("平台表现", 6),
        ("情感与内容分析", 6),
        ("热门帖子TOP", 8),
        ("达人投放表现", 6),
        ("自然传播与受众", 6),
        ("洞察与建议", 4),
        ("方法论", 4),
    ]
    for name, columns in sheet_specs:
        ws = wb.create_sheet(name)
        for column in range(1, columns + 1):
            ws.column_dimensions[chr(64 + column)].width = 16
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
        title_cell = ws.cell(1, 1)
        title_cell.value = name
        title_cell.font = TITLE_FONT
        ws.sheet_view.showGridLines = False
        ws.print_options.horizontalCentered = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.orientation = "landscape"

    # 表头样式（各 Sheet 第 3 行：渲染器重写内容但保留样式锚点）。
    for ws in wb.worksheets:
        for column in range(1, 9):
            cell = ws.cell(3, column)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = HEADER_BORDER
            cell.alignment = CENTER

    _write_template_metadata(wb, "活动综合概览")
    target = output_dir / "campaign_report_v2.xlsx"
    wb.save(target)
    return target


def main() -> None:
    parser = argparse.ArgumentParser(description="清洗来源模板为受控导出模板")
    parser.add_argument("--brand-source", required=True, help="品牌模板来源 xlsx")
    parser.add_argument("--kol-source", required=True, help="达人模板来源 xlsx")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUT), help="输出目录")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    brand = build_brand(Path(args.brand_source), output_dir)
    kol = build_kol(Path(args.kol_source), output_dir)
    campaign = build_campaign(output_dir)
    print(brand)
    print(kol)
    print(campaign)


if __name__ == "__main__":
    main()
