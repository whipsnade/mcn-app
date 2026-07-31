"""一次性模板生成脚本：仓库根 brand_report.xlsx → app/reporting/templates/brand_report_v2.xlsx。

做法：逐 Sheet 清除样例数据行（保留表头行、合并拓扑、列宽、数字格式、章节标题），
删除 3 个样例图表，移除样例品牌「昊来了」与「Python openpyxl」等旧生成说明文字。
脚本入库，产物模板随仓库提交；模板结构变化时改本脚本重跑。

运行：cd backend && .venv/bin/python scripts/build_brand_report_template.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

SOURCE_PATH = Path(__file__).resolve().parents[2] / "brand_report.xlsx"
TARGET_PATH = (
    Path(__file__).resolve().parents[1] / "app" / "reporting" / "templates" / "brand_report_v2.xlsx"
)

SHEET_ORDER = (
    "综合概览", "情感分析", "日趋势", "内容类型与达人",
    "地域分布", "热门帖子TOP", "舆情洞察", "方法论",
)

# 逐 Sheet 需要清空的样例数据区域（(起始行, 结束行, 列数)，含首尾行）；
# 区域外的表头行、章节标题、合并拓扑、列宽与数字格式全部保留。
CLEAR_RANGES: dict[str, list[tuple[int, int, int]]] = {
    "综合概览": [
        (1, 1, 6),      # A1 标题含样例品牌，渲染器重写
        (2, 4, 6),      # A2:B4 分析周期/数据来源/搜索方式样例值
        (7, 15, 6),     # 指标矩阵样例值与跨行公式（A 列行标签随后由渲染器重写/清除）
        (19, 22, 6),    # 环比同比样例值
        (26, 28, 6),    # 声量构成样例值与占比公式
    ],
    "情感分析": [
        (4, 9, 5),      # 平台×情感样例行与占比公式
        (12, 17, 5),    # 样例关键发现
    ],
    "日趋势": [
        (1, 1, 4),      # A1 标题含样例周期，渲染器重写
        (4, 35, 4),     # 日数据样例 + 合计行 SUM 公式
    ],
    "内容类型与达人": [
        (4, 12, 4),     # 内容类型样例行
        (16, 19, 4),    # 达人层级样例行
        (23, 24, 4),    # 商单 vs 自然样例行与占比公式
    ],
    "地域分布": [
        (4, 23, 4),     # 地域样例行与占比公式
    ],
    "热门帖子TOP": [
        (4, 18, 12),    # 小红书段样例帖
        (24, 38, 12),   # 抖音段样例帖
    ],
    "舆情洞察": [
        (5, 11, 4),     # 好评样例行
        (15, 17, 4),    # 槽点样例行
        (21, 25, 4),    # 扩张信号样例行
        (28, 29, 4),    # 含样例品牌的噪音说明
    ],
    "方法论": [
        (4, 15, 2),     # 全部样例口径说明（含「Python openpyxl」生成说明）
    ],
}

# 模板中不得残留的样例痕迹。
FORBIDDEN_FRAGMENTS = ("昊来了", "李昊", "openpyxl", "Python")


def _clear_range(sheet, start_row: int, end_row: int, columns: int) -> None:
    for row in range(start_row, end_row + 1):
        for column in range(1, columns + 1):
            cell = sheet.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None


def build_template() -> Path:
    if not SOURCE_PATH.exists():
        raise SystemExit(f"样例工作簿不存在：{SOURCE_PATH}")
    workbook = load_workbook(SOURCE_PATH)
    if tuple(workbook.sheetnames) != SHEET_ORDER:
        raise SystemExit(f"样例 Sheet 结构不符：{workbook.sheetnames}")
    for name in workbook.sheetnames:
        sheet = workbook[name]
        for start_row, end_row, columns in CLEAR_RANGES[name]:
            _clear_range(sheet, start_row, end_row, columns)
        # 删除样例图表（日趋势 2 个 LineChart、地域 1 个 BarChart）；
        # 渲染器按 payload 数据决定是否新建图表。
        sheet._charts = []
    # 自检：无样例品牌/旧生成说明残留、无图表残留。
    for name in workbook.sheetnames:
        sheet = workbook[name]
        if sheet._charts:
            raise SystemExit(f"图表清除失败：{name}")
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for fragment in FORBIDDEN_FRAGMENTS:
                        if fragment in cell.value:
                            raise SystemExit(f"样例残留 {fragment!r}：{name}!{cell.coordinate}")
    TARGET_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(TARGET_PATH)
    return TARGET_PATH


if __name__ == "__main__":
    target = build_template()
    print(f"模板已生成：{target}")
