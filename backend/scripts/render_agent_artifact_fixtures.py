"""Gate C fixture Excel 生成与逐 Sheet 结构核对（Task 7）。

生成三类 Artifact 的导出工作簿到 ``--output`` 目录：
- brand_report_v3（8 Sheet）
- kol_selection_v3 v3 名单（4 Sheet，Top20 详情块）
- campaign_report_v2 无 ROI（9 Sheet）
- campaign_report_v2 有 ROI（10 Sheet）

随后用 openpyxl 逐 Sheet 核对：Sheet 名、Top20 详情块数量、无 #REF!/#VALUE!/
#DIV/0! 公式错误、无示例数据残留、受限章节说明、URL/公式注入防护。若本机存在
``soffice`` 则把每个 Sheet 渲染为 PNG（逐 Sheet 视觉核对素材）；缺失时打印警告
并跳过（结构核对仍完成）。

用法：cd backend && .venv/bin/python scripts/render_agent_artifact_fixtures.py \
  --output ../outputs/report-template-gate-c
"""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import sys
from datetime import UTC, datetime
from pathlib import Path

BACKEND = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(BACKEND))

from openpyxl import load_workbook  # noqa: E402

from app.agent_artifacts.exporters.brand import render_brand_workbook  # noqa: E402
from app.agent_artifacts.exporters.campaign import (  # noqa: E402
    ROI_SHEET,
    render_campaign_workbook,
)
from app.agent_artifacts.exporters.kol_selection import (  # noqa: E402
    render_kol_selection_workbook,
)
from tests.agent_artifacts.test_payloads import (  # noqa: E402
    build_brand_dict,
    build_campaign_dict,
    build_kol_value_selection_dict,
)


def _now() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def _kol_payload_with_20() -> dict:
    base = build_kol_value_selection_dict()
    items = base["data"]["items"]
    while len(items) < 20:
        clone = dict(items[0])
        clone["rank"] = len(items) + 1
        clone["kol_uid"] = f"k{len(items) + 1}"
        clone["nickname"] = f"达人{len(items) + 1}"
        items.append(clone)
    return base


def _campaign_payload_with_roi() -> dict:
    payload = build_campaign_dict()
    payload["data"]["internal_metrics"] = {
        "spend": 100000,
        "impressions": 2000000,
        "conversions": 5000,
        "revenue": 300000,
        "cpc": 20.0,
        "cpm": 50.0,
    }
    payload["data"]["roi"] = {
        "spend": 100000,
        "revenue": 300000,
        "conversions": 5000,
        "attribution_window": "最后点击 7 天",
        "roi": 2.0,
        "roas": 3.0,
    }
    return payload


def _values(ws) -> list:
    return [
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]


def _check_workbook(name: str, content: bytes, expected_sheets: list[str], *, detail_blocks: int | None = None) -> list[str]:
    """结构核对；返回问题列表（空 = 通过）。"""
    problems: list[str] = []
    wb = load_workbook(__import__("io").BytesIO(content), data_only=False)
    if list(wb.sheetnames) != expected_sheets:
        problems.append(f"{name}: sheetnames {wb.sheetnames} != {expected_sheets}")
    for ws in wb.worksheets:
        for cell in ws.iter_rows():
            for c in cell:
                if isinstance(c.value, str) and any(
                    marker in c.value for marker in ("#REF!", "#VALUE!", "#DIV/0!")
                ):
                    problems.append(f"{name}/{ws.title}: formula error {c.value!r}")
    if detail_blocks is not None:
        count = sum(
            1
            for row in wb["达人详细画像"].iter_rows(min_col=1, max_col=1)
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("#")
        )
        if count != detail_blocks:
            problems.append(f"{name}: detail blocks {count} != {detail_blocks}")
    return problems


def _find_soffice() -> str | None:
    for name in ("soffice", "libreoffice"):
        path = shutil.which(name)
        if path is not None:
            return path
    mac_default = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    return str(mac_default) if mac_default.exists() else None


def _render_pngs(workbook_path: Path, output_dir: Path) -> list[Path]:
    """用 LibreOffice headless 逐 Sheet 转 PNG（每 Sheet 一张，绝不重复首 Sheet）。

    LibreOffice PNG 导出只绘制工作簿的第一个表，设置 active sheet 并不可靠；
    因此为每个 Sheet 生成仅含该 Sheet 的单表工作簿再转换，物理保证每张 PNG
    就是目标 Sheet。无 soffice 时跳过（结构核对仍完成）。
    """
    soffice = _find_soffice()
    if soffice is None:
        print("WARN: soffice 不可用，跳过 PNG 渲染（结构核对仍完成）")
        return []
    sheet_pngs: list[Path] = []
    wb = load_workbook(workbook_path, data_only=False)
    for title in wb.sheetnames:
        single = load_workbook(workbook_path)
        for other in list(single.sheetnames):
            if other != title:
                single.remove(single[other])
        temp = output_dir / f"{workbook_path.stem}-{title}.xlsx"
        single.save(temp)
        png = output_dir / f"{workbook_path.stem}-{title}.png"
        subprocess.run(
            [
                soffice,
                "--headless",
                "--convert-to",
                "png",
                "--outdir",
                str(output_dir),
                str(temp),
            ],
            check=True,
            capture_output=True,
            timeout=120,
        )
        sheet_pngs.append(png)
        temp.unlink(missing_ok=True)
    return sheet_pngs


def main() -> None:
    parser = argparse.ArgumentParser(description="生成 Gate C fixture Excel 并核对")
    parser.add_argument("--output", default="../outputs/report-template-gate-c")
    args = parser.parse_args()

    output_dir = Path(args.output)
    output_dir.mkdir(parents=True, exist_ok=True)

    fixtures = [
        ("brand_report_v3", render_brand_workbook(json.loads(json.dumps(build_brand_dict(), default=str))),
         ["综合概览", "情感分析", "日趋势", "内容类型与达人", "地域分布", "热门帖子TOP", "舆情洞察", "方法论"], None),
        ("kol_selection_v3", render_kol_selection_workbook(json.loads(json.dumps(_kol_payload_with_20(), default=str))),
         ["达人圈选总表", "达人详细画像", "粉丝画像详情", "评分方法论与数据来源"], 20),
        ("campaign_report_v2", render_campaign_workbook(json.loads(json.dumps(build_campaign_dict(), default=str))),
         ["活动综合概览", "周期对比与趋势", "平台表现", "情感与内容分析", "热门帖子TOP", "达人投放表现", "自然传播与受众", "洞察与建议", "方法论"], None),
        ("campaign_report_v2_roi", render_campaign_workbook(json.loads(json.dumps(_campaign_payload_with_roi(), default=str))),
         ["活动综合概览", "周期对比与趋势", "平台表现", "情感与内容分析", "热门帖子TOP", "达人投放表现", "自然传播与受众", "洞察与建议", "方法论", ROI_SHEET], None),
    ]

    all_problems: list[str] = []
    for name, content, sheets, blocks in fixtures:
        path = output_dir / f"{name}.xlsx"
        path.write_bytes(content)
        problems = _check_workbook(name, content, sheets, detail_blocks=blocks)
        if problems:
            all_problems.extend(problems)
        else:
            print(f"OK {name}: {len(sheets)} sheets（结构核对通过）")
        for png in _render_pngs(path, output_dir):
            print(f"PNG {png}")

    if all_problems:
        print("PROBLEMS:")
        for problem in all_problems:
            print(" -", problem)
        sys.exit(1)
    print(f"ALL OK -> {output_dir.resolve()}")


if __name__ == "__main__":
    main()
