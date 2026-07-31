"""GET /sessions/{id}/reports/{report_id}/export：品牌报告 Excel 导出端点。

端点不注入模型依赖（只覆写 get_db 即可跑通），不调 MCP/积分系统。
"""

from __future__ import annotations

from collections.abc import AsyncIterator
from datetime import UTC, datetime, timedelta
from io import BytesIO
from urllib.parse import quote
from uuid import uuid4

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.billing.models import Wallet
from app.core.security import create_access_token
from app.db.session import get_db
from app.identity.models import LoginSession, User
from app.main import create_app
from app.reporting.models import AnalysisReport
from app.workspace.models import WorkspaceSession

from tests.reporting.test_brand_exporter import _payload


@pytest_asyncio.fixture
async def export_client_factory(db_session: AsyncSession):
    clients: list[AsyncClient] = []

    async def create() -> tuple[AsyncClient, User, str]:
        app = create_app()

        async def override_get_db() -> AsyncIterator[AsyncSession]:
            yield db_session

        app.dependency_overrides[get_db] = override_get_db
        now = datetime.now(UTC).replace(tzinfo=None)
        user = User(
            id=str(uuid4()),
            nickname="导出用户",
            role="user",
            status="active",
            created_at=now,
            updated_at=now,
        )
        db_session.add(user)
        await db_session.flush()
        db_session.add(
            Wallet(user_id=user.id, balance=1000, reserved=0, version=1, updated_at=now)
        )
        session = WorkspaceSession(
            id=str(uuid4()),
            user_id=user.id,
            title="导出会话",
            brand="肯德基",
            campaign_name=None,
            status="active",
            platforms=["xiaohongshu"],
            category="美食",
            target_audience="",
            budget_min=None,
            budget_max=None,
            filters_snapshot={},
            is_starred=False,
            last_accessed_at=now,
            created_at=now,
            updated_at=now,
        )
        db_session.add(session)
        login_session = LoginSession(
            id=str(uuid4()),
            user_id=user.id,
            refresh_token_hash=uuid4().hex + uuid4().hex,
            expires_at=now + timedelta(days=1),
            revoked_at=None,
            created_at=now,
            last_seen_at=now,
        )
        db_session.add(login_session)
        await db_session.flush()
        token = create_access_token(user_id=user.id, session_id=login_session.id, role="user")
        test_client = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
        test_client.headers["Authorization"] = f"Bearer {token}"
        clients.append(test_client)
        return test_client, user, session.id

    yield create
    for test_client in clients:
        await test_client.aclose()


def _seed_report(
    db_session: AsyncSession,
    session_id: str,
    *,
    report_type: str = "brand_analysis",
    template_version: str | None = "brand_report_v2",
    payload: dict | None = None,
    version: int = 1,
) -> AnalysisReport:
    now = datetime.now(UTC).replace(tzinfo=None)
    if payload is None and template_version == "brand_report_v2":
        payload = _payload().model_dump(mode="json")
    report = AnalysisReport(
        id=str(uuid4()),
        task_id=None,
        session_id=session_id,
        report_type=report_type,
        scope_json={},
        version=version,
        title="肯德基 品牌分析报告",
        blocks_json=[],
        conclusion_text=None,
        status="completed",
        payload_json=payload,
        template_version=template_version,
        created_at=now,
        updated_at=now,
    )
    db_session.add(report)
    return report


def _export_url(session_id: str, report_id: str) -> str:
    return f"/api/v1/sessions/{session_id}/reports/{report_id}/export"


@pytest.mark.asyncio
async def test_export_success_returns_workbook(
    export_client_factory, db_session: AsyncSession
) -> None:
    client, user, session_id = await export_client_factory()
    report = _seed_report(db_session, session_id, version=3)
    await db_session.flush()

    response = await client.get(_export_url(session_id, report.id))

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    expected_filename = "肯德基_品牌社媒分析报告_2026-06-01-2026-06-30_v3.xlsx"
    disposition = response.headers["content-disposition"]
    assert disposition.startswith("attachment;")
    assert quote(expected_filename) in disposition
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames[0] == "综合概览"
    assert workbook["综合概览"]["A1"].value == "肯德基 品牌社交媒体表现分析报告"
    # 零积分：钱包不变。
    wallet = await db_session.get(Wallet, user.id)
    assert wallet is not None
    assert wallet.balance == 1000
    assert wallet.reserved == 0


@pytest.mark.asyncio
async def test_export_foreign_user_returns_404(
    export_client_factory, db_session: AsyncSession
) -> None:
    _owner_client, _owner, session_id = await export_client_factory()
    report = _seed_report(db_session, session_id)
    await db_session.flush()
    other_client, _other, _other_session = await export_client_factory()

    response = await other_client.get(_export_url(session_id, report.id))

    assert response.status_code == 404
    assert response.json()["detail"] == "report_not_found"


@pytest.mark.asyncio
async def test_export_cross_session_report_returns_404(
    export_client_factory, db_session: AsyncSession
) -> None:
    client, _user, session_id = await export_client_factory()
    report = _seed_report(db_session, session_id)
    # 同一用户的第二个会话：report 挂在 session_id，用第二会话 id 请求应 404。
    now = datetime.now(UTC).replace(tzinfo=None)
    second_session = WorkspaceSession(
        id=str(uuid4()),
        user_id=_user.id,
        title="第二会话",
        brand="肯德基",
        campaign_name=None,
        status="active",
        platforms=["xiaohongshu"],
        category="美食",
        target_audience="",
        budget_min=None,
        budget_max=None,
        filters_snapshot={},
        is_starred=False,
        last_accessed_at=now,
        created_at=now,
        updated_at=now,
    )
    db_session.add(second_session)
    await db_session.flush()

    response = await client.get(_export_url(second_session.id, report.id))

    assert response.status_code == 404
    assert response.json()["detail"] == "report_not_found"


@pytest.mark.asyncio
async def test_export_kol_analysis_type_returns_404(
    export_client_factory, db_session: AsyncSession
) -> None:
    client, _user, session_id = await export_client_factory()
    report = _seed_report(
        db_session, session_id, report_type="kol_analysis", template_version=None, payload=None
    )
    await db_session.flush()

    response = await client.get(_export_url(session_id, report.id))

    assert response.status_code == 404
    assert response.json()["detail"] == "report_not_found"


@pytest.mark.asyncio
async def test_export_legacy_report_without_template_version_returns_404(
    export_client_factory, db_session: AsyncSession
) -> None:
    client, _user, session_id = await export_client_factory()
    report = _seed_report(db_session, session_id, template_version=None, payload=None)
    await db_session.flush()

    response = await client.get(_export_url(session_id, report.id))

    assert response.status_code == 404
    assert response.json()["detail"] == "report_not_found"


@pytest.mark.asyncio
async def test_export_corrupt_payload_returns_404(
    export_client_factory, db_session: AsyncSession
) -> None:
    client, _user, session_id = await export_client_factory()
    report = _seed_report(db_session, session_id, payload={"unexpected": True})
    await db_session.flush()

    response = await client.get(_export_url(session_id, report.id))

    assert response.status_code == 404
    assert response.json()["detail"] == "report_not_found"


@pytest.mark.asyncio
async def test_export_unknown_report_returns_404(
    export_client_factory, db_session: AsyncSession
) -> None:
    client, _user, session_id = await export_client_factory()

    response = await client.get(_export_url(session_id, str(uuid4())))

    assert response.status_code == 404
    assert response.json()["detail"] == "report_not_found"
