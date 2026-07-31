"""brand_report_v2 Excel 渲染器单测（brand_exporter，纯同步 openpyxl，不触库/模型）。"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart

from app.reporting.brand_exporter import (
    SHEET_ORDER,
    render_brand_workbook,
    sanitize_report_filename,
)
from app.reporting.brand_payload import (
    ALL_CHAPTERS,
    BrandReportData,
    BrandReportNarrative,
    BrandReportPayload,
    ChapterAvailability,
    ContentTypeRow,
    CreatorTierRow,
    DailyTrendSection,
    MetricComparison,
    OrganicVsPaid,
    OverviewSection,
    PeriodValue,
    PlatformOverview,
    QuerySpec,
    RegionRow,
    ReportScope,
    SentimentRow,
    SentimentSection,
    SentimentSplit,
    SourceEntry,
    TopPostRow,
    TrendPoint,
)


def _payload() -> BrandReportPayload:
    return BrandReportPayload(
        data_status="complete",
        scope=ReportScope(
            brand="肯德基",
            period_start="2026-06-01",
            period_end="2026-06-30",
            platforms=["xiaohongshu", "douyin"],
            comparison_mode="mom_yoy",
            data_as_of="2026-06-29",
        ),
        query_spec=QuerySpec(
            original_term="肯德基",
            matched_tag="肯德基",
            comparison_definition="环比 2026-05-02~2026-05-31；同比 2025-06-01~2025-06-30",
        ),
        data=BrandReportData(
            overview=OverviewSection(
                platforms=[
                    PlatformOverview(
                        platform="xiaohongshu", mentions=1000.0,
                        interactions=8000.0,
                    ),
                    PlatformOverview(
                        platform="douyin", mentions=2000.0,
                        interactions=12000.0,
                    ),
                ],
                total_mentions=MetricComparison(
                    current=3000.0,
                    mom=PeriodValue(value=2500.0),
                    yoy=PeriodValue(value=2000.0),
                    mom_change_pct=20.0,
                    yoy_change_pct=50.0,
                ),
                total_interactions=MetricComparison(
                    current=20000.0, mom_change_pct=-5.5, yoy_change_pct=30.0
                ),
                sentiment_split=SentimentSplit(positive=1800.0, neutral=900.0, negative=300.0),
            ),
            sentiment=SentimentSection(
                rows=[
                    SentimentRow(
                        platform="xiaohongshu", sentiment="正面", mentions=700.0,
                        interactions=3000.0, share_pct=70.0,
                    ),
                    SentimentRow(
                        platform="xiaohongshu", sentiment="负面", mentions=300.0,
                        interactions=1000.0, share_pct=30.0,
                    ),
                    SentimentRow(
                        platform="douyin", sentiment="正面", mentions=1500.0,
                        interactions=9000.0, share_pct=75.0,
                    ),
                ]
            ),
            daily_trend=DailyTrendSection(
                points=[
                    TrendPoint(date="2026-06-01", mentions=100.0, interactions=800.0),
                    TrendPoint(date="2026-06-02", mentions=150.0, interactions=1200.0),
                    TrendPoint(date="2026-06-03", mentions=120.0, interactions=900.0),
                ],
                peak_date="2026-06-02",
                peak_mentions=150.0,
            ),
            content_types=[ContentTypeRow(content_type="探店", mentions=400.0, share_pct=40.0)],
            creator_tiers=[CreatorTierRow(tier="尾部(1w-5w)", mentions=500.0, share_pct=50.0)],
            organic_vs_paid=OrganicVsPaid(
                organic_mentions=2900.0, paid_mentions=100.0,
                organic_share_pct=96.67, paid_share_pct=3.33,
            ),
            regions=[
                RegionRow(region="广东省", mentions=300.0, interactions=2000.0, share_pct=30.0),
                RegionRow(region="江苏省", mentions=200.0, interactions=1500.0, share_pct=20.0),
            ],
            top_posts=[
                TopPostRow(
                    platform="xiaohongshu", post_id="x1", title="肯德基新品测评", author="美食家",
                    interactions=5000, exposure_count=20000, like_count=3000, comment_count=500,
                    collect_count=800, share_count=700, sentiment="正面", creator_tier="尾部",
                    url="https://www.xiaohongshu.com/explore/abc",
                ),
                TopPostRow(platform="xiaohongshu", post_id="x2"),
                TopPostRow(
                    platform="douyin", post_id="d1", title="抖音测评", author="达人",
                    interactions=9000, url="notaurl",
                ),
            ],
        ),
        narrative=BrandReportNarrative(
            praise_points=["新品测评内容互动表现好"],
            complaint_points=["部分门店服务吐槽"],
            impact_level="中",
            expansion_signals=["下沉市场声量上升"],
            noise_notes="存在明星同名噪音。",
            key_findings=["正面声量占主导", "抖音互动高于小红书"],
            conclusion="品牌整体声量环比增长。",
            recommendations=["延续新品测评内容节奏"],
        ),
        availability={chapter: ChapterAvailability(status="complete") for chapter in ALL_CHAPTERS},
        sources=[SourceEntry(tool="datatap.insight.social.statistic.overview.v1", step_id="step_1")],
    )


def _load(data: bytes):
    return load_workbook(BytesIO(data))


def _column_values(sheet, column: int = 1) -> list[str]:
    return [
        str(sheet.cell(row, column).value)
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, column).value is not None
    ]


# ---------- sanitize_report_filename ----------


def test_sanitize_strips_invalid_characters() -> None:
    name = sanitize_report_filename('昊<来>:了/寿司\\|?*"', "2026-06-01", "2026-06-30", 2)
    assert name == "昊来了寿司_品牌社媒分析报告_2026-06-01-2026-06-30_v2.xlsx"


def test_sanitize_collapses_whitespace_and_underscores() -> None:
    name = sanitize_report_filename("肯德基__疯狂  星期四", "2026-06-01", "2026-06-30", 1)
    assert name.startswith("肯德基_疯狂 星期四_品牌社媒分析报告_")


def test_sanitize_strips_trailing_dots_and_spaces() -> None:
    name = sanitize_report_filename("品牌... ", "2026-06-01", "2026-06-30", 1)
    assert name.startswith("品牌_品牌社媒分析报告_")


def test_sanitize_truncates_brand_segment() -> None:
    name = sanitize_report_filename("牌" * 100, "2026-06-01", "2026-06-30", 1)
    assert name.startswith("牌" * 80 + "_品牌社媒分析报告_")


def test_sanitize_empty_brand_falls_back() -> None:
    assert sanitize_report_filename("///", "s", "e", 1).startswith("未命名品牌_")
    assert sanitize_report_filename("   ", "s", "e", 1).startswith("未命名品牌_")
    assert sanitize_report_filename("", "s", "e", 1).startswith("未命名品牌_")


def test_sanitize_strips_control_characters() -> None:
    name = sanitize_report_filename("ab\x00c\x1fd", "s", "e", 1)
    assert name.startswith("abcd_")


# ---------- 完整 payload 渲染 ----------


def test_render_full_payload_sheet_order_and_overview() -> None:
    workbook = _load(render_brand_workbook(_payload()))
    assert workbook.sheetnames == list(SHEET_ORDER)
    overview = workbook["综合概览"]
    assert overview["A1"].value == "肯德基 品牌社交媒体表现分析报告"
    assert "2026-06-01" in overview["B2"].value
    assert "2026-06-30" in overview["B2"].value
    assert "2026-06-29" in overview["B2"].value  # data_as_of 截至日
    assert [overview.cell(6, col).value for col in range(1, 5)] == [
        "指标", "小红书", "抖音", "合计",
    ]
    assert overview["B7"].value == 1000.0
    assert overview["C7"].value == 2000.0
    assert overview["D7"].value == 3000.0
    # 占比直接写 payload/服务端算好的数值，不写跨行公式。
    assert overview["D7"].value != "=B7+C7"
    # 环比/同比：百分比格式。
    assert overview["C19"].value == 20.0
    assert "%" in overview["C19"].number_format
    assert overview["D19"].value == 50.0
    assert "%" in overview["D19"].number_format
    # 声量构成：服务端占比（1800/3000=60）。
    assert overview["B26"].value == 1800.0
    assert overview["C26"].value == 60.0
    assert "%" in overview["C26"].number_format


def test_render_sentiment_sheet_rows_and_findings() -> None:
    workbook = _load(render_brand_workbook(_payload()))
    sheet = workbook["情感分析"]
    assert [sheet.cell(3, col).value for col in range(1, 6)] == [
        "平台", "情感", "声量", "互动数", "占比",
    ]
    assert sheet["A4"].value == "小红书"
    assert sheet["B4"].value == "正面"
    assert sheet["C4"].value == 700.0
    assert "%" in sheet["E4"].number_format
    texts = _column_values(sheet)
    assert "关键发现" in texts
    assert "1. 正面声量占主导" in texts
    assert "2. 抖音互动高于小红书" in texts


def test_render_daily_trend_charts_point_at_written_range() -> None:
    workbook = _load(render_brand_workbook(_payload()))
    sheet = workbook["日趋势"]
    assert sheet["A4"].value == "2026-06-01"
    assert sheet["B5"].value == 150.0
    assert "峰值" in sheet["D5"].value
    # 合计行紧跟数据（3 个点 → 第 7 行）。
    assert sheet["A7"].value == "合计"
    assert sheet["B7"].value == 370.0
    assert sheet["C7"].value == 2900.0
    charts = sheet._charts
    assert len(charts) == 2
    assert all(isinstance(chart, LineChart) for chart in charts)
    references = [chart.ser[0].val.numRef.f for chart in charts]
    assert any("$B$4:$B$6" in ref for ref in references)
    assert any("$C$4:$C$6" in ref for ref in references)


def test_render_overview_five_platforms_dynamic_columns() -> None:
    """5 平台 payload：表头/合计列按实际列数排布，不截断、不落到无表头列。"""
    payload = _payload()
    payload.scope.platforms = ["xiaohongshu", "douyin", "weibo", "bilibili", "wechat"]
    payload.data.overview.platforms = [
        PlatformOverview(platform=platform, mentions=float(index * 100))
        for index, platform in enumerate(
            ["xiaohongshu", "douyin", "weibo", "bilibili", "wechat"], start=1
        )
    ]
    workbook = _load(render_brand_workbook(payload))
    overview = workbook["综合概览"]
    assert [overview.cell(6, col).value for col in range(1, 8)] == [
        "指标", "小红书", "抖音", "微博", "哔哩哔哩", "微信", "合计",
    ]
    # 合计落在第 7 列（有表头），数值为五平台之和。
    assert overview.cell(7, 7).value == 1500.0
    assert overview.cell(6, 8).value is None
    # 互动数各平台缺失 → 合计「未提供」。
    assert overview.cell(8, 7).value == "未提供"


def test_render_regions_bar_chart() -> None:
    workbook = _load(render_brand_workbook(_payload()))
    sheet = workbook["地域分布"]
    assert sheet["A4"].value == "广东省"
    assert sheet["B4"].value == 300.0
    assert "%" in sheet["D4"].number_format
    assert len(sheet._charts) == 1
    chart = sheet._charts[0]
    assert isinstance(chart, BarChart)
    assert "$B$4:$B$5" in chart.ser[0].val.numRef.f


def test_render_content_creators_dynamic_rows_and_reason_placement() -> None:
    """内容类型与达人 Sheet：三段动态行、占比数值格式、reason 落位不覆盖 A1 标题。"""
    payload = _payload()
    payload.availability["content_creators"] = ChapterAvailability(
        status="partial", reason="达人层级数据部分缺失"
    )
    workbook = _load(render_brand_workbook(payload))
    sheet = workbook["内容类型与达人"]
    # A1 模板标题不被覆盖；reason 写在第 2 行（模板该行无内容）。
    assert sheet["A1"].value == "内容类型分布"
    assert "达人层级数据部分缺失" in sheet["A2"].value
    # 内容类型段：表头第 3 行，数据第 4 行。
    assert [sheet.cell(3, col).value for col in (1, 2, 3)] == ["内容类型", "声量", "占比"]
    assert sheet["A4"].value == "探店"
    assert sheet["B4"].value == 400.0
    assert sheet["C4"].value == 40.0
    assert "%" in sheet["C4"].number_format
    # 达人层级段（动态：内容类型 1 行数据 → 标题 6、表头 7、数据 8）。
    assert sheet["A6"].value == "达人层级分布 (品牌相关内容)"
    assert sheet["A8"].value == "尾部(1w-5w)"
    assert sheet["C8"].value == 50.0
    # 商单 vs 自然段（标题 10、表头 11、数据 12-13）。
    assert sheet["A10"].value == "商单 vs 自然内容"
    assert sheet["A12"].value == "自然内容(非商单)"
    assert sheet["C12"].value == 96.67
    assert "%" in sheet["C12"].number_format
    assert sheet["A13"].value == "商单内容"
    assert sheet["C13"].value == 3.33


def test_render_top_posts_dual_platform_sections() -> None:
    workbook = _load(render_brand_workbook(_payload()))
    sheet = workbook["热门帖子TOP"]
    # 小红书段在前：标题、13 列表头（含链接列）、数据行。
    assert "小红书" in sheet["A1"].value
    assert sheet.cell(2, 13).value == "链接"
    assert sheet["A3"].value == 1
    assert sheet["C3"].value == "肯德基新品测评"
    # 抖音段排在小红书段之后。
    assert "抖音" in sheet["A6"].value
    assert sheet["C8"].value == "抖音测评"
    # 合法 URL：单元格带超链接。
    url_cell = sheet.cell(3, 13)
    assert url_cell.value == "https://www.xiaohongshu.com/explore/abc"
    assert url_cell.hyperlink is not None


def test_render_methodology_and_insights() -> None:
    workbook = _load(render_brand_workbook(_payload()))
    methodology = workbook["方法论"]
    pairs = {
        methodology.cell(row, 1).value: methodology.cell(row, 2).value
        for row in range(1, methodology.max_row + 1)
        if methodology.cell(row, 1).value and methodology.cell(row, 2).value
    }
    assert "环比 2026-05-02~2026-05-31" in pairs["对比口径"]
    assert "datatap.insight.social.statistic.overview.v1" in pairs["数据工具"]
    assert "brand_report_v2" in pairs["报告生成"]
    insights = _column_values(workbook["舆情洞察"])
    assert "1. 新品测评内容互动表现好" in insights
    assert any("负面影响程度" in text and "中" in text for text in insights)
    assert "存在明星同名噪音。" in insights


# ---------- 空章节降级 ----------


def test_render_empty_trend_and_regions_no_charts_with_reason() -> None:
    payload = _payload()
    payload.data.daily_trend = DailyTrendSection()
    payload.data.regions = []
    payload.availability["daily_trend"] = ChapterAvailability(
        status="unavailable", reason="趋势工具调用失败，未采集到数据"
    )
    payload.availability["regions"] = ChapterAvailability(
        status="unavailable", reason="地域分布数据缺失"
    )
    workbook = _load(render_brand_workbook(payload))
    trend = workbook["日趋势"]
    assert trend._charts == []
    assert any("趋势工具调用失败" in text for text in _column_values(trend))
    regions = workbook["地域分布"]
    assert regions._charts == []
    assert any("地域分布数据缺失" in text for text in _column_values(regions))
    # 空章节保留 Sheet 与列头，不隐藏。
    assert [trend.cell(3, col).value for col in range(1, 5)] == ["日期", "声量", "互动数", "备注"]
    assert [regions.cell(3, col).value for col in range(1, 5)] == [
        "省份", "声量", "互动数", "声量占比",
    ]


# ---------- 热帖平台表头 / partial 受限说明 ----------


def test_render_top_posts_platform_specific_headers() -> None:
    """热帖表头按平台段切换：小红书段「阅读数/转发」，抖音段「播放数/分享」。"""
    workbook = _load(render_brand_workbook(_payload()))
    sheet = workbook["热门帖子TOP"]
    # 小红书段表头（第 2 行）。
    assert sheet.cell(2, 6).value == "阅读数"
    assert sheet.cell(2, 10).value == "转发"
    # 抖音段表头（标题第 6 行 → 表头第 7 行）。
    assert sheet.cell(7, 6).value == "播放数"
    assert sheet.cell(7, 10).value == "分享"


def test_render_partial_chapter_reason_includes_missing_fields() -> None:
    """partial 章节只填 missing_fields（无 reason）时，受限说明也要落到 Sheet。"""
    payload = _payload()
    payload.data.regions = []
    payload.availability["regions"] = ChapterAvailability(
        status="partial", missing_fields=["interactions"]
    )
    workbook = _load(render_brand_workbook(payload))
    texts = _column_values(workbook["地域分布"])
    assert any(
        "数据受限" in text and "缺失字段：interactions" in text for text in texts
    )


# ---------- 热帖缺失字段 ----------


def test_render_top_post_missing_fields_and_no_fake_link() -> None:
    workbook = _load(render_brand_workbook(_payload()))
    sheet = workbook["热门帖子TOP"]
    # 第二条小红书帖子全部业务字段为 null。
    assert sheet["A4"].value == 2
    assert sheet["C4"].value == "未提供"
    assert sheet["D4"].value == "未提供"
    assert sheet["E4"].value == "未提供"
    assert sheet.cell(4, 13).value == "未提供"
    assert sheet.cell(4, 13).hyperlink is None
    # 非法 URL（非 http/https）不出现链接。
    assert sheet.cell(8, 13).value == "未提供"
    assert sheet.cell(8, 13).hyperlink is None


def test_render_top_post_javascript_url_no_hyperlink() -> None:
    """带 scheme 的协议注入（javascript:）同样不生成 hyperlink，显示「未提供」。"""
    payload = _payload()
    payload.data.top_posts[2].url = "javascript:alert(1)"
    workbook = _load(render_brand_workbook(payload))
    sheet = workbook["热门帖子TOP"]
    assert sheet["C8"].value == "抖音测评"
    assert sheet.cell(8, 13).value == "未提供"
    assert sheet.cell(8, 13).hyperlink is None


def test_render_empty_top_posts_keeps_headers_and_reason() -> None:
    """热帖空章节：保留 Sheet + 13 列通用列头 + availability.reason 受限说明。"""
    payload = _payload()
    payload.data.top_posts = []
    payload.availability["top_posts"] = ChapterAvailability(
        status="unavailable", reason="原帖查询工具调用失败"
    )
    workbook = _load(render_brand_workbook(payload))
    sheet = workbook["热门帖子TOP"]
    assert [sheet.cell(3, col).value for col in range(1, 14)] == [
        "排名", "平台", "标题", "用户昵称", "互动数", "阅读数",
        "点赞", "评论", "收藏", "转发", "情感", "达人层级", "链接",
    ]
    assert any("原帖查询工具调用失败" in text for text in _column_values(sheet))


# ---------- 往返 ----------


def test_render_roundtrip_loadable() -> None:
    payload = _payload()
    first = render_brand_workbook(payload)
    again = render_brand_workbook(
        BrandReportPayload.model_validate(payload.model_dump(mode="json"))
    )
    for data in (first, again):
        workbook = _load(data)
        assert workbook.sheetnames == list(SHEET_ORDER)
        assert workbook["综合概览"]["B7"].value == 1000.0
