from io import BytesIO

from openpyxl import load_workbook

from app.reporting.exporter import ExportCandidate, render_workbook


def _candidate(index: int) -> ExportCandidate:
    return ExportCandidate(
        rank=index,
        platform="xiaohongshu" if index % 2 else "douyin",
        nickname=f"达人{index}",
        followers=20_000 + index,
        city="湖州",
        total_score=80 - index,
        rating="推荐",
        stars="★★★★",
        dimension_scores={
            "industry_interest": 18,
            "target_region": 12,
            "target_age": 14,
            "engagement": 10,
            "active_follower": 8,
            "content": 9,
            "followers": 8,
            "engagement_follower_ratio": 4,
        },
        values={
            "industry_interest_rate": "18.00%",
            "target_region_rate": "12.00%",
            "target_age_rate": "70.00%",
            "engagement_rate": "3.20%",
            "active_follower_rate": "40.00%",
            "content_tags": "美妆,护肤",
        },
        score_reason="字段完整，匹配度较高",
        source_names=("聆媒洞察",),
        collected_at="2026-07-16T05:00:00",
    )


def test_render_workbook_contains_all_candidates_and_template_sheets() -> None:
    content = render_workbook(
        metadata={
            "brand": "测试品牌",
            "category": "美妆",
            "target_audience": "20-30女性",
            "locations": ["浙江", "湖州"],
            "generated_at": "2026-07-16T05:00:00",
        },
        candidates=[_candidate(index) for index in range(1, 13)],
    )

    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)

    assert workbook.sheetnames == [
        "KOL匹配度筛选",
        "达人详细画像",
        "粉丝画像详情",
        "评分方法论与数据来源",
    ]
    summary = workbook["KOL匹配度筛选"]
    headers = [cell.value for cell in summary[4]]
    assert "平台" in headers
    assert summary.max_row >= 4 + 12
    assert summary[5][1].value in {"小红书", "抖音"}
    assert "平台: 小红书、抖音" in (summary[2][0].value or "")
    assert workbook["粉丝画像详情"].max_row >= 1 + 12
    detail = workbook["达人详细画像"]
    methodology = workbook["评分方法论与数据来源"]
    assert "B3:F3" in {str(item) for item in detail.merged_cells.ranges}
    assert "A13:F13" in {str(item) for item in detail.merged_cells.ranges}
    assert "A15:D15" in {str(item) for item in methodology.merged_cells.ranges}
    assert "A24:D24" in {str(item) for item in methodology.merged_cells.ranges}
    assert "A31:D31" in {str(item) for item in methodology.merged_cells.ranges}


def test_render_workbook_uses_chinese_placeholder_for_missing_candidate_fields() -> None:
    candidate = ExportCandidate(
        rank=1,
        platform="xiaohongshu",
        nickname="缺失达人",
        followers=None,
        city=None,
        total_score=None,
        rating="数据缺失",
        stars="数据缺失",
        dimension_scores={key: None for key in (
            "industry_interest", "target_region", "target_age", "engagement",
            "active_follower", "content", "followers", "engagement_follower_ratio",
        )},
    )
    content = render_workbook(
        metadata={"brand": "缺失字段", "category": "美妆", "generated_at": "2026-07-16"},
        candidates=[candidate],
    )
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    summary = workbook["KOL匹配度筛选"]
    assert summary[5][4].value == "数据缺失"  # 粉丝数
    assert summary[5][5].value == "数据缺失"  # 城市
    assert summary[5][6].value == "数据缺失"  # 首个维度分
    assert summary[5][14].value == "数据缺失"  # 综合评分
    fan_profile = workbook["粉丝画像详情"]
    assert fan_profile[2][3].value == "数据缺失"
    detail = workbook["达人详细画像"]
    assert detail["B5"].value == "数据缺失"


def test_render_workbook_writes_only_merged_range_anchors_when_rows_expand() -> None:
    """Rows beyond the template reservation must not assign to MergedCell objects."""
    content = render_workbook(
        metadata={"brand": "合并单元格测试", "category": "餐饮", "generated_at": "2026-07-16"},
        candidates=[_candidate(index) for index in range(1, 81)],
    )

    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    summary = workbook["KOL匹配度筛选"]
    assert summary[84][0].value == 80
    assert any(str(merged) == "A1:Q1" for merged in summary.merged_cells.ranges)
    assert any(str(merged) == "A2:Q2" for merged in summary.merged_cells.ranges)
    # The dynamically moved rating table keeps the template's row styles.
    assert summary.cell(88, 1)._style == summary.cell(22, 1)._style


def test_detail_blocks_rebuild_standard_merges_at_late_and_dynamic_rows() -> None:
    content = render_workbook(
        metadata={"brand": "详情块测试", "category": "餐饮", "generated_at": "2026-07-16"},
        candidates=[_candidate(index) for index in range(1, 81)],
    )
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    detail = workbook["达人详细画像"]
    merged = {str(item) for item in detail.merged_cells.ranges}
    for start in (218, 249, 2450):
        assert f"A{start}:F{start}" in merged
        assert f"A{start + 1}:F{start + 1}" in merged
        assert f"B{start + 2}:F{start + 2}" in merged
        assert f"A{start + 29}:F{start + 29}" in merged
        assert str(detail.cell(start + 29, 1).value).startswith("报告摘要：")
        assert detail.cell(start + 29, 2).value is None


def test_detail_blocks_copy_complete_standard_row_height_matrix() -> None:
    content = render_workbook(
        metadata={"brand": "详情行高测试", "category": "餐饮", "generated_at": "2026-07-16"},
        candidates=[_candidate(index) for index in range(1, 81)],
    )
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    detail = workbook["达人详细画像"]
    base_heights = [detail.row_dimensions[1 + offset].height for offset in range(31)]
    for start in (1, 218, 249, 2450):
        actual = [detail.row_dimensions[start + offset].height for offset in range(31)]
        assert actual == base_heights


def test_export_candidate_keeps_public_profile_url_and_platform() -> None:
    candidate = _candidate(1)
    candidate = ExportCandidate(**{**candidate.__dict__, "profile_url": "https://example.com/达人1"})
    content = render_workbook(
        metadata={"brand": "链接测试", "category": "美妆", "generated_at": "2026-07-16"},
        candidates=[candidate],
    )
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=False)
    detail = workbook["达人详细画像"]
    values = [cell.value for row in detail.iter_rows() for cell in row if cell.value]
    assert any("小红书" in str(value) for value in values)
    assert any("https://example.com/达人1" in str(value) for value in values)
