"""品牌报告全链路集成：plan_json 证据 → run_brand_analysis 落库 → export_brand_report 渲染。

覆盖集成盲区：assemble_brand_report 真实产物 → 叙事（FakeModel）→ model_dump 落库 →
export_brand_report 读回渲染 → openpyxl 读回 workbook 断言 8 Sheet 与关键数值；
payload 形态含 platform="all" 合计行（防双计）与 partial + data_as_of 组合。
全程不调真实 MCP/模型。
"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
import json
from types import SimpleNamespace
from uuid import uuid4

import pytest
from openpyxl import load_workbook

from app.model.contracts import StructuredResult
from app.reporting.brand_exporter import SHEET_ORDER, export_brand_report
from app.reporting.brand_payload import BrandReportNarrative, BrandReportPayload
from app.reporting.builders import run_brand_analysis
from app.workspace.models import WorkspaceSession

TOOL_OVERVIEW = "datatap.insight.social.statistic.overview.v1"
TOOL_TREND = "datatap.insight.social.statistic.trend.v1"
TOOL_RAW_POSTS = "datatap.insight.query.raw.posts.v1"

_NARRATIVE_OUTPUT = {
    "praise_points": ["新品测评内容互动表现好"],
    "complaint_points": [],
    "impact_level": "低",
    "expansion_signals": [],
    "noise_notes": None,
    "key_findings": ["正面声量占主导"],
    "conclusion": "品牌声量稳步上升。",
    "recommendations": ["延续新品测评内容节奏"],
}


class FakeModel:
    def __init__(self) -> None:
        self.requests: list = []

    async def complete_json(self, request):
        self.requests.append(request)
        assert request.output_model is BrandReportNarrative
        return StructuredResult(
            value=BrandReportNarrative.model_validate(_NARRATIVE_OUTPUT),
            usage=None,
            request_id="req-integration",
            regeneration_count=0,
        )


def _summary(rows: object) -> dict[str, str]:
    return {"result": json.dumps(rows, ensure_ascii=False)}


def _step(step_id: str, tool: str, goal: str) -> dict[str, object]:
    return {
        "id": step_id,
        "internal_tool_name": tool,
        "arguments": {"start_time": "2026-06-01", "end_time": "2026-06-30"},
        "evidence_goal": goal,
    }


def _note(step_id: str, tool: str, summary: object) -> dict[str, object]:
    return {"step_id": step_id, "tool": tool, "status": "settled", "summary": summary}


async def _create_session(db_session, user_factory) -> tuple[str, str]:
    user = await user_factory()
    now = datetime.now(UTC).replace(tzinfo=None)
    session = WorkspaceSession(
        id=str(uuid4()),
        user_id=user.id,
        title="品牌报告集成会话",
        brand="海底捞",
        campaign_name=None,
        status="active",
        platforms=["xiaohongshu", "douyin"],
        category="美食",
        target_audience="",
        budget_min=None,
        budget_max=None,
        filters_snapshot={},
        is_starred=False,
        last_accessed_at=now,
        created_at=now,
        updated_at=now,
    )
    db_session.add(session)
    await db_session.flush()
    return user.id, session.id


def _plan_json() -> dict[str, object]:
    """v1 轨迹：overview（含「合计」行防双计）+ 趋势（尾部缺 2 天）+ 热帖。"""
    trend_rows = [
        {"日期": f"2026-06-{day:02d}", "平台": "小红书", "声量": 30, "互动数": 100}
        for day in range(1, 29)  # 6-01~6-28：周期末 6-29/6-30 无数据
    ]
    steps = [
        _step("step_1", TOOL_OVERVIEW, "current: 当期概览"),
        _step("step_2", TOOL_TREND, "current: 当期日趋势"),
        _step("step_3", TOOL_RAW_POSTS, "current: 品牌热门原帖"),
    ]
    notes = [
        _note(
            "step_1",
            TOOL_OVERVIEW,
            _summary(
                [
                    {"平台": "小红书", "声量": 1000, "曝光量": 50000, "互动数": 8000},
                    {"平台": "抖音", "声量": 2000, "曝光量": 90000, "互动数": 12000},
                    # DataTap 常见合计行：与平台明细并存，必须跳过防双计。
                    {"平台": "合计", "声量": 99999, "曝光量": 99999, "互动数": 99999},
                ]
            ),
        ),
        _note("step_2", TOOL_TREND, _summary(trend_rows)),
        _note(
            "step_3",
            TOOL_RAW_POSTS,
            _summary(
                [
                    {
                        "平台": "小红书",
                        "帖子ID": "xhs-1",
                        "标题": "海底捞新品测评",
                        "昵称": "美食家",
                        "互动数": 5000,
                        "阅读数": 20000,
                        "点赞数": 3000,
                        "评论数": 500,
                        "收藏数": 800,
                        "转发数": 700,
                        "发布时间": "2026-06-15",
                        "帖子链接": "https://www.xiaohongshu.com/explore/abc",
                    }
                ]
            ),
        ),
    ]
    return {"schema": "agent_trajectory_v1", "steps": steps, "results": notes}


@pytest.mark.asyncio
async def test_brand_report_end_to_end_xlsx(db_session, user_factory) -> None:
    user_id, session_id = await _create_session(db_session, user_factory)
    task = SimpleNamespace(id=str(uuid4()), plan_json=_plan_json())
    goal = SimpleNamespace(
        id=str(uuid4()),
        params_json={
            "brand": "海底捞",
            "period": {"start": "2026-06-01", "end": "2026-06-30"},
            "platforms": ["xiaohongshu", "douyin"],
        },
    )

    report = await run_brand_analysis(
        db_session,
        FakeModel(),
        user_id=user_id,
        session_id=session_id,
        task=task,
        goal=goal,
    )

    # 落库形态：v2 快照 + partial + data_as_of（趋势尾部缺 6-29/6-30）。
    assert report.template_version == "brand_report_v2"
    payload = BrandReportPayload.model_validate(report.payload_json)
    assert payload.data_status == "partial"
    assert payload.scope.data_as_of == "2026-06-28"
    # 「合计」行被跳过：总声量 1000+2000，无双计。
    assert payload.data.overview.total_mentions.current == 3000.0

    exported = await export_brand_report(db_session, user_id, session_id, report.id)
    assert exported.filename.startswith("海底捞_品牌社媒分析报告_2026-06-01-2026-06-30_v1")

    workbook = load_workbook(BytesIO(exported.content))
    assert workbook.sheetnames == list(SHEET_ORDER)

    overview = workbook["综合概览"]
    assert overview["A1"].value == "海底捞 品牌社交媒体表现分析报告"
    assert "数据截至 2026-06-28" in overview["B2"].value
    assert [overview.cell(6, col).value for col in range(1, 5)] == [
        "指标", "小红书", "抖音", "合计",
    ]
    assert overview["B7"].value == 1000.0
    assert overview["C7"].value == 2000.0
    assert overview["D7"].value == 3000.0  # 合计行未被「合计」证据行污染

    trend = workbook["日趋势"]
    assert trend["A4"].value == "2026-06-01"
    assert trend["A31"].value == "2026-06-28"
    assert trend["A32"].value == "合计"  # 28 个点：数据行 4-31，合计行 32
    assert trend["B32"].value == 30.0 * 28
    assert len(trend._charts) == 2

    posts = workbook["热门帖子TOP"]
    assert "小红书" in posts["A1"].value
    assert posts.cell(2, 6).value == "阅读数"  # 小红书段表头
    assert posts.cell(2, 10).value == "转发"
    assert posts["C3"].value == "海底捞新品测评"
    assert posts.cell(3, 13).hyperlink is not None

    methodology = workbook["方法论"]
    pairs = {
        methodology.cell(row, 1).value: methodology.cell(row, 2).value
        for row in range(1, methodology.max_row + 1)
        if methodology.cell(row, 1).value and methodology.cell(row, 2).value
    }
    assert "数据截至 2026-06-28" in pairs["时间范围"]
    assert "日趋势" in pairs["章节可用性"]  # partial 章节进入可用性汇总
