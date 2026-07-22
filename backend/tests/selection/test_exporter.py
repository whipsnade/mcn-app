from __future__ import annotations

import re
from collections.abc import AsyncIterator
from datetime import UTC, datetime, timedelta
from io import BytesIO
from typing import Any
from uuid import uuid4

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.security import create_access_token
from app.db.session import get_db
from app.identity.models import LoginSession, User
from app.main import create_app
from app.selection.exporter import (
    CONTENT_TYPE,
    ExportCandidate,
    export_session_selection,
    render_workbook,
)
from app.selection.models import SessionKolSelection
from app.workspace.models import WorkspaceSession


def _candidate(index: int) -> ExportCandidate:
    return ExportCandidate(
        rank=index,
        platform="xiaohongshu" if index % 2 else "douyin",
        nickname=f"达人{index}",
        followers=20_000 + index,
        city="湖州",
        total_score=80 - index,
        rating="推荐",
        stars="★★★★",
        dimension_scores={
            "industry_interest": 18,
            "target_region": 12,
            "target_age": 14,
            "engagement": 10,
            "active_follower": 8,
            "content": 9,
            "followers": 8,
            "engagement_follower_ratio": 4,
        },
        values={
            "industry_interest_rate": "18.00%",
            "target_region_rate": "12.00%",
            "target_age_rate": "70.00%",
            "engagement_rate": "3.20%",
            "active_follower_rate": "40.00%",
            "content_tags": "美妆,护肤",
        },
        score_reason="字段完整，匹配度较高",
        source_names=("已授权数据服务",),
        collected_at="2026-07-16T05:00:00",
    )


def _tier_candidate(rank: int, total: float, rating: str, stars: str) -> ExportCandidate:
    return ExportCandidate(
        rank=rank,
        platform="xiaohongshu",
        nickname=f"{rating}达人",
        followers=10_000,
        city="杭州市",
        total_score=total,
        rating=rating,
        stars=stars,
    )


def _four_tier_candidates() -> list[ExportCandidate]:
    return [
        _tier_candidate(1, 90.0, "重点推荐", "★★★★★"),
        _tier_candidate(2, 70.0, "推荐", "★★★★"),
        _tier_candidate(3, 55.0, "可考虑", "★★★"),
        _tier_candidate(4, 30.0, "观察", "★★"),
    ]


def test_rating_summary_counts_all_candidates_across_four_tiers() -> None:
    """汇总表与图表辅助列必须覆盖新四档评级的全部候选人。"""
    candidates = _four_tier_candidates()
    content = render_workbook(
        metadata={"brand": "评级测试", "category": "美食", "generated_at": "2026-07-22"},
        candidates=candidates,
    )

    summary = load_workbook(BytesIO(content))["KOL匹配度筛选"]
    # _render_summary: rating_title_row = max(20, 5 + n + 2)，评级表头在其下一行。
    start_row = max(20, 5 + len(candidates) + 2) + 1

    labels = [summary.cell(start_row + 1 + index, 1).value for index in range(4)]
    assert labels == ["重点推荐", "推荐", "可考虑", "观察"]
    counts = [summary.cell(start_row + 1 + index, 4).value for index in range(4)]
    assert counts == [1, 1, 1, 1]
    assert summary.cell(start_row + 5, 1).value is None  # 旧第五档不再写入

    # 图表辅助列（R=评级, S=数量）：按 rating 文本匹配计数，四档都要统计到。
    aux_labels = [summary.cell(start_row + index, 18).value for index in range(4)]
    assert aux_labels == ["重点推荐", "推荐", "可考虑", "观察"]
    aux_counts = [summary.cell(start_row + index, 19).value for index in range(4)]
    assert sum(aux_counts) == len(candidates)
    assert summary.cell(start_row + 4, 18).value is None


def test_rating_fill_colors_cover_four_tiers() -> None:
    candidates = _four_tier_candidates()
    content = render_workbook(
        metadata={"brand": "填充测试", "category": "美食", "generated_at": "2026-07-22"},
        candidates=candidates,
    )

    summary = load_workbook(BytesIO(content))["KOL匹配度筛选"]
    fills = [str(summary.cell(5 + index, 16).fill.fgColor.rgb) for index in range(4)]

    assert fills[0].endswith("318B32")  # 重点推荐：最深强调色
    assert fills[1].endswith("63BE7B")  # 推荐
    assert fills[2].endswith("FFDD35")  # 可考虑
    assert fills[3].endswith("FFEB84")  # 观察：最浅


def test_methodology_rating_table_uses_four_tiers() -> None:
    content = render_workbook(
        metadata={"brand": "方法论测试", "category": "美食", "generated_at": "2026-07-22"},
        candidates=[_candidate(1)],
    )

    sheet = load_workbook(BytesIO(content))["评分方法论与数据来源"]
    rows = [(sheet.cell(17 + index, 1).value, sheet.cell(17 + index, 2).value,
             sheet.cell(17 + index, 3).value) for index in range(4)]
    assert rows == [
        ("重点推荐", "★★★★★", "≥78"),
        ("推荐", "★★★★", "62-77"),
        ("可考虑", "★★★", "48-61"),
        ("观察", "★★", "<48"),
    ]
    assert sheet.cell(21, 1).value is None  # 旧「不推荐」行不再写入


def test_render_workbook_escapes_formula_injection_in_third_party_strings() -> None:
    """KOL 昵称/标签等第三方可控文本以 =、+、-、@ 开头时必须转义，不得写成公式。"""
    candidate = ExportCandidate(
        rank=1,
        platform="xiaohongshu",
        nickname='=HYPERLINK("http://evil.example","click")',
        followers=10_000,
        city="杭州市",
        total_score=80.0,
        rating="重点推荐",
        stars="★★★★★",
        values={"content_tags": "+cmd", "engagement_rate": "@evil", "summary": "-1+1"},
    )
    content = render_workbook(
        metadata={"brand": "=cmd", "category": "美食", "generated_at": "2026-07-22"},
        candidates=[candidate],
    )

    workbook = load_workbook(BytesIO(content))
    summary = workbook["KOL匹配度筛选"]
    title_cell = summary.cell(1, 1)
    assert title_cell.data_type != "f"
    assert str(title_cell.value).startswith("'=cmd")
    nickname_cell = summary.cell(5, 3)
    assert nickname_cell.data_type != "f"
    assert nickname_cell.value == "'=HYPERLINK(\"http://evil.example\",\"click\")"

    detail = workbook["达人详细画像"]
    detail_values = [cell.value for row in detail.iter_rows() for cell in row if cell.value]
    assert any(str(value).startswith("'+cmd") for value in detail_values)
    assert any(str(value).startswith("'@evil") for value in detail_values)
    assert all(cell.data_type != "f" for row in detail.iter_rows() for cell in row)
    fan_profile = workbook["粉丝画像详情"]
    assert all(cell.data_type != "f" for row in fan_profile.iter_rows() for cell in row)


def test_rating_block_and_chart_move_below_large_candidate_pool() -> None:
    """超过模板预留行数时评级块下移，图表锚点与引用区间同步更新。"""
    candidates = [_candidate(index) for index in range(1, 17)]
    content = render_workbook(
        metadata={"brand": "大名单", "category": "美食", "generated_at": "2026-07-22"},
        candidates=candidates,
    )

    workbook = load_workbook(BytesIO(content))
    summary = workbook["KOL匹配度筛选"]
    rating_title_row = 5 + len(candidates) + 2  # 23，超过模板默认的 20
    start_row = rating_title_row + 1
    assert summary.cell(rating_title_row, 1).value == "评级分布汇总"
    assert summary.cell(start_row, 1).value == "评级"
    assert [summary.cell(start_row + 1 + i, 1).value for i in range(4)] == [
        "重点推荐", "推荐", "可考虑", "观察",
    ]
    assert [summary.cell(start_row + i, 18).value for i in range(4)] == [
        "重点推荐", "推荐", "可考虑", "观察",
    ]
    # 候选人评级均为「推荐」，辅助列计数总和覆盖全部候选人。
    assert sum(summary.cell(start_row + i, 19).value for i in range(4)) == len(candidates)

    charts = summary._charts
    assert len(charts) == 1
    assert charts[0].anchor._from.row == rating_title_row + 7 - 1
    refs = [
        (series.cat.numRef.f, series.val.numRef.f)
        for series in charts[0].ser
        if getattr(getattr(series, "cat", None), "numRef", None) is not None
    ]
    assert refs == [
        (
            f"'KOL匹配度筛选'!$R${start_row}:$R${start_row + 3}",
            f"'KOL匹配度筛选'!$S${start_row}:$S${start_row + 3}",
        )
    ]


def test_render_workbook_contains_all_candidates_and_template_sheets() -> None:
    content = render_workbook(
        metadata={
            "brand": "测试品牌",
            "category": "美妆",
            "target_audience": "20-30女性",
            "locations": ["浙江", "湖州"],
            "generated_at": "2026-07-16T05:00:00",
        },
        candidates=[_candidate(index) for index in range(1, 13)],
    )

    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)

    assert workbook.sheetnames == [
        "KOL匹配度筛选",
        "达人详细画像",
        "粉丝画像详情",
        "评分方法论与数据来源",
    ]
    summary = workbook["KOL匹配度筛选"]
    headers = [cell.value for cell in summary[4]]
    assert "平台" in headers
    assert summary.max_row >= 4 + 12
    assert summary[5][1].value in {"小红书", "抖音"}
    assert "平台: 小红书、抖音" in (summary[2][0].value or "")
    assert workbook["粉丝画像详情"].max_row >= 1 + 12
    detail = workbook["达人详细画像"]
    methodology = workbook["评分方法论与数据来源"]
    assert "B3:F3" in {str(item) for item in detail.merged_cells.ranges}
    assert "A13:F13" in {str(item) for item in detail.merged_cells.ranges}
    assert "A15:D15" in {str(item) for item in methodology.merged_cells.ranges}
    assert "A24:D24" in {str(item) for item in methodology.merged_cells.ranges}
    assert "A31:D31" in {str(item) for item in methodology.merged_cells.ranges}


def test_render_workbook_uses_chinese_placeholder_for_missing_candidate_fields() -> None:
    candidate = ExportCandidate(
        rank=1,
        platform="xiaohongshu",
        nickname="缺失达人",
        followers=None,
        city=None,
        total_score=None,
        rating="数据缺失",
        stars="数据缺失",
        dimension_scores={key: None for key in (
            "industry_interest", "target_region", "target_age", "engagement",
            "active_follower", "content", "followers", "engagement_follower_ratio",
        )},
    )
    content = render_workbook(
        metadata={"brand": "缺失字段", "category": "美妆", "generated_at": "2026-07-16"},
        candidates=[candidate],
    )
    workbook = load_workbook(BytesIO(content), read_only=False, data_only=False)
    summary = workbook["KOL匹配度筛选"]
    assert summary[5][4].value == "数据缺失"  # 粉丝数
    assert summary[5][5].value == "数据缺失"  # 城市
    assert summary[5][6].value == "数据缺失"  # 首个维度分
    assert summary[5][14].value == "数据缺失"  # 综合评分
    fan_profile = workbook["粉丝画像详情"]
    assert fan_profile[2][3].value == "数据缺失"
    detail = workbook["达人详细画像"]
    assert detail["B5"].value == "数据缺失"


def _selection_row(
    user_id: str,
    session_id: str,
    uid: str,
    *,
    platform: str = "xiaohongshu",
    nickname: str = "圈选达人",
    followers: int | None = 125000,
    city: str | None = "杭州市",
    total: float = 80.0,
    rating: str = "重点推荐",
    stars: str = "★★★★★",
    dimensions: dict[str, Any] | None = None,
    export_fields: dict[str, Any] | None = None,
) -> SessionKolSelection:
    now = datetime.now(UTC).replace(tzinfo=None)
    return SessionKolSelection(
        id=str(uuid4()),
        user_id=user_id,
        session_id=session_id,
        platform=platform,
        kol_uid=uid,
        nickname=nickname,
        followers=followers,
        city=city,
        profile_url=f"https://example.com/{uid}",
        fields_json={"export_fields": dict(export_fields or {}), "missing_fields": []},
        score_json={
            "total": total,
            "rating": rating,
            "stars": stars,
            "dimensions": dimensions
            or {"content": {"raw_score": 88.0, "weight": 20, "weighted_score": 17.6}},
        },
        source_tool="tool",
        first_task_id="t1",
        last_task_id="t1",
        created_at=now,
        updated_at=now,
    )


async def _create_session(
    db_session: AsyncSession,
    user: User,
    *,
    brand: str = "海底捞",
    category: str | None = "美食",
    filters_snapshot: dict[str, Any] | None = None,
) -> WorkspaceSession:
    now = datetime.now(UTC).replace(tzinfo=None)
    session = WorkspaceSession(
        id=str(uuid4()),
        user_id=user.id,
        title="圈选会话",
        brand=brand,
        campaign_name=None,
        status="active",
        platforms=["xiaohongshu"],
        category=category,
        target_audience="25-35岁",
        budget_min=None,
        budget_max=None,
        filters_snapshot=filters_snapshot or {},
        is_starred=False,
        last_accessed_at=now,
        created_at=now,
        updated_at=now,
    )
    db_session.add(session)
    await db_session.flush()
    return session


@pytest.mark.asyncio
async def test_export_session_selection_renders_rows_sorted_by_score(
    db_session: AsyncSession, user_factory
) -> None:
    user = await user_factory()
    session = await _create_session(db_session, user)
    db_session.add(
        _selection_row(user.id, session.id, "low", nickname="低分达人", total=50.0,
                       rating="可考虑", stars="★★★")
    )
    db_session.add(
        _selection_row(user.id, session.id, "high", nickname="高分达人", total=90.0)
    )
    await db_session.flush()

    workbook_out = await export_session_selection(db_session, user.id, session.id)

    assert workbook_out.content_type == CONTENT_TYPE
    assert re.fullmatch(
        r"海底捞_美食_KOL匹配度分析_\d{8}_\d{4}\.xlsx", workbook_out.filename
    )
    workbook = load_workbook(BytesIO(workbook_out.content), read_only=False)
    assert workbook.sheetnames == [
        "KOL匹配度筛选",
        "达人详细画像",
        "粉丝画像详情",
        "评分方法论与数据来源",
    ]
    summary = workbook["KOL匹配度筛选"]
    # 表头第 4 行 + 2 行候选人，按综合评分倒序。
    assert summary.cell(5, 1).value == 1
    assert summary.cell(5, 2).value == "小红书"
    assert summary.cell(5, 3).value == "高分达人"
    assert summary.cell(5, 15).value == 90.0
    assert summary.cell(5, 16).value == "重点推荐"
    assert summary.cell(6, 1).value == 2
    assert summary.cell(6, 3).value == "低分达人"
    assert summary.cell(6, 15).value == 50.0
    assert summary.cell(6, 16).value == "可考虑"
    assert workbook["达人详细画像"].max_row >= 31 * 2
    assert workbook["粉丝画像详情"].max_row >= 1 + 2


@pytest.mark.asyncio
async def test_export_session_selection_empty_raises_no_kol_selection(
    db_session: AsyncSession, user_factory
) -> None:
    user = await user_factory()
    session = await _create_session(db_session, user)

    with pytest.raises(LookupError, match="no_kol_selection"):
        await export_session_selection(db_session, user.id, session.id)


@pytest.mark.asyncio
async def test_export_session_selection_checks_ownership(
    db_session: AsyncSession, user_factory
) -> None:
    user = await user_factory()
    other = await user_factory()
    session = await _create_session(db_session, user)
    db_session.add(_selection_row(user.id, session.id, "a"))
    await db_session.flush()

    with pytest.raises(LookupError, match="session_not_found"):
        await export_session_selection(db_session, other.id, session.id)


@pytest.mark.asyncio
async def test_export_session_selection_prefers_brainstorm_region(
    db_session: AsyncSession, user_factory
) -> None:
    user = await user_factory()
    session = await _create_session(
        db_session,
        user,
        filters_snapshot={
            "brainstorm_profile": {"region": "杭州市"},
            "target_fan_locations": ["旧地区"],
        },
    )
    db_session.add(_selection_row(user.id, session.id, "a"))
    await db_session.flush()

    workbook_out = await export_session_selection(db_session, user.id, session.id)

    summary = load_workbook(BytesIO(workbook_out.content))["KOL匹配度筛选"]
    region_header = summary.cell(4, 8).value or ""
    assert "杭州市" in region_header
    assert "旧地区" not in region_header
    assert "杭州市" in (summary.cell(2, 1).value or "")


@pytest.mark.asyncio
async def test_export_session_selection_falls_back_to_target_fan_locations(
    db_session: AsyncSession, user_factory
) -> None:
    user = await user_factory()
    session = await _create_session(
        db_session, user, filters_snapshot={"target_fan_locations": ["浙江", "湖州"]}
    )
    db_session.add(_selection_row(user.id, session.id, "a"))
    await db_session.flush()

    workbook_out = await export_session_selection(db_session, user.id, session.id)

    summary = load_workbook(BytesIO(workbook_out.content))["KOL匹配度筛选"]
    assert "浙江、湖州" in (summary.cell(2, 1).value or "")


@pytest.mark.asyncio
async def test_export_filename_sanitizes_brand_and_defaults_category(
    db_session: AsyncSession, user_factory
) -> None:
    user = await user_factory()
    session = await _create_session(db_session, user, brand='品/牌"A', category=None)
    db_session.add(_selection_row(user.id, session.id, "a"))
    await db_session.flush()

    workbook_out = await export_session_selection(db_session, user.id, session.id)

    assert "/" not in workbook_out.filename
    assert '"' not in workbook_out.filename
    assert re.fullmatch(r"品_牌_A_未分类_KOL匹配度分析_\d{8}_\d{4}\.xlsx", workbook_out.filename)


@pytest_asyncio.fixture
async def export_client_factory(db_session: AsyncSession):
    clients: list[AsyncClient] = []

    async def create() -> tuple[AsyncClient, User, str]:
        app = create_app()

        async def override_get_db() -> AsyncIterator[AsyncSession]:
            yield db_session

        app.dependency_overrides[get_db] = override_get_db
        now = datetime.now(UTC).replace(tzinfo=None)
        user = User(
            id=str(uuid4()),
            nickname="导出用户",
            role="user",
            status="active",
            created_at=now,
            updated_at=now,
        )
        db_session.add(user)
        await db_session.flush()
        session = await _create_session(db_session, user)
        login_session = LoginSession(
            id=str(uuid4()),
            user_id=user.id,
            refresh_token_hash=uuid4().hex + uuid4().hex,
            expires_at=now + timedelta(days=1),
            revoked_at=None,
            created_at=now,
            last_seen_at=now,
        )
        db_session.add(login_session)
        await db_session.flush()
        token = create_access_token(user_id=user.id, session_id=login_session.id, role="user")
        test_client = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
        test_client.headers["Authorization"] = f"Bearer {token}"
        clients.append(test_client)
        return test_client, user, session.id

    yield create
    for test_client in clients:
        await test_client.aclose()


@pytest.mark.asyncio
async def test_list_kol_selection_returns_total_and_items(
    export_client_factory, db_session: AsyncSession
) -> None:
    client, user, session_id = await export_client_factory()
    db_session.add(
        _selection_row(user.id, session_id, "low", nickname="低分达人", total=50.0,
                       rating="可考虑", stars="★★★")
    )
    db_session.add(
        _selection_row(user.id, session_id, "high", nickname="高分达人", total=90.0)
    )
    await db_session.flush()

    response = await client.get(f"/api/v1/sessions/{session_id}/kol-selection")

    assert response.status_code == 200
    body = response.json()
    assert body["total"] == 2
    assert [item["kol_uid"] for item in body["items"]] == ["high", "low"]
    first = body["items"][0]
    assert first["platform"] == "xiaohongshu"
    assert first["nickname"] == "高分达人"
    assert first["followers"] == 125000
    assert first["city"] == "杭州市"
    assert first["profile_url"] == "https://example.com/high"
    assert first["score"]["total"] == 90.0
    assert first["score"]["rating"] == "重点推荐"
    assert first["fields"]["export_fields"] == {}


@pytest.mark.asyncio
async def test_list_kol_selection_empty_and_foreign_session(export_client_factory) -> None:
    client, _user, session_id = await export_client_factory()

    empty = await client.get(f"/api/v1/sessions/{session_id}/kol-selection")
    assert empty.status_code == 200
    assert empty.json() == {"total": 0, "items": []}

    other_client, _other, _other_session = await export_client_factory()
    foreign = await other_client.get(f"/api/v1/sessions/{session_id}/kol-selection")
    assert foreign.status_code == 404


@pytest.mark.asyncio
async def test_export_kol_selection_streams_workbook(
    export_client_factory, db_session: AsyncSession
) -> None:
    client, user, session_id = await export_client_factory()
    db_session.add(_selection_row(user.id, session_id, "a", nickname="导出达人"))
    await db_session.flush()

    response = await client.get(f"/api/v1/sessions/{session_id}/kol-selection/export")

    assert response.status_code == 200
    assert response.headers["content-type"] == CONTENT_TYPE
    disposition = response.headers["content-disposition"]
    assert disposition.startswith("attachment; filename*=UTF-8''")
    assert "KOL" in disposition
    workbook = load_workbook(BytesIO(response.content), read_only=False)
    assert workbook.sheetnames == [
        "KOL匹配度筛选",
        "达人详细画像",
        "粉丝画像详情",
        "评分方法论与数据来源",
    ]
    summary = workbook["KOL匹配度筛选"]
    assert summary.cell(5, 3).value == "导出达人"


@pytest.mark.asyncio
async def test_export_kol_selection_empty_returns_409(export_client_factory) -> None:
    client, _user, session_id = await export_client_factory()

    response = await client.get(f"/api/v1/sessions/{session_id}/kol-selection/export")

    assert response.status_code == 409
    assert response.json()["detail"] == "NO_KOL_SELECTION"


@pytest.mark.asyncio
async def test_export_kol_selection_foreign_session_returns_404(
    export_client_factory, db_session: AsyncSession
) -> None:
    _client, user, session_id = await export_client_factory()
    db_session.add(_selection_row(user.id, session_id, "a"))
    await db_session.flush()
    other_client, _other, _other_session = await export_client_factory()

    response = await other_client.get(f"/api/v1/sessions/{session_id}/kol-selection/export")

    assert response.status_code == 404
