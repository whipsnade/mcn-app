"""analysis_report_v1 → workbook_v1 的确定性导出契约测试。"""

from __future__ import annotations

from hashlib import sha256
from io import BytesIO

import pytest
from openpyxl import load_workbook
from pydantic import ValidationError

from app.agent_artifacts.exporters import (
    ArtifactExportUnsupported,
    WorkbookLimits,
    WorkbookTechnicalLimitExceeded,
    export_artifact,
    render_workbook_v1,
    workbook_layout_digest,
)
from app.agent_artifacts.payloads.analysis_report import AnalysisReportV1


def build_analysis_report_payload(*, row_count: int = 3) -> dict:
    rows = [
        [
            f"平台-{index}",
            index,
            index + 0.5,
            (index + 1) / 10,
            f"2026-01-{index % 28 + 1:02d}",
            f"2026-01-{index % 28 + 1:02d}T12:00:00",
            f"https://example.com/posts/{index}",
            index % 2 == 0,
        ]
        for index in range(row_count)
    ]
    blocks = [
        {
            "block_type": "metric_cards",
            "id": "summary_metrics",
            "title": "综合指标",
            "cards": [{"key": "volume", "label": "声量", "value": 100, "value_type": "integer"}],
        },
        {
            "block_type": "typed_table",
            "id": "platform_rows",
            "title": "跨平台明细",
            "columns": [
                {"key": "platform", "label": "平台", "type": "string"},
                {"key": "rank", "label": "排名", "type": "integer"},
                {"key": "score", "label": "得分", "type": "number"},
                {"key": "share", "label": "占比", "type": "percent"},
                {"key": "published", "label": "发布日期", "type": "date"},
                {"key": "observed_at", "label": "观测时间", "type": "datetime"},
                {"key": "url", "label": "链接", "type": "url"},
                {"key": "is_paid", "label": "付费", "type": "boolean"},
            ],
            "rows": rows,
        },
        {
            "block_type": "time_series",
            "id": "trend",
            "title": "趋势",
            "points": [{"timestamp": "2026-01-01", "values": {"volume": 100}}],
        },
        {
            "block_type": "link_list",
            "id": "sources",
            "title": "来源",
            "items": [{"label": "帖子", "url": "https://example.com/posts/0"}],
        },
        {
            "block_type": "chart",
            "id": "volume_chart",
            "title": "声量图",
            "chart_type": "bar",
            "categories": ["小红书", "抖音"],
            "series": [{"key": "volume", "label": "声量", "values": [100, 80]}],
        },
        {
            "block_type": "narrative",
            "id": "summary",
            "title": "摘要",
            "content": "真实数据已按平台统一列展示。",
        },
        {
            "block_type": "methodology_limitations",
            "id": "methodology",
            "title": "方法与限制",
            "methodology": "来源为已审核营销数据工具。",
            "limitations": [],
        },
    ]
    return {
        "schema_version": "analysis_report_v1",
        "module": "report",
        "data_status": "complete",
        "availability": {
            "blocks": {"status": "complete", "reason_codes": []},
            "fulfillment": {"status": "complete", "reason_codes": []},
        },
        "limitations": [],
        "methodology": {
            "data_as_of": "2026-01-15T12:00:00",
            "source_names": ["DataTap"],
            "notes": [],
        },
        "title": "跨平台营销分析",
        "subject_type": "mixed",
        "scope": {"brand": "示例品牌", "platforms": ["xiaohongshu", "douyin"]},
        "blocks": blocks,
        "fulfillment": [{
            "key": "requested_items",
            "requested_min": row_count,
            "actual_count": row_count,
            "status": "complete",
            "reason": "真实返回数量满足要求",
        }],
        "workbook": {
            "schema_version": "workbook_v1",
            "sheets": [{
                "key": "overview",
                "title": "总览",
                "block_ids": ["summary_metrics", "summary"],
                "freeze_rows": 1,
                "auto_filter": False,
            }, {
                "key": "platforms",
                "title": "平台明细",
                "block_ids": ["platform_rows"],
                "columns": [
                    {"key": "platform", "label": "渠道", "width": 18},
                    {"key": "rank", "label": "序号", "width": 10},
                    {"key": "score", "label": "统一得分", "number_format": "0.0"},
                    {"key": "share", "label": "占比", "number_format": "0.0%"},
                    {"key": "published", "label": "发布日期"},
                    {"key": "observed_at", "label": "观测时间"},
                    {"key": "url", "label": "原文"},
                    {"key": "is_paid", "label": "是否付费"},
                ],
                "freeze_rows": 1,
                "auto_filter": True,
                "sort_by": ["rank"],
            }],
        },
    }


class _Version:
    def __init__(self, schema_version: str, payload_json: dict | None) -> None:
        self.schema_version = schema_version
        self.payload_json = payload_json
        self.data_status = "complete" if payload_json else "draft"


def test_workbook_layout_digest_is_stable_and_changes_with_layout() -> None:
    report = AnalysisReportV1.model_validate(build_analysis_report_payload())
    assert workbook_layout_digest(report.workbook) == workbook_layout_digest(report.workbook)
    changed = report.workbook.model_copy(update={"sheets": ()})
    assert workbook_layout_digest(report.workbook) != workbook_layout_digest(changed)


def test_analysis_report_exports_deterministic_workbook_with_custom_columns_and_links() -> None:
    payload = build_analysis_report_payload()
    report = AnalysisReportV1.model_validate(payload)
    limits = WorkbookLimits(
        max_sheets=8,
        max_rows_per_sheet=100,
        max_columns=16,
        max_cell_chars=32767,
        max_bytes=2_000_000,
    )
    first = render_workbook_v1(report, exporter_version="analysis-report-v1", limits=limits)
    second = render_workbook_v1(report, exporter_version="analysis-report-v1", limits=limits)
    assert sha256(first).hexdigest() == sha256(second).hexdigest()
    workbook = load_workbook(BytesIO(first), data_only=False)
    assert tuple(workbook.sheetnames) == ("总览", "平台明细")
    sheet = workbook["平台明细"]
    assert sheet["A1"].value == "渠道"
    assert sheet["C1"].number_format == "General"
    assert sheet["A2"].value == "平台-0"
    assert sheet["G2"].hyperlink.target == "https://example.com/posts/0"
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:H4"
    assert sheet["D2"].number_format == "0.0%"


def test_analysis_report_paginates_without_dropping_business_rows() -> None:
    payload = build_analysis_report_payload(row_count=45)
    report = AnalysisReportV1.model_validate(payload)
    content = render_workbook_v1(
        report,
        exporter_version="analysis-report-v1",
        limits=WorkbookLimits(
            max_sheets=16,
            max_rows_per_sheet=10,
            max_columns=16,
            max_cell_chars=128,
            max_bytes=2_000_000,
        ),
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    values = [cell.value for sheet in workbook.worksheets for row in sheet.iter_rows() for cell in row]
    assert len([value for value in values if isinstance(value, str) and value.startswith("平台-")]) == 45
    assert len(workbook.worksheets) >= 5
    assert all(sheet.max_row <= 10 for sheet in workbook.worksheets)


def test_analysis_report_export_escapes_formula_like_text_and_rejects_unsafe_urls() -> None:
    payload = build_analysis_report_payload()
    report = AnalysisReportV1.model_validate(payload).model_copy(update={"title": "@外部输入"})
    content = render_workbook_v1(
        report,
        exporter_version="analysis-report-v1",
        limits=WorkbookLimits(
            max_sheets=8,
            max_rows_per_sheet=100,
            max_columns=16,
            max_cell_chars=32767,
            max_bytes=2_000_000,
        ),
    )
    workbook = load_workbook(BytesIO(content), data_only=False)
    values = [cell.value for row in workbook["总览"].iter_rows() for cell in row]
    assert any(isinstance(value, str) and value.startswith("'") for value in values)
    with pytest.raises(ValidationError):
        invalid = build_analysis_report_payload()
        invalid["blocks"][1]["rows"][0][6] = "javascript:alert(1)"
        AnalysisReportV1.model_validate(invalid)


def test_analysis_report_export_rejects_technical_limit_without_truncating() -> None:
    report = AnalysisReportV1.model_validate(build_analysis_report_payload(row_count=45))
    with pytest.raises(WorkbookTechnicalLimitExceeded) as error:
        render_workbook_v1(
            report,
            exporter_version="analysis-report-v1",
            limits=WorkbookLimits(
                max_sheets=2,
                max_rows_per_sheet=10,
                max_columns=16,
                max_cell_chars=32767,
                max_bytes=2_000_000,
            ),
        )
    assert error.value.code == "workbook_technical_limit_exceeded"


def test_export_dispatch_supports_analysis_report_and_keeps_draft_unsupported() -> None:
    content = export_artifact(_Version("analysis_report_v1", build_analysis_report_payload()))
    assert content[:2] == b"PK"
    with pytest.raises(ArtifactExportUnsupported):
        export_artifact(_Version("analysis_report_v1", None))
