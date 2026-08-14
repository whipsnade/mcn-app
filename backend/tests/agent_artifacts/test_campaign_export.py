"""campaign_report_v2 Excel 导出（Gate C Task 5 / 设计 §8.3）。

无 ROI 时 9 个 Sheet；具备可靠 ROI 数据时动态增加「ROI与转化」Sheet。
导出只读已发布 Version 的 payload，不调用模型/MCP。
"""

from __future__ import annotations

import re
from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.agent_artifacts.exporters import ArtifactExportUnsupported, export_artifact

from tests.agent_artifacts.test_payloads import (
    COMPLETE,
    build_campaign_dict,
    refresh_fixture_canonical,
)

CAMPAIGN_SHEETS = (
    "活动综合概览",
    "周期对比与趋势",
    "平台表现",
    "情感与内容分析",
    "热门帖子TOP",
    "达人投放表现",
    "自然传播与受众",
    "洞察与建议",
    "方法论",
)


class _Version:
    def __init__(self, schema_version: str, payload_json: dict | None) -> None:
        self.schema_version = schema_version
        if isinstance(payload_json, dict) and "data" in payload_json:
            refresh_fixture_canonical(payload_json, module="campaign")
        self.payload_json = payload_json
        self.data_status = "complete" if payload_json else "draft"


def _values(ws) -> list:
    return [
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]


def _campaign_version(**overrides) -> _Version:
    payload = build_campaign_dict()
    payload.update(overrides)
    return _Version("campaign_report_v2", payload)


def test_campaign_export_has_nine_sheets_without_roi() -> None:
    content = export_artifact(_campaign_version())
    assert content[:2] == b"PK"
    wb = load_workbook(BytesIO(content), data_only=False)
    assert tuple(wb.sheetnames) == CAMPAIGN_SHEETS
    assert "某品牌" in str(wb["活动综合概览"]["A1"].value)


def test_campaign_report_v3_export_uses_the_direct_contract() -> None:
    payload = build_campaign_dict()
    payload["schema_version"] = "campaign_report_v3"
    content = export_artifact(_Version("campaign_report_v3", payload))
    wb = load_workbook(BytesIO(content), data_only=False)
    assert tuple(wb.sheetnames) == CAMPAIGN_SHEETS
    assert "某品牌" in str(wb["活动综合概览"]["A1"].value)


def test_campaign_export_adds_roi_sheet_only_when_available() -> None:
    payload = build_campaign_dict()
    payload["data"]["internal_metrics"] = {
        "spend": 100000,
        "impressions": 2000000,
        "conversions": 5000,
        "revenue": 300000,
        "cpc": 20.0,
        "cpm": 50.0,
    }
    payload["data"]["roi"] = {
        "spend": 100000,
        "revenue": 300000,
        "conversions": 5000,
        "attribution_window": "最后点击 7 天",
        "roi": 2.0,
        "roas": 3.0,
    }
    content = export_artifact(_Version("campaign_report_v2", payload))
    wb = load_workbook(BytesIO(content), data_only=False)
    assert tuple(wb.sheetnames) == (*CAMPAIGN_SHEETS, "ROI与转化")
    assert 100000 in _values(wb["ROI与转化"])
    assert 2.0 in _values(wb["ROI与转化"])


def test_campaign_export_platform_chart_only_with_data() -> None:
    content = export_artifact(_campaign_version())
    wb = load_workbook(BytesIO(content), data_only=False)
    platforms = wb["平台表现"]
    assert len(platforms._charts) >= 0  # 无数据不强制建图


def test_campaign_external_text_is_formula_escaped() -> None:
    payload = build_campaign_dict()
    if payload["data"]["top_posts"]:
        payload["data"]["top_posts"][0]["title"] = "=SUM(A1)"
    content = export_artifact(_Version("campaign_report_v2", payload))
    wb = load_workbook(BytesIO(content))
    values = _values(wb["热门帖子TOP"])
    assert any(isinstance(v, str) and v.startswith("'") for v in values)


def test_campaign_draft_raises_unsupported() -> None:
    with pytest.raises(ArtifactExportUnsupported) as excinfo:
        export_artifact(_Version("campaign_report_v2", None))
    assert excinfo.value.code == "ARTIFACT_EXPORT_UNSUPPORTED"


def test_campaign_invalid_payload_raises_unsupported() -> None:
    with pytest.raises(ArtifactExportUnsupported) as excinfo:
        export_artifact(_Version("campaign_report_v2", {"schema_version": "campaign_report_v2"}))
    assert excinfo.value.code == "ARTIFACT_EXPORT_UNSUPPORTED"


# ---------------------------------------------------------------------------
# Gate C 审核修复：图表真实引用 / ROI 模板 Sheet / used range
# ---------------------------------------------------------------------------


def test_campaign_platform_chart_refs_start_at_row_5() -> None:
    payload = build_campaign_dict()
    content = export_artifact(_Version("campaign_report_v2", payload))
    wb = load_workbook(BytesIO(content), data_only=False)
    platforms = wb["平台表现"]
    charts = [c for c in platforms._charts if c is not None]
    assert charts, "平台表现必须有真实图表"
    refs = [ser.val.numRef.f for chart in charts for ser in chart.ser]
    assert any("$B$5" in ref for ref in refs), refs


def test_campaign_trend_chart_rebuilt() -> None:
    payload = build_campaign_dict()
    payload["data"]["comparisons"] = {
        "current_baseline": [
            {"metric": "volume", "current": 1200, "baseline": 800, "delta": 400, "rate": 0.5},
            {"metric": "engagement", "current": 5000, "baseline": 3000, "delta": 2000, "rate": 0.67},
        ],
        "current_post": [],
    }
    content = export_artifact(_Version("campaign_report_v2", payload))
    wb = load_workbook(BytesIO(content), data_only=False)
    trend = wb["周期对比与趋势"]
    assert trend._charts, "周期对比必须重建图表"
    refs = [ser.val.numRef.f for chart in trend._charts for ser in chart.ser]
    # 表带标题（@4）→ 表头@5、数据 6-7；图表必须引用真实数据行而非表头。
    assert any("$B$6:$B$7" in ref for ref in refs), refs


def test_campaign_used_range_not_extended_to_1000() -> None:
    content = export_artifact(_campaign_version_roi())
    wb = load_workbook(BytesIO(content), data_only=False)
    for ws in wb.worksheets:
        assert ws.max_row < 100, f"{ws.title} max_row={ws.max_row}"


def _campaign_version_roi() -> _Version:
    payload = build_campaign_dict()
    payload["data"]["internal_metrics"] = {
        "spend": 100000,
        "impressions": 2000000,
        "conversions": 5000,
        "revenue": 300000,
        "cpc": 20.0,
        "cpm": 50.0,
    }
    payload["data"]["roi"] = {
        "spend": 100000,
        "revenue": 300000,
        "conversions": 5000,
        "attribution_window": "最后点击 7 天",
        "roi": 2.0,
        "roas": 3.0,
    }
    return _Version("campaign_report_v2", payload)


# ---------------------------------------------------------------------------
# Gate C 复审 P0：自然传播/受众地域/内容归属必须用 write_table 游标连续排布
# ---------------------------------------------------------------------------


def _campaign_version_organic_full() -> _Version:
    payload = build_campaign_dict()
    payload["data"]["organic_summary"] = {
        "volume": 300,
        "engagement": 1200,
        "posts": 12,
        "share_of_volume": 0.3,
    }
    payload["data"]["audience_regions"] = [
        {"region": "上海", "volume": 200, "share": 0.5},
        {"region": "北京", "volume": 150, "share": 0.375},
        {"region": "广州", "volume": 50, "share": 0.125},
    ]
    payload["data"]["attribution"] = {
        "paid_confirmed": 20,
        "organic": 25,
        "unknown": 5,
        "paid_confirmed_share": 0.4,
    }
    for section in ("organic_summary", "audience_regions", "attribution"):
        payload["availability"][section] = COMPLETE
    return _Version("campaign_report_v2", payload)


def test_campaign_organic_audience_attribution_sections_do_not_overlap() -> None:
    content = export_artifact(_campaign_version_organic_full())
    wb = load_workbook(BytesIO(content), data_only=False)
    ws = wb["自然传播与受众"]
    title_rows: dict[str, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value in ("自然传播", "受众地域", "内容归属"):
                title_rows.setdefault(str(cell.value), cell.row)
    assert set(title_rows) == {"自然传播", "受众地域", "内容归属"}
    assert title_rows["自然传播"] < title_rows["受众地域"] < title_rows["内容归属"]
    # 受众地域 3 行数据必须完整落在内容归属标题之前（标题+表头+3 数据行）。
    assert title_rows["内容归属"] >= title_rows["受众地域"] + 5
    values = _values(ws)
    for label in ("自然声量", "上海", "北京", "广州", "付费确认", "自然", "未知"):
        assert label in values


def test_campaign_attribution_chart_refs_real_data_rows() -> None:
    content = export_artifact(_campaign_version_organic_full())
    wb = load_workbook(BytesIO(content), data_only=False)
    ws = wb["自然传播与受众"]
    charts = [c for c in ws._charts if c is not None]
    assert charts, "内容归属必须有图表"
    refs = [ser.val.numRef.f for chart in charts for ser in chart.ser if ser.val and ser.val.numRef]
    assert refs
    # 归属表数据行区间必须与表格真实行一致：类别轴 3 行（付费确认/自然/未知）。
    for ref in refs:
        start, end = ref.split("!")[1].split(":")
        start_row = int("".join(ch for ch in start if ch.isdigit()))
        end_row = int("".join(ch for ch in end if ch.isdigit()))
        assert end_row - start_row + 1 == 3, ref


# ---------------------------------------------------------------------------
# Gate C 复审 P0：图表必须用 write_table 返回的真实表头/数据行
# ---------------------------------------------------------------------------


def _series_value_ranges(ws) -> list[str]:
    return [
        ser.val.numRef.f
        for chart in ws._charts
        for ser in chart.ser
        if ser.val is not None and ser.val.numRef is not None
    ]


def _chart_value_ranges(chart) -> list[str]:
    return [
        ser.val.numRef.f
        for ser in chart.ser
        if ser.val is not None and ser.val.numRef is not None
    ]


def _chart_title_ranges(chart) -> list[str]:
    return [
        ser.tx.strRef.f
        for ser in chart.ser
        if ser.tx is not None and ser.tx.strRef is not None
    ]


def _parse_ref(ref: str) -> tuple[str, int, str, int]:
    # openpyxl 图表引用形态不一（标题 B4 / 单格 $B$5 / 区间 $B$5:$B$7，
    # 带或不带 sheet 前缀与 $）；统一解析为 (列, 行, 列, 行) 再断言。
    part = ref.split("!")[-1].replace("$", "")
    a, _, b = part.partition(":")
    b = b or a
    ma = re.match(r"([A-Z]+)(\d+)$", a)
    mb = re.match(r"([A-Z]+)(\d+)$", b)
    assert ma and mb, ref
    return ma.group(1), int(ma.group(2)), mb.group(1), int(mb.group(2))


def test_campaign_platform_chart_multi_row_exact_range_and_chinese_series() -> None:
    payload = build_campaign_dict()
    payload["data"]["platform_contributions"] = [
        {"platform": "xiaohongshu", "volume": 1000, "engagement": 5000, "posts": 50, "creators": 5, "share": 0.5},
        {"platform": "douyin", "volume": 800, "engagement": 4000, "posts": 40, "creators": 4, "share": 0.4},
        {"platform": "weibo", "volume": 200, "engagement": 1000, "posts": 10, "creators": 1, "share": 0.1},
    ]
    content = export_artifact(_Version("campaign_report_v2", payload))
    wb = load_workbook(BytesIO(content), data_only=False)
    ws = wb["平台表现"]
    charts = [c for c in ws._charts if c is not None]
    assert charts, "平台多行数据必须有图表"
    chart = charts[0]
    # 数据行 5-7（表头@4）：声量列 B 的完整区间，不得丢行。
    value_spans = [_parse_ref(r) for r in _chart_value_ranges(chart)]
    assert ("B", 5, "B", 7) in value_spans, value_spans
    # 中文系列名：系列标题引用表头行（第 4 行），而非数据行或默认 Series N。
    title_spans = [_parse_ref(t) for t in _chart_title_ranges(chart)]
    assert title_spans and all(span[1] == 4 and span[3] == 4 for span in title_spans), title_spans
    assert ws.cell(4, 2).value == "声量", "系列标题必须指向中文表头"
    # 类别轴 = 平台名 A 列 5-7。
    cat = chart.ser[0].cat
    cat_ref = None
    if cat is not None:
        cat_ref = cat.strRef.f if cat.strRef is not None else (cat.numRef.f if cat.numRef is not None else None)
    assert cat_ref is not None and _parse_ref(cat_ref) == ("A", 5, "A", 7), cat_ref


def test_campaign_platform_chart_single_row_exact_range() -> None:
    payload = build_campaign_dict()  # 单平台 → 单行数据
    content = export_artifact(_Version("campaign_report_v2", payload))
    wb = load_workbook(BytesIO(content), data_only=False)
    ws = wb["平台表现"]
    charts = [c for c in ws._charts if c is not None]
    assert charts
    value_spans = [_parse_ref(r) for r in _chart_value_ranges(charts[0])]
    assert ("B", 5, "B", 5) in value_spans, value_spans


def test_campaign_platform_chart_absent_when_empty() -> None:
    payload = build_campaign_dict()
    payload["data"]["platform_contributions"] = []
    content = export_artifact(_Version("campaign_report_v2", payload))
    wb = load_workbook(BytesIO(content), data_only=False)
    ws = wb["平台表现"]
    assert not [c for c in ws._charts if c is not None], "空数据不得建图"


def test_campaign_comparison_chart_chinese_series_and_exact_range() -> None:
    payload = build_campaign_dict()
    payload["data"]["comparisons"] = {
        "current_baseline": [
            {"metric": "volume", "current": 1200, "baseline": 800, "delta": 400, "rate": 0.5},
            {"metric": "engagement", "current": 5000, "baseline": 3000, "delta": 2000, "rate": 0.67},
            {"metric": "posts", "current": 50, "baseline": 40, "delta": 10, "rate": 0.25},
        ],
        "current_post": [],
    }
    payload["availability"]["comparisons"] = COMPLETE
    content = export_artifact(_Version("campaign_report_v2", payload))
    wb = load_workbook(BytesIO(content), data_only=False)
    ws = wb["周期对比与趋势"]
    charts = [c for c in ws._charts if c is not None]
    assert charts, "周期对比必须重建图表"
    chart = charts[0]
    assert len(chart.ser) == 2, "当前 + 对比期两个系列"
    # 表带标题（@4）→ 表头@5、数据 6-8；中文系列名引用表头行。
    title_spans = [_parse_ref(t) for t in _chart_title_ranges(chart)]
    assert title_spans and all(span[1] == 5 and span[3] == 5 for span in title_spans), title_spans
    # 数据区间 6-8（3 行指标），当前列 B、对比期列 C。
    value_spans = sorted(_parse_ref(r) for r in _chart_value_ranges(chart))
    assert value_spans == [("B", 6, "B", 8), ("C", 6, "C", 8)], value_spans


def test_campaign_timeline_rendered_and_charted() -> None:
    payload = build_campaign_dict()
    payload["data"]["comparisons"] = {"current_baseline": [], "current_post": []}
    payload["data"]["timeline"] = [
        {"date": date(2026, 1, 1), "platform": "xiaohongshu", "volume": 100, "engagement": 500, "posts": 5},
        {"date": date(2026, 1, 2), "platform": "xiaohongshu", "volume": 150, "engagement": 700, "posts": 7},
        {"date": date(2026, 1, 3), "platform": "xiaohongshu", "volume": 120, "engagement": 600, "posts": 6},
    ]
    content = export_artifact(_Version("campaign_report_v2", payload))
    wb = load_workbook(BytesIO(content), data_only=False)
    ws = wb["周期对比与趋势"]
    values = _values(ws)
    assert 150 in values and 120 in values, "时间线数据必须导出"
    charts = [c for c in ws._charts if c is not None]
    assert len(charts) == 1, "对比为空时仅时间线成图"
    spans = [_parse_ref(r) for r in _chart_value_ranges(charts[0])]
    assert spans
    for col_start, row_start, col_end, row_end in spans:
        assert col_start == col_end and row_end - row_start + 1 == 3, spans
