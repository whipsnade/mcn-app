"""Task 19 API 测试：/api/v1/agent 的 Artifact 读取 API（设计 §15.2）。

覆盖：
1. 列表（module / parent_artifact_id 过滤）；
2. 详情 + 版本；
3. 未读水位：monotonic max 且后端校验不超过当前 Session sequence；
4. 导出：支持类型返回 xlsx bytes，不支持 409 ARTIFACT_EXPORT_UNSUPPORTED；
5. 全部归属失败 → 404（无存在泄露）。
"""

from __future__ import annotations

import json
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.agent_artifacts.models import (
    AgentArtifact,
    AgentArtifactVersion,
    ArtifactDraft,
    ArtifactDraftRevision,
    ArtifactEvent,
)
from app.agent_runtime.models import AgentRun
from app.agent_runtime.repository import utc_now
from tests.agent_artifacts.test_payloads import build_brand_dict


def _json_safe(payload: dict) -> dict:
    """把含 date/datetime 的 payload 转成 JSON 可落库形式（ISO 字符串）。"""
    return json.loads(json.dumps(payload, ensure_ascii=False, default=str))


async def _me_id(client) -> str:
    resp = await client.get("/api/v1/users/me")
    assert resp.status_code == 200, resp.text
    return resp.json()["id"]


async def _create_session(client) -> str:
    resp = await client.post("/api/v1/agent/sessions", json={})
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


async def _make_published_artifact(
    db: AsyncSession,
    user_id: str,
    session_id: str,
    *,
    module: str = "brand",
    artifact_type: str = "brand_report_v3",
    artifact_key: str | None = None,
    payload: dict | None = None,
    parent_artifact_id: str | None = None,
    latest_version: int = 1,
    data_status: str = "complete",
) -> AgentArtifact:
    """创建 user/session 下的一条已发布 Artifact（含 run + draft + revision + version）。"""
    if payload is not None:
        payload = _json_safe(payload)
    now = utc_now()
    run = AgentRun(
        id=str(uuid4()),
        session_id=session_id,
        user_id=user_id,
        run_kind="user",
        visibility="user",
        profile_name="session_analyst_v1",
        profile_version="v1",
        model="test-model",
        status="completed",
        decision_count=0,
        review_count=0,
        revision_count=0,
        completed_at=now,
    )
    db.add(run)
    artifact = AgentArtifact(
        id=str(uuid4()),
        session_id=session_id,
        user_id=user_id,
        module=module,
        artifact_type=artifact_type,
        parent_artifact_id=parent_artifact_id,
        artifact_key=artifact_key or f"{module}:{uuid4().hex[:8]}",
        status="published",
        latest_version=latest_version,
        activity_sequence=0,
        created_at=now,
        updated_at=now,
    )
    db.add(artifact)
    draft = ArtifactDraft(
        id=str(uuid4()),
        artifact_id=artifact.id,
        session_id=session_id,
        owner_run_id=None,
        current_revision=1,
        status="idle",
        review_count=0,
        revision_count=1,
        updated_at=now,
    )
    db.add(draft)
    # 依赖序 flush：AgentRun 自引用 parent_run_id 且 revision↔version 存在循环 FK
    # （parent_artifact_version_id use_alter），unit-of-work 不保证插入顺序，必须
    # 逐层落库保证 FK 目标存在。
    await db.flush()  # run / artifact / draft
    revision = ArtifactDraftRevision(
        id=str(uuid4()),
        draft_id=draft.id,
        artifact_id=artifact.id,
        run_id=run.id,
        revision=1,
        schema_version=artifact_type,
        payload_json=payload,
        evidence_refs_json=[],
        parent_artifact_version_id=None,
        payload_hash="h",
        created_at=now,
    )
    db.add(revision)
    await db.flush()  # revision 先落库，version.source_draft_revision_id 才有目标
    version = AgentArtifactVersion(
        id=str(uuid4()),
        artifact_id=artifact.id,
        version=1,
        source_run_id=run.id,
        source_draft_revision_id=revision.id,
        parent_artifact_version_id=None,
        schema_version=artifact_type,
        payload_json=payload,
        evidence_refs_json=[],
        review_json=None,
        data_status=data_status,
        created_at=now,
    )
    db.add(version)
    await db.flush()
    return artifact


async def _add_artifact_event(
    db: AsyncSession, user_id: str, session_id: str, artifact_id: str, sequence: int
) -> None:
    db.add(
        ArtifactEvent(
            id=str(uuid4()),
            session_id=session_id,
            user_id=user_id,
            sequence=sequence,
            module="brand",
            artifact_id=artifact_id,
            event_type="published",
            draft_revision=None,
            artifact_version_id=None,
            created_at=utc_now(),
        )
    )
    await db.flush()


async def _add_published_version(
    db: AsyncSession, artifact: AgentArtifact, *, version: int, payload: dict
) -> AgentArtifactVersion:
    """给已有 Artifact 追加一个已发布 Version（新 revision + version，latest_version 前进）。"""
    payload = _json_safe(payload)
    now = utc_now()
    first = await db.scalar(
        select(AgentArtifactVersion).where(
            AgentArtifactVersion.artifact_id == artifact.id,
            AgentArtifactVersion.version == 1,
        )
    )
    assert first is not None
    draft = await db.scalar(
        select(ArtifactDraft).where(ArtifactDraft.artifact_id == artifact.id)
    )
    assert draft is not None
    revision = ArtifactDraftRevision(
        id=str(uuid4()),
        draft_id=draft.id,
        artifact_id=artifact.id,
        run_id=first.source_run_id,
        revision=version,
        schema_version=artifact.artifact_type,
        payload_json=payload,
        evidence_refs_json=[],
        parent_artifact_version_id=first.id,
        payload_hash="h",
        created_at=now,
    )
    db.add(revision)
    await db.flush()  # revision 先落库，version.source_draft_revision_id 才有目标
    row = AgentArtifactVersion(
        id=str(uuid4()),
        artifact_id=artifact.id,
        version=version,
        source_run_id=first.source_run_id,
        source_draft_revision_id=revision.id,
        parent_artifact_version_id=first.id,
        schema_version=artifact.artifact_type,
        payload_json=payload,
        evidence_refs_json=[],
        review_json=None,
        data_status="complete",
        created_at=now,
    )
    db.add(row)
    artifact.latest_version = version
    await db.flush()
    return row


# ---------------------------------------------------------------------------
# 列表 / 详情 / 版本
# ---------------------------------------------------------------------------


async def test_list_artifacts_filters_by_module_and_parent(auth_client_factory, db_session) -> None:
    client = await auth_client_factory("13700000001")
    session_id = await _create_session(client)
    user_id = await _me_id(client)

    brand = await _make_published_artifact(
        db_session, user_id, session_id, module="brand", artifact_type="brand_report_v3"
    )
    campaign = await _make_published_artifact(
        db_session,
        user_id,
        session_id,
        module="campaign",
        artifact_type="campaign_report_v2",
        artifact_key="campaign:brand:x",
    )
    insight = await _make_published_artifact(
        db_session,
        user_id,
        session_id,
        module="brand",
        artifact_type="insight_board_v1",
        artifact_key=f"insight:{uuid4().hex[:8]}",
        parent_artifact_id=brand.id,
    )

    listed = await client.get(f"/api/v1/agent/sessions/{session_id}/artifacts")
    assert listed.status_code == 200
    assert {item["id"] for item in listed.json()} == {brand.id, campaign.id, insight.id}

    brand_only = await client.get(
        f"/api/v1/agent/sessions/{session_id}/artifacts", params={"module": "brand"}
    )
    assert {item["id"] for item in brand_only.json()} == {brand.id, insight.id}

    children = await client.get(
        f"/api/v1/agent/sessions/{session_id}/artifacts",
        params={"parent_artifact_id": brand.id},
    )
    assert [item["id"] for item in children.json()] == [insight.id]


async def test_get_artifact_and_version(auth_client_factory, db_session) -> None:
    client = await auth_client_factory("13700000002")
    session_id = await _create_session(client)
    user_id = await _me_id(client)
    artifact = await _make_published_artifact(db_session, user_id, session_id)

    detail = await client.get(f"/api/v1/agent/artifacts/{artifact.id}")
    assert detail.status_code == 200
    assert detail.json()["artifact_type"] == "brand_report_v3"
    assert detail.json()["status"] == "published"

    version = await client.get(f"/api/v1/agent/artifacts/{artifact.id}/versions/1")
    assert version.status_code == 200
    assert version.json()["schema_version"] == "brand_report_v3"
    assert version.json()["data_status"] == "complete"

    missing = await client.get(f"/api/v1/agent/artifacts/{artifact.id}/versions/99")
    assert missing.status_code == 404


# ---------------------------------------------------------------------------
# 未读水位
# ---------------------------------------------------------------------------


async def test_read_state_monotonic_max_and_sequence_validation(
    auth_client_factory, db_session
) -> None:
    client = await auth_client_factory("13700000003")
    session_id = await _create_session(client)
    user_id = await _me_id(client)
    artifact = await _make_published_artifact(db_session, user_id, session_id)
    # Session 级 artifact sequence 当前为 3
    await _add_artifact_event(db_session, user_id, session_id, artifact.id, 1)
    await _add_artifact_event(db_session, user_id, session_id, artifact.id, 2)
    await _add_artifact_event(db_session, user_id, session_id, artifact.id, 3)

    first = await client.put(
        f"/api/v1/agent/sessions/{session_id}/artifact-read-state",
        json={"module": "brand", "last_seen_sequence": 2},
    )
    assert first.status_code == 200
    assert first.json()["last_seen_sequence"] == 2

    # 更小的 stale 写入不后退（monotonic max）
    stale = await client.put(
        f"/api/v1/agent/sessions/{session_id}/artifact-read-state",
        json={"module": "brand", "last_seen_sequence": 1},
    )
    assert stale.status_code == 200
    assert stale.json()["last_seen_sequence"] == 2

    # 超过当前 Session sequence → 422
    overshoot = await client.put(
        f"/api/v1/agent/sessions/{session_id}/artifact-read-state",
        json={"module": "brand", "last_seen_sequence": 5},
    )
    assert overshoot.status_code == 422


async def test_get_read_states_returns_current_user_watermarks(
    auth_client_factory, db_session
) -> None:
    """GET artifact-read-states 返回该会话当前用户全部模块水位（含空会话/跨会话隔离）。"""
    client = await auth_client_factory("13700000013")
    session_id = await _create_session(client)
    other_session_id = await _create_session(client)
    user_id = await _me_id(client)
    artifact = await _make_published_artifact(db_session, user_id, session_id)
    await _add_artifact_event(db_session, user_id, session_id, artifact.id, 1)
    await _add_artifact_event(db_session, user_id, session_id, artifact.id, 2)

    # 空会话（无任何水位写入）→ 空列表
    empty = await client.get(f"/api/v1/agent/sessions/{other_session_id}/artifact-read-states")
    assert empty.status_code == 200
    assert empty.json() == []

    await client.put(
        f"/api/v1/agent/sessions/{session_id}/artifact-read-state",
        json={"module": "brand", "last_seen_sequence": 2},
    )
    await client.put(
        f"/api/v1/agent/sessions/{session_id}/artifact-read-state",
        json={"module": "kol", "last_seen_sequence": 1},
    )
    # 同用户另一会话的水位不应混入
    await client.put(
        f"/api/v1/agent/sessions/{other_session_id}/artifact-read-state",
        json={"module": "campaign", "last_seen_sequence": 0},
    )

    resp = await client.get(f"/api/v1/agent/sessions/{session_id}/artifact-read-states")
    assert resp.status_code == 200
    states = {item["module"]: item["last_seen_sequence"] for item in resp.json()}
    assert states == {"brand": 2, "kol": 1}


# ---------------------------------------------------------------------------
# 导出
# ---------------------------------------------------------------------------


async def test_export_supported_type_returns_xlsx(auth_client_factory, db_session) -> None:
    client = await auth_client_factory("13700000004")
    session_id = await _create_session(client)
    user_id = await _me_id(client)
    artifact = await _make_published_artifact(
        db_session, user_id, session_id, payload=build_brand_dict()
    )

    resp = await client.get(f"/api/v1/agent/artifacts/{artifact.id}/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "filename" in resp.headers["content-disposition"]
    assert resp.content[:2] == b"PK"
    wb = load_workbook(BytesIO(resp.content))
    assert "综合概览" in wb.sheetnames


async def test_export_explicit_version_returns_that_version(
    auth_client_factory, db_session
) -> None:
    """?version=N 导出指定历史版本；不传时保持最新版本语义。"""
    client = await auth_client_factory("13700000014")
    session_id = await _create_session(client)
    user_id = await _me_id(client)

    payload_v1 = build_brand_dict()
    payload_v1["scope"]["brand"] = "旧品牌"
    artifact = await _make_published_artifact(
        db_session, user_id, session_id, payload=payload_v1
    )
    payload_v2 = build_brand_dict()
    payload_v2["scope"]["brand"] = "新品牌"
    await _add_published_version(db_session, artifact, version=2, payload=payload_v2)

    # 不传 version → 最新版本（v2）
    latest = await client.get(f"/api/v1/agent/artifacts/{artifact.id}/export")
    assert latest.status_code == 200
    assert "_v2.xlsx" in latest.headers["content-disposition"]
    assert "新品牌" in load_workbook(BytesIO(latest.content))["综合概览"]["A1"].value

    # 显式 version=1 → 历史版本内容
    historical = await client.get(
        f"/api/v1/agent/artifacts/{artifact.id}/export", params={"version": 1}
    )
    assert historical.status_code == 200
    assert "_v1.xlsx" in historical.headers["content-disposition"]
    assert "旧品牌" in load_workbook(BytesIO(historical.content))["综合概览"]["A1"].value

    # 显式 version=2 → 与最新一致
    explicit = await client.get(
        f"/api/v1/agent/artifacts/{artifact.id}/export", params={"version": 2}
    )
    assert explicit.status_code == 200
    assert "新品牌" in load_workbook(BytesIO(explicit.content))["综合概览"]["A1"].value


async def test_export_unknown_version_returns_404(auth_client_factory, db_session) -> None:
    """导出指定版本不存在 → 404（沿用归属失败统一 404 惯例，不区分归属/版本缺失）。"""
    client = await auth_client_factory("13700000015")
    session_id = await _create_session(client)
    user_id = await _me_id(client)
    artifact = await _make_published_artifact(
        db_session, user_id, session_id, payload=build_brand_dict()
    )

    resp = await client.get(
        f"/api/v1/agent/artifacts/{artifact.id}/export", params={"version": 99}
    )
    assert resp.status_code == 404


async def test_export_unsupported_type_with_explicit_version_still_conflicts(
    auth_client_factory, db_session
) -> None:
    """显式版本存在但类型不支持导出 → 仍为 409 ARTIFACT_EXPORT_UNSUPPORTED。"""
    client = await auth_client_factory("13700000016")
    session_id = await _create_session(client)
    user_id = await _me_id(client)
    unsupported = await _make_published_artifact(
        db_session,
        user_id,
        session_id,
        module="kol",
        artifact_type="kol_detail_v2",
        artifact_key="kol-detail:x:v",
    )

    resp = await client.get(
        f"/api/v1/agent/artifacts/{unsupported.id}/export", params={"version": 1}
    )
    assert resp.status_code == 409
    assert resp.json()["detail"] == "ARTIFACT_EXPORT_UNSUPPORTED"


async def test_export_unsupported_type_conflicts(auth_client_factory, db_session) -> None:
    client = await auth_client_factory("13700000005")
    session_id = await _create_session(client)
    user_id = await _me_id(client)

    unsupported = await _make_published_artifact(
        db_session,
        user_id,
        session_id,
        module="kol",
        artifact_type="kol_detail_v2",
        artifact_key="kol-detail:x:k",
    )
    resp = await client.get(f"/api/v1/agent/artifacts/{unsupported.id}/export")
    assert resp.status_code == 409
    assert resp.json()["detail"] == "ARTIFACT_EXPORT_UNSUPPORTED"


async def test_export_draft_without_version_conflicts(auth_client_factory, db_session) -> None:
    client = await auth_client_factory("13700000006")
    session_id = await _create_session(client)
    user_id = await _me_id(client)
    now = utc_now()
    draft_only = AgentArtifact(
        id=str(uuid4()),
        session_id=session_id,
        user_id=user_id,
        module="brand",
        artifact_type="brand_report_v3",
        parent_artifact_id=None,
        artifact_key=f"brand:{uuid4().hex[:8]}",
        status="draft",
        latest_version=0,
        activity_sequence=0,
        created_at=now,
        updated_at=now,
    )
    db_session.add(draft_only)
    await db_session.flush()

    resp = await client.get(f"/api/v1/agent/artifacts/{draft_only.id}/export")
    assert resp.status_code == 409
    assert resp.json()["detail"] == "ARTIFACT_EXPORT_UNSUPPORTED"


async def test_export_invalid_published_payload_conflicts_not_500(
    auth_client_factory, db_session
) -> None:
    """历史/旁路非法 payload（强类型 ValidationError）→ 稳定 409，不泄漏 500。"""
    client = await auth_client_factory("13700000007")
    session_id = await _create_session(client)
    user_id = await _me_id(client)
    bad_payload = build_brand_dict()
    bad_payload["data"]["overview"] = {"totally": "wrong"}
    artifact = await _make_published_artifact(
        db_session, user_id, session_id, payload=bad_payload
    )

    resp = await client.get(f"/api/v1/agent/artifacts/{artifact.id}/export")
    assert resp.status_code == 409
    assert resp.json()["detail"] == "ARTIFACT_EXPORT_UNSUPPORTED"


# ---------------------------------------------------------------------------
# 归属隔离（无存在泄露 → 404）
# ---------------------------------------------------------------------------


async def test_artifact_ownership_isolation_returns_404(
    auth_client_factory, db_session
) -> None:
    alice = await auth_client_factory("13700000007")
    bob = await auth_client_factory("13700000008")
    session_id = await _create_session(alice)
    alice_id = await _me_id(alice)
    artifact = await _make_published_artifact(
        db_session, alice_id, session_id, payload=build_brand_dict()
    )

    assert (
        await bob.get(f"/api/v1/agent/sessions/{session_id}/artifacts")
    ).status_code == 404
    assert (await bob.get(f"/api/v1/agent/artifacts/{artifact.id}")).status_code == 404
    assert (
        await bob.get(f"/api/v1/agent/artifacts/{artifact.id}/versions/1")
    ).status_code == 404
    assert (
        await bob.get(f"/api/v1/agent/artifacts/{artifact.id}/export")
    ).status_code == 404
    assert (
        await bob.get(
            f"/api/v1/agent/artifacts/{artifact.id}/export", params={"version": 1}
        )
    ).status_code == 404

    # read-state 归属失败 → 404（不泄露会话存在）
    assert (
        await bob.put(
            f"/api/v1/agent/sessions/{session_id}/artifact-read-state",
            json={"module": "brand", "last_seen_sequence": 1},
        )
    ).status_code == 404
    # 水位查询归属失败 → 404（不泄露会话存在与其他用户水位）
    assert (
        await bob.get(f"/api/v1/agent/sessions/{session_id}/artifact-read-states")
    ).status_code == 404


async def test_artifact_under_archived_session_is_404(
    auth_client_factory, db_session
) -> None:
    """§15.2：删除（软删除）Session 后，其下的 Artifact 一律 404，不泄露存在。"""
    alice = await auth_client_factory("13700000009")
    session_id = await _create_session(alice)
    user_id = await _me_id(alice)
    artifact = await _make_published_artifact(
        db_session, user_id, session_id, payload=build_brand_dict()
    )

    assert (await alice.delete(f"/api/v1/agent/sessions/{session_id}")).status_code == 204

    assert (await alice.get(f"/api/v1/agent/artifacts/{artifact.id}")).status_code == 404
    assert (
        await alice.get(f"/api/v1/agent/artifacts/{artifact.id}/versions/1")
    ).status_code == 404
    assert (
        await alice.get(f"/api/v1/agent/artifacts/{artifact.id}/export")
    ).status_code == 404
