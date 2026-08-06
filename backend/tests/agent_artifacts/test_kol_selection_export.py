"""kol_selection_v3 Excel 导出（Gate C Task 3 / 设计 §7.3）。

4 个 Sheet：达人圈选总表 / 达人详细画像（Top20 全详情块）/ 粉丝画像详情 /
评分方法论与数据来源。v3 名单展示效果/价格效率/价值/报价/评级/完整度；
历史 kol_score_v2 快照仍可导出。导出器不调用模型/MCP。
"""

from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.agent_artifacts.exporters import ArtifactExportUnsupported, export_artifact

from tests.agent_artifacts.test_payloads import (
    build_insight_dict,
    build_kol_analysis_dict,
    build_kol_detail_dict,
    build_kol_selection_dict,
    build_kol_value_selection_dict,
)

KOL_SHEETS = (
    "达人圈选总表",
    "达人详细画像",
    "粉丝画像详情",
    "评分方法论与数据来源",
)


class _Version:
    """轻量 Version 桩：模拟已发布 AgentArtifactVersion 的读取面。"""

    def __init__(self, schema_version: str, payload_json: dict | None) -> None:
        self.schema_version = schema_version
        self.payload_json = payload_json
        self.data_status = "complete" if payload_json else "draft"


def _values(ws) -> list:
    return [
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]


def _count_detail_blocks(ws) -> int:
    """统计「达人详细画像」中 #N 详情块标题行数。"""
    return sum(
        1
        for row in ws.iter_rows(min_col=1, max_col=1)
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("#")
    )


def _kol_version(payload_factory=build_kol_value_selection_dict) -> _Version:
    return _Version("kol_selection_v3", payload_factory())


def _kol_version_with_20_items() -> _Version:
    base = build_kol_value_selection_dict()
    items = base["data"]["items"]
    while len(items) < 20:
        clone = dict(items[0])
        clone["rank"] = len(items) + 1
        clone["kol_uid"] = f"k{len(items) + 1}"
        clone["nickname"] = f"达人{len(items) + 1}"
        items.append(clone)
    return _Version("kol_selection_v3", base)


def test_kol_v3_export_has_four_sheets_and_v3_headers() -> None:
    content = export_artifact(_kol_version())
    assert content[:2] == b"PK"
    wb = load_workbook(BytesIO(content), data_only=False)
    assert tuple(wb.sheetnames) == KOL_SHEETS

    summary = wb["达人圈选总表"]
    values = _values(summary)
    for header in ("效果分", "价格效率分", "价值总分", "报价", "价格样本数", "数据完整度"):
        assert header in values
    assert "某品牌" in str(summary["A1"].value)


def test_kol_export_contains_twenty_detail_blocks() -> None:
    content = export_artifact(_kol_version_with_20_items())
    wb = load_workbook(BytesIO(content), data_only=False)
    detail = wb["达人详细画像"]
    assert _count_detail_blocks(detail) == 20
    detail_values = _values(detail)
    assert "价值总分" in detail_values
    assert "价格样本数" in detail_values


def test_kol_v2_history_export_remains_compatible() -> None:
    """历史 kol_score_v2 快照仍可导出：总表展示综合分/星级。"""
    content = export_artifact(_kol_version(build_kol_selection_dict))
    wb = load_workbook(BytesIO(content), data_only=False)
    summary = wb["达人圈选总表"]
    values = _values(summary)
    assert "综合分" in values
    assert "星级" in values
    assert "效果分" not in values
    assert _count_detail_blocks(wb["达人详细画像"]) >= 1


def test_kol_methodology_sheet_records_v3_weights() -> None:
    content = export_artifact(_kol_version())
    wb = load_workbook(BytesIO(content), data_only=False)
    method = wb["评分方法论与数据来源"]
    values = _values(method)
    assert any("kol_value_score_v3" in str(v) for v in values)
    assert "价格效率" in values
    assert 30 in values


def test_kol_external_text_is_formula_escaped() -> None:
    payload = build_kol_value_selection_dict()
    payload["data"]["items"][0]["nickname"] = "=1+1"
    content = export_artifact(_Version("kol_selection_v3", payload))
    wb = load_workbook(BytesIO(content))
    values = _values(wb["达人圈选总表"])
    assert any(isinstance(v, str) and v.startswith("'") for v in values)


def test_draft_kol_selection_raises_unsupported() -> None:
    with pytest.raises(ArtifactExportUnsupported) as excinfo:
        export_artifact(_Version("kol_selection_v3", None))
    assert excinfo.value.code == "ARTIFACT_EXPORT_UNSUPPORTED"


def test_invalid_published_kol_payload_raises_unsupported() -> None:
    with pytest.raises(ArtifactExportUnsupported) as excinfo:
        export_artifact(_Version("kol_selection_v3", {"schema_version": "kol_selection_v3"}))
    assert excinfo.value.code == "ARTIFACT_EXPORT_UNSUPPORTED"


def test_other_artifact_types_raise_unsupported() -> None:
    for schema, factory in (
        ("kol_analysis_v2", build_kol_analysis_dict),
        ("kol_detail_v2", build_kol_detail_dict),
        ("insight_board_v1", build_insight_dict),
    ):
        payload = factory()
        with pytest.raises(ArtifactExportUnsupported) as excinfo:
            export_artifact(_Version(schema, payload))
        assert excinfo.value.code == "ARTIFACT_EXPORT_UNSUPPORTED"
