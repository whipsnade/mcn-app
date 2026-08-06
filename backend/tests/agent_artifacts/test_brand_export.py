"""brand_report_v3 Excel 导出（Gate C Task 3 / 设计 §12.1 消费边界）。

导出器只读已发布不可变 Version 的 payload，不调用模型/MCP；brand_report_v3
之外的类型（含 draft）抛 ``ArtifactExportUnsupported``（路由映射 409）。
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.agent_artifacts.exporters import ArtifactExportUnsupported, export_artifact

from tests.agent_artifacts.test_payloads import build_brand_dict

BRAND_SHEETS = (
    "综合概览",
    "情感分析",
    "日趋势",
    "内容类型与达人",
    "地域分布",
    "热门帖子TOP",
    "舆情洞察",
    "方法论",
)


class _Version:
    """轻量 Version 桩：模拟已发布 AgentArtifactVersion 的读取面。"""

    def __init__(self, schema_version: str, payload_json: dict | None) -> None:
        self.schema_version = schema_version
        self.payload_json = payload_json
        self.data_status = "complete" if payload_json else "draft"


def _brand_version(**overrides) -> _Version:
    payload = build_brand_dict()
    payload.update(overrides)
    return _Version("brand_report_v3", payload)


def _values(ws) -> list:
    return [
        cell.value
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    ]


def _brand_version_multi() -> _Version:
    """多点位 fixture：3 个日趋势点 + 2 个地域，覆盖图表区间计算。"""
    payload = build_brand_dict()
    payload["data"]["daily_trend"] = [
        {
            "date": date(2026, 1, day),
            "platform": "xiaohongshu",
            "volume": day * 100,
            "engagement": day * 500,
            "positive": day * 70,
            "neutral": 20,
            "negative": 10,
        }
        for day in (1, 2, 3)
    ]
    payload["data"]["regions"] = [
        {"region": "上海", "volume": 300, "share": 0.3, "sentiment_score": 0.7},
        {"region": "北京", "volume": 700, "share": 0.7, "sentiment_score": 0.6},
    ]
    return _Version("brand_report_v3", payload)


def test_published_brand_report_exports_valid_workbook() -> None:
    """已发布 brand_report_v3 导出为可重载 .xlsx，八章节/关键单元格齐全。"""
    content = export_artifact(_brand_version())
    assert content[:2] == b"PK"
    wb = load_workbook(BytesIO(content), data_only=False)
    assert tuple(wb.sheetnames) == BRAND_SHEETS

    overview = wb["综合概览"]
    assert "某品牌" in str(overview["A1"].value)
    assert 1000 in _values(overview)  # total_volume
    assert 5000 in _values(overview)  # total_engagement

    sentiment = wb["情感分析"]
    assert 10 in _values(sentiment)  # 每个情感桶计数
    assert 0.5 in _values(sentiment)  # 占比

    trend = wb["日趋势"]
    assert 100 in _values(trend)  # volume
    assert len(trend._charts) == 1

    content_sheet = wb["内容类型与达人"]
    assert "图文" in _values(content_sheet)
    insights = wb["舆情洞察"]
    assert "咖啡" in _values(insights)

    top_posts = wb["热门帖子TOP"]
    assert "热帖" in _values(top_posts)
    assert "author" in _values(top_posts)
    assert 115 in _values(top_posts)

    methodology = wb["方法论"]
    assert "DataTap" in _values(methodology)


def test_region_chart_covers_region_values() -> None:
    """地域分布图表由导出器现场重建，数值区间覆盖真实行。"""
    content = export_artifact(_brand_version_multi())
    wb = load_workbook(BytesIO(content), data_only=False)
    sheet = wb["地域分布"]
    assert len(sheet._charts) == 1
    chart = sheet._charts[0]
    assert chart.ser[0].val.numRef.f == "'地域分布'!$B$5:$B$6"
    assert chart.ser[0].cat.numRef.f == "'地域分布'!$A$5:$A$6"
    assert sheet["B5"].value == 300
    assert sheet["B6"].value == 700


def test_daily_trend_chart_rebuilt() -> None:
    """日趋势折线图现场重建，声量/互动两系列引用真实区间。"""
    content = export_artifact(_brand_version_multi())
    wb = load_workbook(BytesIO(content), data_only=False)
    sheet = wb["日趋势"]
    assert len(sheet._charts) == 1
    chart = sheet._charts[0]
    refs = [ser.val.numRef.f for ser in chart.ser]
    assert "'日趋势'!$B$4:$B$6" in refs  # 声量（表头@3 作系列名）
    assert "'日趋势'!$C$4:$C$6" in refs  # 互动


def test_daily_trend_platform_cell_is_escaped() -> None:
    """平台名等第三方可控文本以 = 开头时前缀 ' 转义，防公式注入。"""
    payload = build_brand_dict()
    payload["data"]["daily_trend"] = [
        {
            "date": date(2026, 1, 1),
            "platform": "=1+1",
            "volume": 10,
            "engagement": 20,
            "positive": 1,
            "neutral": 1,
            "negative": 1,
        }
    ]
    content = export_artifact(_Version("brand_report_v3", payload))
    wb = load_workbook(BytesIO(content))
    trend = wb["日趋势"]
    cells = [c.value for row in trend.iter_rows() for c in row if c.value is not None]
    assert any(isinstance(v, str) and v.startswith("'") for v in cells)


def test_draft_brand_report_raises_unsupported() -> None:
    with pytest.raises(ArtifactExportUnsupported) as excinfo:
        export_artifact(_Version("brand_report_v3", None))
    assert excinfo.value.code == "ARTIFACT_EXPORT_UNSUPPORTED"


def test_invalid_published_payload_raises_unsupported() -> None:
    with pytest.raises(ArtifactExportUnsupported) as excinfo:
        export_artifact(_Version("brand_report_v3", {"schema_version": "brand_report_v3"}))
    assert excinfo.value.code == "ARTIFACT_EXPORT_UNSUPPORTED"
    assert excinfo.value.schema_version == "brand_report_v3"


def test_restricted_brand_discloses_restricted_sections() -> None:
    """受限章节：保留表头、写「数据受限」说明、不建误导性图表。"""
    payload = build_brand_dict()
    payload["data_status"] = "restricted"
    payload["data"]["daily_trend"] = []
    payload["availability"]["daily_trend"] = {
        "status": "partial",
        "reason_codes": ["source_unavailable"],
    }
    payload["limitations"] = [
        {
            "code": "source_unavailable",
            "message": "数据源暂不可用",
            "affected_paths": ["data.daily_trend"],
        }
    ]
    content = export_artifact(_Version("brand_report_v3", payload))
    wb = load_workbook(BytesIO(content))
    trend = wb["日趋势"]
    assert "数据受限" in str(_values(trend))
    assert "数据源暂不可用" in str(_values(trend))
    assert trend._charts == []  # 空章节不建图
    assert 1000 in _values(wb["综合概览"])


# ---------------------------------------------------------------------------
# Gate C 审核修复：模板无示例残留 / 图表真实引用 / used range
# ---------------------------------------------------------------------------


def test_no_fixed_sample_values_in_exported_workbook() -> None:
    """来源模板的固定品牌/日期/「小红书 TOP15」标题不得出现在最终文件。"""
    content = export_artifact(_brand_version())
    wb = load_workbook(BytesIO(content), data_only=False)
    blob = "\n".join(
        str(cell.value)
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.value is not None
    )
    assert "昊来了" not in blob
    assert "2026-06-30" not in blob
    assert "TOP 15" not in blob
    assert "某品牌" in blob  # 动态标题存在


def test_daily_trend_series_names_and_date_text() -> None:
    """趋势图系列名来自表头（非 Series N）；日期轴为文本而非 Excel 序列号。"""
    content = export_artifact(_brand_version_multi())
    wb = load_workbook(BytesIO(content), data_only=False)
    sheet = wb["日趋势"]
    assert sheet["B3"].value == "声量"
    assert sheet["C3"].value == "互动数"
    assert sheet["A4"].value == "2026-01-01"  # 日期文本
    assert len(sheet._charts) == 1
    chart = sheet._charts[0]
    refs = [ser.val.numRef.f for ser in chart.ser]
    assert "'日趋势'!$B$4:$B$6" in refs
    assert "'日趋势'!$C$4:$C$6" in refs


def test_used_range_not_extended_to_row_1000() -> None:
    content = export_artifact(_brand_version_multi())
    wb = load_workbook(BytesIO(content), data_only=False)
    for ws in wb.worksheets:
        assert ws.max_row < 100, f"{ws.title} max_row={ws.max_row}"


# ---------------------------------------------------------------------------
# Gate C 复审：品牌 6 列达人分层表 / 概览固定标签
# ---------------------------------------------------------------------------


def test_creator_tiers_table_uses_six_columns() -> None:
    """达人分层表必须 6 列（平台/层级/达人数量/发帖数/声量/互动数），标题合并跨 6 列。"""
    payload = build_brand_dict()
    payload["data"]["creator_tiers"] = [
        {"platform": "xiaohongshu", "tier": "头部", "creator_count": 3, "posts": 30, "volume": 600, "engagement": 3000},
        {"platform": "xiaohongshu", "tier": "腰部", "creator_count": 5, "posts": 50, "volume": 400, "engagement": 2000},
    ]
    content = export_artifact(_Version("brand_report_v3", payload))
    wb = load_workbook(BytesIO(content), data_only=False)
    ws = wb["内容类型与达人"]
    values = _values(ws)
    for header in ("平台", "层级", "达人数量", "发帖数", "声量", "互动数"):
        assert header in values, header
    # 达人分层标题合并必须跨到第 6 列（互动数），不得只跨 4 列。
    tier_rows = [cell.row for row in ws.iter_rows() for cell in row if cell.value == "达人分层"]
    assert tier_rows
    tier_row = tier_rows[0]
    merges = [r for r in ws.merged_cells.ranges if r.min_row == tier_row]
    assert merges and any(r.max_col >= 6 for r in merges), [str(r) for r in merges]
    # 第 6 列互动数数据真实落位。
    assert 3000 in values and 2000 in values


def test_overview_has_fixed_metric_and_total_labels() -> None:
    """概览指标表固定标签：A6=指标（行名列）、E6=总计（合计列）。"""
    content = export_artifact(_brand_version())
    wb = load_workbook(BytesIO(content), data_only=False)
    overview = wb["综合概览"]
    assert overview["A6"].value == "指标"
    assert overview["E6"].value == "总计"
