"""安全上传与 upload Evidence 测试（Gate B Task 3）。

覆盖：CSV/XLSX 上传解析为不可变 upload Evidence（XOR：tool_call_id 为
NULL、upload_id 有值）；xlsm/未知扩展名 415；超 20 MiB 413；超 50,000
数据行 400 + error_code；跨用户查询统一 404；GET 归属读取。
"""

from __future__ import annotations

import io
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from uuid import uuid4

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from sqlalchemy import select

from app.agent_runtime.models import AgentUpload, EvidenceItem
from app.agent_runtime.uploads import UploadRejectedError, _parse_xlsx
from app.db.session import get_db
from app.main import create_app

MAX_BYTES = 20 * 1024 * 1024
MAX_ROWS = 50000


def _now() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def _csv_bytes(rows: list[list[object]]) -> bytes:
    return ("\n".join(",".join(str(cell) for cell in row) for row in rows) + "\n").encode(
        "utf-8"
    )


def _xlsx_bytes() -> bytes:
    """生成最小合法 xlsx（openpyxl 输出到内存）。"""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["平台", "声量"])
    sheet.append(["小红书", 100])
    sheet.append(["抖音", 200])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _with_rels(base: bytes, rels: dict[str, bytes]) -> bytes:
    """在合法 xlsx 基础上新增/覆盖指定 ZIP 部件（用于构造 .rels 测试夹具）。"""
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(base)) as zin, zipfile.ZipFile(
        out, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for info in zin.infolist():
            zout.writestr(info, zin.read(info.filename))
        for name, content in rels.items():
            zout.writestr(name, content)
    return out.getvalue()


def _repack_with_override(base: bytes, path: str, content: bytes) -> bytes:
    """把 zip 中某个条目替换为 content（同名覆盖）。"""
    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(base)) as zin, zipfile.ZipFile(
        out, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for info in zin.infolist():
            if info.filename == path:
                continue
            zout.writestr(info, zin.read(info.filename))
        zout.writestr(path, content)
    return out.getvalue()


# OOXML relationships 命名空间与合法内部 drawing 关系。
_RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_INTERNAL_DRAWING_RELS = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
    f'<Relationships xmlns="{_RELS_NS}">'
    '<Relationship Id="rId1" '
    'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
    'Target="../drawings/drawing1.xml"/>'
    "</Relationships>"
).encode("utf-8")


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[3]


def _custom_xlsx(sheet_xml: str) -> bytes:
    """构造最小合法 xlsx（自定义 worksheet XML），用于物理维度攻击夹具。"""
    cts = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    wbrels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        "</Relationships>"
    )
    wb = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    out = io.BytesIO()
    with zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        zout.writestr("[Content_Types].xml", cts)
        zout.writestr("_rels/.rels", rels)
        zout.writestr("xl/_rels/workbook.xml.rels", wbrels)
        zout.writestr("xl/workbook.xml", wb)
        zout.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return out.getvalue()


@pytest_asyncio.fixture
async def upload_client(db_session):
    """带鉴权头的测试客户端；上传服务注入临时目录。"""
    import tempfile

    from app.agent_runtime.router import get_upload_service

    temp_dir = tempfile.mkdtemp(prefix="agent-uploads-test-")
    clients: list[AsyncClient] = []

    async def _make(phone: str):
        app = create_app()

        async def _override_db():
            yield db_session

        async def _override_upload_service():
            from app.agent_runtime.uploads import UploadService

            yield UploadService(
                db_session,
                storage_dir=temp_dir,
                max_bytes=MAX_BYTES,
                max_rows=MAX_ROWS,
            )

        app.dependency_overrides[get_db] = _override_db
        app.dependency_overrides[get_upload_service] = _override_upload_service
        tc = AsyncClient(transport=ASGITransport(app=app), base_url="http://test")
        login = await tc.post(
            "/api/v1/auth/mock/sms/login",
            json={"phone": phone, "code": "000000"},
        )
        assert login.status_code == 200, login.text
        tc.headers["Authorization"] = f"Bearer {login.json()['access_token']}"
        clients.append(tc)
        return tc, app

    yield _make
    for tc in clients:
        await tc.aclose()


async def _create_session(client: AsyncClient) -> str:
    resp = await client.post("/api/v1/agent/sessions", json={})
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


async def _seed_upload(db_session, user_id: str, session_id: str, **overrides) -> object:
    from app.agent_runtime.models import AgentUpload

    upload = AgentUpload(
        id=str(uuid4()),
        user_id=user_id,
        session_id=session_id,
        original_filename="投放数据.csv",
        mime_type="text/csv",
        size_bytes=100,
        sha256="a" * 64,
        storage_key=f"{user_id}/{uuid4()}.csv",
        status="parsed",
        created_at=_now(),
        **overrides,
    )
    db_session.add(upload)
    await db_session.flush()
    return upload


async def _load_upload_evidence(db_session, upload_id: str) -> EvidenceItem | None:
    return await db_session.scalar(
        select(EvidenceItem).where(EvidenceItem.upload_id == upload_id)
    )


async def test_upload_csv_creates_upload_evidence(upload_client, db_session) -> None:
    client, _app = await upload_client("13800000001")
    session_id = await _create_session(client)
    csv_bytes = _csv_bytes([["平台", "声量"], ["小红书", 100], ["抖音", 200]])

    response = await client.post(
        f"/api/v1/agent/sessions/{session_id}/uploads",
        files={"file": ("投放数据.csv", csv_bytes, "text/csv")},
    )
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "parsed"
    assert body["original_filename"] == "投放数据.csv"
    assert body["mime_type"] == "text/csv"
    assert len(body["sha256"]) == 64

    evidence = await _load_upload_evidence(db_session, body["id"])
    assert evidence is not None
    assert evidence.source_type == "user_upload"
    assert evidence.tool_call_id is None
    assert evidence.upload_id == body["id"]
    assert evidence.raw_payload_json["rows"][0] == {"平台": "小红书", "声量": "100"}


async def test_upload_xlsx_creates_upload_evidence(upload_client, db_session) -> None:
    client, _app = await upload_client("13800000002")
    session_id = await _create_session(client)

    response = await client.post(
        f"/api/v1/agent/sessions/{session_id}/uploads",
        files={
            "file": (
                "投放数据.xlsx",
                _xlsx_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 201, response.text
    body = response.json()
    assert body["status"] == "parsed"
    evidence = await _load_upload_evidence(db_session, body["id"])
    assert evidence is not None
    assert evidence.source_type == "user_upload"
    assert evidence.tool_call_id is None


async def test_macro_workbook_is_rejected(upload_client, db_session) -> None:
    client, _app = await upload_client("13800000003")
    session_id = await _create_session(client)

    response = await client.post(
        f"/api/v1/agent/sessions/{session_id}/uploads",
        files={"file": ("投放数据.xlsm", b"PK\x03\x04fake", "application/vnd.ms-excel.sheet.macroEnabled.12")},
    )
    assert response.status_code == 415, response.text


async def test_upload_unknown_extension_is_rejected(upload_client, db_session) -> None:
    client, _app = await upload_client("13800000004")
    session_id = await _create_session(client)

    response = await client.post(
        f"/api/v1/agent/sessions/{session_id}/uploads",
        files={"file": ("notes.txt", b"hello", "text/plain")},
    )
    assert response.status_code == 415, response.text


async def test_upload_too_large_is_rejected(upload_client, db_session) -> None:
    client, _app = await upload_client("13800000005")
    session_id = await _create_session(client)

    response = await client.post(
        f"/api/v1/agent/sessions/{session_id}/uploads",
        files={"file": ("big.csv", b"x" * (MAX_BYTES + 1), "text/csv")},
    )
    assert response.status_code == 413, response.text


async def test_upload_over_rows_limit_is_failed(upload_client, db_session) -> None:
    client, _app = await upload_client("13800000006")
    session_id = await _create_session(client)
    header = ["平台", "声量"]
    rows = [header] + [["平台", str(index)] for index in range(MAX_ROWS + 1)]
    csv_bytes = _csv_bytes(rows)

    response = await client.post(
        f"/api/v1/agent/sessions/{session_id}/uploads",
        files={"file": ("超大.csv", csv_bytes, "text/csv")},
    )
    assert response.status_code == 400, response.text
    assert response.json()["detail"] == "rows_exceeded"


async def test_upload_cross_user_is_404(upload_client, db_session) -> None:
    client_a, _app_a = await upload_client("13800000007")
    client_b, _app_b = await upload_client("13800000008")
    session_a = await _create_session(client_a)
    session_b = await _create_session(client_b)
    csv_bytes = _csv_bytes([["平台", "声量"], ["小红书", 100]])

    created = await client_a.post(
        f"/api/v1/agent/sessions/{session_a}/uploads",
        files={"file": ("投放数据.csv", csv_bytes, "text/csv")},
    )
    assert created.status_code == 201, created.text
    upload_id = created.json()["id"]

    # B 用户读取 A 的上传 → 404；B 用户在自己的 Session 上传 → 正常。
    fetched = await client_b.get(f"/api/v1/agent/uploads/{upload_id}")
    assert fetched.status_code == 404, fetched.text
    own = await client_b.post(
        f"/api/v1/agent/sessions/{session_b}/uploads",
        files={"file": ("我的.csv", csv_bytes, "text/csv")},
    )
    assert own.status_code == 201, own.text


async def test_get_upload_owned(upload_client, db_session) -> None:
    client, _app = await upload_client("13800000009")
    session_id = await _create_session(client)
    csv_bytes = _csv_bytes([["平台", "声量"], ["小红书", 100]])

    created = await client.post(
        f"/api/v1/agent/sessions/{session_id}/uploads",
        files={"file": ("投放数据.csv", csv_bytes, "text/csv")},
    )
    assert created.status_code == 201, created.text
    upload_id = created.json()["id"]

    fetched = await client.get(f"/api/v1/agent/uploads/{upload_id}")
    assert fetched.status_code == 200, fetched.text
    assert fetched.json()["id"] == upload_id
    assert fetched.json()["status"] == "parsed"


# ---------------------------------------------------------------------------
# P1-1: XLSX 合法内部关系不被误判（XML 解析 TargetMode=external 才拒绝）
# ---------------------------------------------------------------------------


def test_p1_brand_report_xlsx_is_accepted() -> None:
    """仓库品牌模板含 charts/drawings 内部关系，必须能正常解析。"""
    columns, rows = _parse_xlsx(
        (_repo_root() / "brand_report.xlsx").read_bytes(), max_rows=10000
    )
    assert isinstance(columns, list)
    assert isinstance(rows, list)
    assert rows


def test_p1_internal_drawing_relationship_is_accepted() -> None:
    """worksheet→drawing 是合法内部关系（TargetMode 缺失），不得拒绝。"""
    content = _with_rels(
        _xlsx_bytes(),
        {"xl/worksheets/_rels/sheet1.xml.rels": _INTERNAL_DRAWING_RELS},
    )
    columns, rows = _parse_xlsx(content, max_rows=100)
    assert rows


@pytest.mark.parametrize(
    "rels_xml",
    [
        # 双引号 external
        (
            '<?xml version="1.0"?>'
            f'<Relationships xmlns="{_RELS_NS}">'
            '<Relationship Id="r1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink" '
            'Target="https://example.com" TargetMode="External"/>'
            "</Relationships>"
        ),
        # 单引号 external（字符串搜索会被绕过）
        (
            "<?xml version=\"1.0\"?>"
            f"<Relationships xmlns=\"{_RELS_NS}\">"
            "<Relationship Id=\"r1\" "
            "Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink\" "
            "Target=\"https://example.com\" TargetMode='External'/>"
            "</Relationships>"
        ),
        # 属性顺序变化（TargetMode 后置）仍被拒绝
        (
            "<?xml version=\"1.0\"?>"
            f"<Relationships xmlns=\"{_RELS_NS}\">"
            "<Relationship Id=\"r1\" Target=\"https://example.com\" "
            "Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/hyperlink\" "
            "TargetMode=\"external\"/>"
            "</Relationships>"
        ),
    ],
)
def test_p1_external_relationship_is_rejected(rels_xml: str) -> None:
    content = _with_rels(
        _xlsx_bytes(),
        {"xl/worksheets/_rels/sheet1.xml.rels": rels_xml.encode("utf-8")},
    )
    with pytest.raises(UploadRejectedError) as exc_info:
        _parse_xlsx(content, max_rows=100)
    assert exc_info.value.error_code == "external_links_detected"


def test_p1_xl_external_links_directory_is_rejected() -> None:
    """xl/externalLinks/ 目录存在即拒绝（外部引用部件）。"""
    content = _with_rels(
        _xlsx_bytes(),
        {"xl/externalLinks/externalLink1.xml": b"<externalLink/>"},
    )
    with pytest.raises(UploadRejectedError) as exc_info:
        _parse_xlsx(content, max_rows=100)
    assert exc_info.value.error_code == "external_links_detected"


def test_p1_macro_part_inside_xlsx_is_rejected() -> None:
    """xlsx 内嵌 vbaProject 宏部件（即使扩展名不是 .xlsm）也必须拒绝。"""
    content = _with_rels(_xlsx_bytes(), {"xl/vbaProject.bin": b"\x01\x02VBA"})
    with pytest.raises(UploadRejectedError) as exc_info:
        _parse_xlsx(content, max_rows=100)
    assert exc_info.value.error_code == "macro_detected"


def test_p1_invalid_relationships_xml_is_rejected() -> None:
    """.rels 不是合法 XML 时按 invalid_relationships_xml 拒绝，不能忽略。"""
    content = _with_rels(
        _xlsx_bytes(),
        {"xl/worksheets/_rels/sheet1.xml.rels": b"<Relationships><unclosed"},
    )
    with pytest.raises(UploadRejectedError) as exc_info:
        _parse_xlsx(content, max_rows=100)
    assert exc_info.value.error_code == "invalid_relationships_xml"


async def test_p1_external_relationship_error_code_persisted(
    upload_client, db_session
) -> None:
    """拒绝的外部引用 xlsx 上传后，失败记录的 error_code 正确持久化。"""
    client, _app = await upload_client("13800000999")
    session_id = await _create_session(client)
    rels_xml = (
        '<?xml version="1.0"?>'
        f'<Relationships xmlns="{_RELS_NS}">'
        '<Relationship Id="r1" Target="https://example.com" TargetMode="External"/>'
        "</Relationships>"
    )
    content = _with_rels(
        _xlsx_bytes(),
        {"xl/worksheets/_rels/sheet1.xml.rels": rels_xml.encode("utf-8")},
    )
    response = await client.post(
        f"/api/v1/agent/sessions/{session_id}/uploads",
        files={
            "file": (
                "恶意.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 415, response.text
    assert response.json()["detail"] == "external_links_detected"
    upload = await db_session.scalar(
        select(AgentUpload)
        .where(AgentUpload.session_id == session_id)
        .order_by(AgentUpload.created_at.desc())
    )
    assert upload is not None
    assert upload.status == "failed"
    assert upload.error_code == "external_links_detected"


# ---------------------------------------------------------------------------
# P1-4: 工作表物理维度防护（超大 dimension / 极大 row r / 超大列坐标 / 行边界）
# ---------------------------------------------------------------------------

_WORKSHEET_NS = "xmlns=\"http://schemas.openxmlformats.org/spreadsheetml/2006/main\""


def test_p1_huge_dimension_rejected_quickly() -> None:
    """声明 A1:XFD1048576 的几 KB XLSX 被快速拒绝。"""
    sheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet {_WORKSHEET_NS}><dimension ref="A1:XFD1048576"/><sheetData/></worksheet>'
    )
    with pytest.raises(UploadRejectedError) as exc_info:
        _parse_xlsx(_custom_xlsx(sheet), max_rows=50000)
    assert exc_info.value.error_code == "worksheet_dimensions_exceeded"


def test_p1_huge_row_r_rejected() -> None:
    """第二个 row 的 r 设为极大值（配合超大 dimension 触发）被拒绝。"""
    sheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet {_WORKSHEET_NS}><dimension ref="A1:A1048576"/><sheetData>'
        '<row r="1"><c r="A1" t="inlineStr"><is><t>平台</t></is></c></row>'
        '<row r="1048576"><c r="A1048576" t="inlineStr"><is><t>x</t></is></c></row>'
        "</sheetData></worksheet>"
    )
    with pytest.raises(UploadRejectedError) as exc_info:
        _parse_xlsx(_custom_xlsx(sheet), max_rows=50000)
    assert exc_info.value.error_code == "worksheet_dimensions_exceeded"


def test_p1_huge_column_coordinate_rejected() -> None:
    """超大列坐标（XFD）被拒绝。"""
    sheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet {_WORKSHEET_NS}><dimension ref="A1:XFD10"/><sheetData/></worksheet>'
    )
    with pytest.raises(UploadRejectedError) as exc_info:
        _parse_xlsx(_custom_xlsx(sheet), max_rows=10)
    assert exc_info.value.error_code == "worksheet_columns_exceeded"


def test_p1_row_count_boundary() -> None:
    """max_rows=5 时：声明 6 行（含表头）通过，7 行拒绝。"""
    ok_sheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet {_WORKSHEET_NS}><dimension ref="A1:A6"/><sheetData/></worksheet>'
    )
    columns, rows = _parse_xlsx(_custom_xlsx(ok_sheet), max_rows=5)
    assert rows == []

    over_sheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet {_WORKSHEET_NS}><dimension ref="A1:A7"/><sheetData/></worksheet>'
    )
    with pytest.raises(UploadRejectedError) as exc_info:
        _parse_xlsx(_custom_xlsx(over_sheet), max_rows=5)
    assert exc_info.value.error_code == "worksheet_dimensions_exceeded"


def test_p1_brand_template_still_passes() -> None:
    """合法品牌模板继续通过（物理维度在允许范围内）。"""
    columns, rows = _parse_xlsx(
        (_repo_root() / "brand_report.xlsx").read_bytes(), max_rows=10000
    )
    assert rows


async def test_p1_worksheet_dimension_error_code_persisted(upload_client, db_session) -> None:
    """超大 dimension 上传经 UploadRejectedError 路径：失败记录 error_code 持久化。"""
    client, _app = await upload_client("13800000888")
    session_id = await _create_session(client)
    sheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet {_WORKSHEET_NS}><dimension ref="A1:XFD1048576"/><sheetData/></worksheet>'
    )
    response = await client.post(
        f"/api/v1/agent/sessions/{session_id}/uploads",
        files={
            "file": (
                "恶意.xlsx",
                _custom_xlsx(sheet),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 400, response.text
    assert response.json()["detail"] == "worksheet_dimensions_exceeded"
    upload = await db_session.scalar(
        select(AgentUpload)
        .where(AgentUpload.session_id == session_id)
        .order_by(AgentUpload.created_at.desc())
    )
    assert upload is not None
    assert upload.status == "failed"
    assert upload.error_code == "worksheet_dimensions_exceeded"


# ---------------------------------------------------------------------------
# P2: 窄表按实际工作表宽度迭代（不再固定 1000 列）
# ---------------------------------------------------------------------------


def test_p2_two_column_narrow_table_parses_two_columns() -> None:
    """两列窄表解析结果仍为两列（不因 max_col 上限膨胀列名）。"""
    columns, rows = _parse_xlsx(_xlsx_bytes(), max_rows=50000)
    assert len(columns) == 2
    assert columns == ["平台", "声量"]
    assert len(rows) == 2


def test_p2_iter_rows_uses_effective_width(monkeypatch) -> None:
    """iter_rows 的 max_col 使用实际工作表宽度（2 列窄表按 2 迭代，而非 1000）。"""
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet

    captured: list[int | None] = []
    original = ReadOnlyWorksheet.iter_rows

    def _spy(self, *args, **kwargs):
        captured.append(kwargs.get("max_col"))
        return original(self, *args, **kwargs)

    monkeypatch.setattr(ReadOnlyWorksheet, "iter_rows", _spy)
    columns, rows = _parse_xlsx(_xlsx_bytes(), max_rows=50000)
    assert captured
    assert all(max_col == 2 for max_col in captured)
    assert len(columns) == 2
    assert len(rows) == 2


def test_p2_many_row_narrow_table_parses_all_rows() -> None:
    """多行窄表（2000 行 × 2 列）解析全部数据行，不被误拒绝。"""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["平台", "声量"])
    for i in range(2000):
        sheet.append([f"平台{i}", i])
    buffer = io.BytesIO()
    workbook.save(buffer)
    columns, rows = _parse_xlsx(buffer.getvalue(), max_rows=50000)
    assert len(columns) == 2
    assert len(rows) == 2000
    assert rows[0] == {"平台": "平台0", "声量": 0}
    assert rows[-1] == {"平台": "平台1999", "声量": 1999}


def test_p2_no_dimension_huge_row_r_bounded(monkeypatch) -> None:
    """dimension 缺失、row r 极大：迭代有界、不无限循环、不抛 CPU 异常。"""
    sheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet {_WORKSHEET_NS}><sheetData>'
        '<row r="1"><c r="A1" t="inlineStr"><is><t>平台</t></is></c></row>'
        '<row r="1048576"><c r="A1048576" t="inlineStr"><is><t>x</t></is></c></row>'
        "</sheetData></worksheet>"
    )
    columns, rows = _parse_xlsx(_custom_xlsx(sheet), max_rows=10)
    assert len(rows) <= 10  # 有界，不迭代到 1048576 行
