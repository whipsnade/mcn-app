"""Pi 自主营销行为的离线集成契约测试。

这些测试不启动真实模型、DataTap 或钱包服务，只锁定生产边界上的可观察语义：
未知 MCP 结果不能机械重放，但模型仍可继续调用独立参数；长尾明细经通用
analysis_report_v1 投影为完整 Workbook，而不是切换到某个业务专用导出器。
"""

from __future__ import annotations

from typing import Any, Mapping

from app.agent_runtime.tools.contracts import ToolContext
from app.agent_runtime.tools.mcp import AgentMcpTool
from app.agent_artifacts.exporters import WorkbookLimits, render_workbook_v1
from app.mcp_gateway.contracts import DataTapService
from app.mcp_gateway.transport import PossiblySentTimeout, RemoteToolResult

from tests.agent_artifacts.test_analysis_report_export import build_analysis_report_payload


INPUT_SCHEMA = {
    "type": "object",
    "properties": {"keyword": {"type": "string"}},
    "required": ["keyword"],
    "additionalProperties": False,
}
OUTPUT_SCHEMA = {
    "type": "object",
    "properties": {"result": {"type": "string"}},
    "required": ["result"],
    "additionalProperties": False,
}


class _RecordingTransport:
    def __init__(self) -> None:
        self.calls: list[dict[str, Any]] = []
        self._outcomes: list[Any] = [
            PossiblySentTimeout("result not confirmed"),
            RemoteToolResult(
                structured_content={"result": '{"rows":[{"keyword":"母婴"}]}'},
                is_error=False,
                upstream_request_id="upstream-independent-2",
            ),
        ]

    async def call_tool(
        self,
        service: DataTapService,
        remote_name: str,
        arguments: Mapping[str, Any],
    ) -> RemoteToolResult:
        self.calls.append(
            {"service": service, "remote_name": remote_name, "arguments": dict(arguments)}
        )
        outcome = self._outcomes.pop(0)
        if isinstance(outcome, Exception):
            raise outcome
        return outcome


class _FakeCoordinator:
    """仅模拟 durable coordinator 的最小接口，测试 dispatch 语义而非数据库。"""

    def __init__(self) -> None:
        self.states: dict[str, str] = {}
        self.prepared: list[tuple[str, dict[str, Any]]] = []

    async def prepare(
        self,
        context: ToolContext,
        *,
        logical_call_id: str,
        args_hash: str,
        normalized_arguments: Mapping[str, Any],
    ) -> None:
        del context, args_hash
        if logical_call_id in self.states:
            raise AssertionError("the test must use an independent logical call")
        self.states[logical_call_id] = "running"
        self.prepared.append((logical_call_id, dict(normalized_arguments)))
        return None

    async def finalize_unknown(self, *, logical_call_id: str, **_: Any) -> None:
        self.states[logical_call_id] = "unknown"

    async def finalize_success(self, *, logical_call_id: str, **_: Any) -> tuple[str, dict[str, Any]]:
        self.states[logical_call_id] = "settled"
        return "evidence-independent-2", {"truncated": False, "status": "available"}


async def test_unknown_call_does_not_replay_when_model_continues_independent_tool() -> None:
    transport = _RecordingTransport()
    tool = AgentMcpTool(
        internal_name="query_analysis_data",
        service=DataTapService.INSIGHT_CUBE,
        remote_name="datatap.insight.query.analysis.v1",
        input_schema=INPUT_SCHEMA,
        output_schema=OUTPUT_SCHEMA,
        transport=transport,
        session_factory=lambda: None,
    )
    coordinator = _FakeCoordinator()
    tool._coordinator = coordinator
    context = ToolContext(
        user_id="user-1",
        session_id="session-1",
        run_id="run-1",
        profile_name="session_analyst_v1",
        step_id="step-1",
    )

    unknown = await tool.execute(context, {"keyword": "美妆"})
    continued = await tool.execute(context, {"keyword": "母婴"})

    assert unknown.status == "unknown"
    assert unknown.error_type == "result_unknown"
    assert continued.status == "success"
    assert len(transport.calls) == 2
    assert [call["arguments"] for call in transport.calls] == [
        {"keyword": "美妆"},
        {"keyword": "母婴"},
    ]
    assert len(coordinator.prepared) == 2
    assert coordinator.prepared[0][0] != coordinator.prepared[1][0]
    assert set(coordinator.states.values()) == {"unknown", "settled"}


def test_long_tail_excel_uses_generic_analysis_report_projection() -> None:
    from openpyxl import load_workbook
    from io import BytesIO

    from app.agent_artifacts.payloads.analysis_report import AnalysisReportV1

    report = AnalysisReportV1.model_validate(build_analysis_report_payload(row_count=7))
    workbook = render_workbook_v1(
        report,
        exporter_version="analysis-report-v1",
        limits=WorkbookLimits(
            max_sheets=16,
            max_rows_per_sheet=100,
            max_columns=16,
            max_cell_chars=32767,
            max_bytes=2_000_000,
        ),
    )

    loaded = load_workbook(BytesIO(workbook), data_only=False)
    assert "平台明细" in loaded.sheetnames
    values = [
        cell.value
        for sheet in loaded.worksheets
        for row in sheet.iter_rows()
        for cell in row
    ]
    assert len([value for value in values if value == "平台-0"]) == 1
    assert len([value for value in values if isinstance(value, str) and value.startswith("平台-")]) == 7
