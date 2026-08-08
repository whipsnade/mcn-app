"""方案 A 六场景对比的纯数据契约与 Gate 汇总。

真实 Runtime 调用由 CLI 注入执行器；本模块不读取凭证、不调用模型或 DataTap，且所有输出
均写入新的 round 目录，避免覆盖既有真实供应商结果。
"""

import json
import re
from collections.abc import Callable
from dataclasses import asdict, dataclass
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from typing import Literal, Protocol
from uuid import uuid4

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field, field_validator
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.agent_artifacts.exporters import ArtifactExportUnsupported, export_artifact
from app.agent_artifacts.models import AgentArtifact, AgentArtifactVersion
from app.agent_runtime.events import AgentEventBroker, AgentEventStream
from app.agent_runtime.executor import SESSION_ANALYST_PROFILE
from app.agent_runtime.models import (
    AgentMessage,
    AgentRun,
    AgentSession,
    AgentToolCall,
    EvidenceItem,
)
from app.core.config import Settings
from app.identity.models import User, UserChannelPermission
from app.identity.service import IdentityService
from app.marketing_capability_pack.runtime import (
    MarketingRunCapability,
    build_marketing_run_capability,
    render_marketing_system_prompt,
)
from app.mcp_gateway.contracts import DataTapService
from app.pi_runtime_poc.gate import HARD_CHECKS, Summary, evaluate_case, finalize_execution
from app.pi_runtime_poc.rpc import PiRpcClient, PiRpcConfig
from app.pi_runtime_poc.runner import PiClientFactory, PiPocRunner

RuntimeName = Literal["pi"]
CaseExecutionStatus = Literal["completed", "failed", "skipped_dependency", "not_run"]
_SECRET_PATTERN = re.compile(r"(?:sk-[A-Za-z0-9._-]+|Bearer\s+\S+)", re.IGNORECASE)
_REPORT_BEHAVIOR = "report"
_DATATAP_SERVICE_SLUGS = frozenset(
    {
        "insight-cube-mcp",
        "social-grow-mcp",
        "social-grow-content-mcp",
        "bilibili-mcp",
    }
)
_PI_DATATAP_AUDIT_SERVICE = "pi_poc_datatap"
_DATATAP_AUDIT_SERVICES = frozenset(
    (*[service.value for service in DataTapService], _PI_DATATAP_AUDIT_SERVICE)
)


def _is_datatap_audit_service(service: str) -> bool:
    """只把真实 DataTap transport 与 Pi 的直接 DataTap 旁路计入调用数。"""
    return service in _DATATAP_AUDIT_SERVICES


@dataclass(frozen=True)
class PocCase:
    case_id: str
    user_question: str
    date_anchor: str
    expected_behavior: Literal["report", "drilldown", "clarify", "refuse"]
    required_artifact_type: str | None
    depends_on_case_id: str | None = None


@dataclass(frozen=True)
class PocCaseResult:
    case_id: str
    runtime: RuntimeName
    run_id: str | None
    status: CaseExecutionStatus
    error_code: str | None
    outcome: str | None
    artifact_versions: tuple[str, ...]
    evidence_ids: tuple[str, ...]
    metrics: dict[str, float | int | bool | None]
    diagnostic_path: str
    hard_checks: dict[str, bool]


class ScoreWithReason(BaseModel):
    model_config = ConfigDict(extra="forbid")

    score: int = Field(ge=1, le=5)
    reason: str

    @field_validator("reason")
    @classmethod
    def validate_reason(cls, value: str) -> str:
        if not value.strip():
            raise ValueError("poc_review_reason_required")
        return value


class ReportReview(BaseModel):
    model_config = ConfigDict(extra="forbid")

    factuality: ScoreWithReason
    insight: ScoreWithReason
    actionability: ScoreWithReason
    limitations: ScoreWithReason


class HumanReview(BaseModel):
    model_config = ConfigDict(extra="forbid")

    reviewer: str
    reviewed_at: str
    reports: dict[str, ReportReview]

    @field_validator("reports")
    @classmethod
    def validate_reports(cls, value: dict[str, ReportReview]) -> dict[str, ReportReview]:
        expected = {"brand-research-v1", "campaign-evaluation-v1", "kol-selection-v1"}
        if set(value) != expected:
            raise ValueError("poc_review_reports_must_be_exact")
        return value


class RuntimeCaseExecutor(Protocol):
    async def execute(self, case: PocCase, *, prior_run_id: str | None = None) -> PocCaseResult: ...


class PocCaseFactory:
    """为每个 Pi case 建立独立、可审计且同内容的 POC Run。"""

    def __init__(
        self,
        session_factory: async_sessionmaker[AsyncSession],
        *,
        round_id: str,
        model_name: str,
    ) -> None:
        if not model_name:
            raise ValueError("poc_model_name_required")
        self._session_factory = session_factory
        self._round_id = round_id
        self._model_name = model_name

    async def create(self, case: PocCase, *, prior_run_id: str | None = None) -> str:
        now = datetime.now(UTC).replace(tzinfo=None)
        async with self._session_factory() as db:
            dependency: dict[str, object] | None = None
            if prior_run_id is not None:
                prior = await db.get(AgentRun, prior_run_id)
                if prior is None:
                    raise ValueError("poc_dependency_run_not_found")
                prior_snapshot = (prior.prompt_snapshot_json or {}).get("pi_runtime_poc")
                if (
                    not isinstance(prior_snapshot, dict)
                    or prior_snapshot.get("runtime") != "pi"
                    or prior_snapshot.get("case_id") != case.depends_on_case_id
                ):
                    raise ValueError("poc_dependency_run_mismatch")
                version_ids = tuple(
                    (
                        await db.scalars(
                            select(AgentArtifactVersion.id)
                            .join(AgentArtifact, AgentArtifact.id == AgentArtifactVersion.artifact_id)
                            .where(
                                AgentArtifactVersion.source_run_id == prior.id,
                                AgentArtifact.session_id == prior.session_id,
                                AgentArtifact.status == "published",
                            )
                        )
                    ).all()
                )
                if not version_ids:
                    raise ValueError("poc_dependency_artifact_required")
                user_id = prior.user_id
                session_id = prior.session_id
                message_sequence = (
                    await db.scalar(
                        select(func.max(AgentMessage.sequence)).where(AgentMessage.session_id == session_id)
                    )
                    or 0
                ) + 1
                dependency = {
                    "case_id": case.depends_on_case_id,
                    "artifact_version_ids": list(version_ids),
                }
            else:
                user = User(
                    id=str(uuid4()),
                    nickname="pi-poc-pi",
                    role="user",
                    status="active",
                    industries=["美食"],
                    created_at=now,
                    updated_at=now,
                )
                db.add(user)
                await db.flush()
                for channel in IdentityService.default_channels:
                    db.add(
                        UserChannelPermission(
                            id=str(uuid4()),
                            user_id=user.id,
                            channel=channel,
                            is_enabled=True,
                            created_at=now,
                            updated_at=now,
                        )
                    )
                session = AgentSession(
                    id=str(uuid4()),
                    user_id=user.id,
                    title=f"Pi POC {case.case_id}",
                    status="active",
                    session_summary=f"date_anchor={case.date_anchor}; case_id={case.case_id}",
                    summary_version=1,
                    created_at=now,
                    updated_at=now,
                )
                db.add(session)
                await db.flush()
                user_id = user.id
                session_id = session.id
                message_sequence = 1

            run = AgentRun(
                id=str(uuid4()),
                session_id=session_id,
                user_id=user_id,
                run_kind="user",
                visibility="user",
                profile_name=SESSION_ANALYST_PROFILE,
                profile_version="v1",
                model=self._model_name,
                status="queued",
                prompt_snapshot_json={
                    "marketing_capability_pack": build_marketing_run_capability(
                        model_version=self._model_name
                    ).model_dump(mode="json"),
                    "pi_runtime_poc": {
                        "round_id": self._round_id,
                        "case_id": case.case_id,
                        "runtime": "pi",
                        "date_anchor": case.date_anchor,
                        "billing_mode": "disabled",
                        **({"dependency": dependency} if dependency is not None else {}),
                    }
                },
                created_at=now,
            )
            # 三层均为标量 FK、没有 ORM relationship：严格按 Session → Run → Message 刷新，
            # 最后再回填 Run.input_message_id，避免 MySQL 任一层被提前插入。
            db.add(run)
            await db.flush()
            message = AgentMessage(
                id=str(uuid4()),
                session_id=session_id,
                run_id=run.id,
                role="user",
                content=case.user_question,
                sequence=message_sequence,
                created_at=now,
            )
            run.input_message_id = message.id
            db.add(message)
            await db.commit()
            return run.id


class PiRuntimeCaseExecutor:
    def __init__(
        self,
        *,
        factory: PocCaseFactory,
        session_factory: async_sessionmaker[AsyncSession],
        settings: Settings,
        client_factory: PiClientFactory,
        output_root: Path,
        worker_id: str = "pi-poc",
    ) -> None:
        self._factory = factory
        self._session_factory = session_factory
        self._settings = settings
        self._client_factory = client_factory
        self._output_root = output_root
        self._worker_id = worker_id

    async def execute(self, case: PocCase, *, prior_run_id: str | None = None) -> PocCaseResult:
        run_id = await self._factory.create(case, prior_run_id=prior_run_id)
        async with self._session_factory() as db:
            runner = PiPocRunner(
                db=db,
                events=AgentEventStream(db, AgentEventBroker()),
                settings=self._settings,
                worker_id=self._worker_id,
                client_factory=self._client_factory,
            )
            outcome = await runner.run(run_id)
        result = await _collect_case_result(
            self._session_factory, case, run_id, outcome, self._output_root
        )
        return result


def load_cases(path: Path) -> tuple[PocCase, ...]:
    """加载版本化、无供应商结果的 fixture。"""
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, list):
        raise TypeError("poc_cases_must_be_array")
    cases = tuple(PocCase(**item) for item in raw)
    if len({case.case_id for case in cases}) != len(cases):
        raise ValueError("poc_case_id_duplicate")
    return cases


def assess_gate_a(
    cases: tuple[PocCase, ...],
    results: tuple[PocCaseResult, ...],
    human_review: HumanReview | None = None,
) -> dict[str, object]:
    """按 Pi-only 绝对事实分类，不计算任何 Current 相对指标。"""
    expected_ids = [case.case_id for case in cases]
    exact = len(cases) == 6 and len(results) == 6 and [result.case_id for result in results] == expected_ids
    if not exact or any(result.runtime != "pi" or result.status in {"failed", "not_run"} for result in results):
        return {"gate": "INFRA_FAILED", "hard_checks": {"exact_pi_cases": exact}}
    if any(result.status != "completed" or result.outcome is None for result in results):
        return {"gate": "INFRA_FAILED", "hard_checks": {"cases_evaluable": False}}
    if human_review is None:
        raise ValueError("poc_human_review_required")
    check_values = [value for result in results for value in result.hard_checks.values()]
    checks = {f"case:{result.case_id}:{key}": value for result in results for key, value in result.hard_checks.items()}
    scores = [
        score.score
        for report in human_review.reports.values()
        for score in (report.factuality, report.insight, report.actionability, report.limitations)
    ]
    passed = bool(check_values) and all(check_values) and all(score >= 3 for score in scores)
    return {"gate": "PASS" if passed else "EVALUATED_FAIL", "hard_checks": checks}


def write_append_only_round(
    output_root: Path, round_id: str, results: tuple[PocCaseResult, ...]
) -> Path:
    """写入新 round 的安全汇总；重复 round 或疑似密钥一律拒绝。"""
    if not round_id or Path(round_id).name != round_id:
        raise ValueError("invalid_poc_round_id")
    round_dir = output_root / round_id
    if round_dir.exists():
        raise FileExistsError(round_dir)
    payload = {"round_id": round_id, "results": [asdict(result) for result in results]}
    serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    if _SECRET_PATTERN.search(serialized):
        raise ValueError("poc_output_contains_secret")
    round_dir.mkdir(parents=True, exist_ok=False)
    summary = round_dir / "summary.json"
    summary.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return summary


def begin_round(output_root: Path, round_id: str) -> Path:
    """建立唯一真实轮次目录；已有目录一律不复用或覆盖。"""
    if not round_id or Path(round_id).name != round_id:
        raise ValueError("invalid_poc_round_id")
    round_dir = output_root / round_id
    round_dir.mkdir(parents=True, exist_ok=False)
    return round_dir


def write_round_summary(round_dir: Path, results: tuple[PocCaseResult, ...], gate: dict[str, object]) -> Path:
    """写入单轮汇总；只允许在已由 ``begin_round`` 建立的新轮次中执行一次。"""
    summary = round_dir / "summary.json"
    if not round_dir.is_dir() or summary.exists():
        raise FileExistsError(summary)
    payload = {"round_id": round_dir.name, "results": [asdict(result) for result in results], "gate": gate}
    _write_safe_json(summary, payload)
    return summary


def write_case_result(round_dir: Path, result: PocCaseResult) -> Path:
    path = round_dir / result.case_id / f"{result.runtime}.json"
    if not round_dir.is_dir() or path.exists():
        raise FileExistsError(path)
    _write_safe_json(path, asdict(result))
    return path


def make_non_executed_result(
    round_dir: Path,
    case: PocCase,
    *,
    status: Literal["failed", "skipped_dependency", "not_run"],
    error_code: str,
) -> PocCaseResult:
    """构造不泄漏异常详情的稳定未执行案例结果。"""
    return PocCaseResult(
        case_id=case.case_id,
        runtime="pi",
        run_id=None,
        status=status,
        error_code=error_code,
        outcome=None,
        artifact_versions=(),
        evidence_ids=(),
        metrics={
            "datatap_tool_calls": 0,
            "points_reserved": 0,
            "points_settled": 0,
            "wallet_rows": 0,
            "wallet_transactions": 0,
            "hanging_tool_calls": 0,
        },
        diagnostic_path=str(round_dir / case.case_id / "pi.json"),
        hard_checks={},
    )


def write_execution_manifest(
    round_dir: Path,
    cases: tuple[PocCase, ...],
    results: tuple[PocCaseResult, ...],
) -> Path:
    """一次性写入完整 Pi-only 执行事实；拒绝缺案例或重复写。"""
    expected = [case.case_id for case in cases]
    actual = [result.case_id for result in results]
    if actual != expected or len(results) != 6 or any(result.runtime != "pi" for result in results):
        raise ValueError("poc_execution_requires_exact_cases")
    path = round_dir / "execution.json"
    if not round_dir.is_dir() or path.exists():
        raise FileExistsError(path)
    _write_safe_json(
        path,
        {"round_id": round_dir.name, "runtime": "pi", "results": [asdict(result) for result in results]},
    )
    return path


def _write_safe_json(path: Path, payload: object) -> None:
    serialized = json.dumps(payload, ensure_ascii=False, sort_keys=True)
    if _SECRET_PATTERN.search(serialized):
        raise ValueError("poc_output_contains_secret")
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("x", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=False, indent=2) + "\n")


def build_real_pi_client_factory(settings: Settings) -> Callable[[AgentRun, str], object]:
    """构造真实 Pi 子进程工厂，仅传入主 Runtime 的同一模型和 DataTap 配置。

    调用者必须先以真实配置初始化 ``Settings``，并将数据库显式覆写为 POC 库。
    任一缺失项直接失败，不以 mock 或替代模型继续。
    """
    if set(settings.datatap_mcp_urls) != _DATATAP_SERVICE_SLUGS:
        raise RuntimeError("pi_poc_datatap_endpoint_mapping_required")
    thinking = settings.tencent_plan_reasoning_effort
    if thinking is None:
        raise RuntimeError("pi_poc_same_thinking_required")
    root = Path(__file__).parents[3]
    pi_root = root / "pi-runtime"
    executable = pi_root / "node_modules" / ".bin" / "pi"
    adapter_extension = pi_root / "node_modules" / "pi-mcp-adapter" / "index.ts"
    extension = pi_root / "src" / "extensions" / "poc-runtime.ts"
    if not executable.is_file() or not adapter_extension.is_file() or not extension.is_file():
        raise RuntimeError("pi_poc_runtime_resources_missing")
    provider = "kol_insight_pi_poc"
    models_json = json.dumps(
        {
            "providers": {
                provider: {
                    "baseUrl": str(settings.tencent_plan_base_url),
                    "apiKey": "$TENCENT_PLAN_API_KEY",
                    "authHeader": True,
                    "api": "openai-completions",
                    "models": [
                        {
                            "id": settings.tencent_plan_model,
                            "reasoning": True,
                            "thinkingLevelMap": {thinking: thinking},
                            "input": ["text"],
                            "compat": {
                                "supportsReasoningEffort": True,
                                "thinkingFormat": "reasoning_effort",
                            },
                        }
                    ],
                }
            }
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )

    def factory(run: AgentRun, token: str) -> object:
        try:
            capability = MarketingRunCapability.model_validate(
                (run.prompt_snapshot_json or {}).get("marketing_capability_pack")
            )
        except ValueError as exc:
            raise ValueError("pi_marketing_capability_snapshot_invalid") from exc
        environment = {
            "TENCENT_PLAN_API_KEY": settings.tencent_plan_api_key.get_secret_value(),
            "DATATAP_MCP_TOKEN": settings.datatap_mcp_token.get_secret_value(),
            "DATATAP_INSIGHT_CUBE_MCP_URL": str(settings.datatap_mcp_urls["insight-cube-mcp"]),
            "DATATAP_SOCIAL_GROW_MCP_URL": str(settings.datatap_mcp_urls["social-grow-mcp"]),
            "DATATAP_SOCIAL_GROW_CONTENT_MCP_URL": str(
                settings.datatap_mcp_urls["social-grow-content-mcp"]
            ),
            # adapter 项目配置中的 aktools 是代理名称；受控 DataTap mapping 的实际服务为 bilibili。
            "DATATAP_AKTOOLS_MCP_URL": str(settings.datatap_mcp_urls["bilibili-mcp"]),
            "PI_RUNTIME_POC_BASE_URL": str(settings.pi_runtime_poc_base_url),
            "PI_RUNTIME_POC_RUN_ID": run.id,
            "PI_RUNTIME_POC_TOKEN": token,
        }
        return PiRpcClient.start(
            PiRpcConfig(
                executable=str(executable),
                extensions=(str(adapter_extension), str(extension)),
                skills=(),
                append_system_prompt=render_marketing_system_prompt(capability),
                timeout_seconds=settings.pi_runtime_poc_run_timeout_seconds,
                environment=environment,
                # adapter 在临时 agent dir 缺少 metadata cache 时会 bootstrap 四个服务，
                # 与单工具 smoke 的一次 tools/list 边界冲突。空 cache 不含 endpoint、token 或工具数据，
                # 仅阻止 bootstrap；首次指定 server 的 mcp proxy call 仍按需 discovery。
                agent_files={
                    "models.json": models_json,
                    "mcp-cache.json": '{"version":1,"servers":{}}',
                },
                provider=provider,
                model=settings.tencent_plan_model,
                thinking=thinking,
                cwd=str(pi_root),
            )
        )

    return factory


async def _collect_case_result(
    session_factory: async_sessionmaker[AsyncSession],
    case: PocCase,
    run_id: str,
    outcome: object,
    output_root: Path,
) -> PocCaseResult:
    """从 POC DB 唯一事实来源读取结果；不读取 Pi stdout 或供应商原文。"""
    async with session_factory() as db:
        run = await db.get(AgentRun, run_id)
        if run is None:
            raise LookupError("poc_run_not_found")
        version_rows = (
            await db.execute(
                select(AgentArtifactVersion, AgentArtifact)
                .join(AgentArtifact, AgentArtifact.id == AgentArtifactVersion.artifact_id)
                .where(AgentArtifactVersion.source_run_id == run_id)
            )
        ).all()
        versions = tuple(version.id for version, _artifact in version_rows)
        evidence = tuple(
            (
                await db.scalars(
                    select(EvidenceItem.id).where(EvidenceItem.run_id == run_id)
                )
            ).all()
        )
        call_total = await db.scalar(
            select(func.count())
            .select_from(AgentToolCall)
            .where(
                AgentToolCall.run_id == run_id,
                AgentToolCall.service.in_(_DATATAP_AUDIT_SERVICES),
            )
        )
        call_settled = await db.scalar(
            select(func.count())
            .select_from(AgentToolCall)
            .where(
                AgentToolCall.run_id == run_id,
                AgentToolCall.service.in_(_DATATAP_AUDIT_SERVICES),
                AgentToolCall.status == "settled",
            )
        )
        points_settled = await db.scalar(
            select(func.coalesce(func.sum(AgentToolCall.points_settled), 0)).where(
                AgentToolCall.run_id == run_id
            )
        )
        points_reserved = await db.scalar(
            select(func.coalesce(func.sum(AgentToolCall.points_reserved), 0)).where(
                AgentToolCall.run_id == run_id
            )
        )
        value = outcome.value if hasattr(outcome, "value") else outcome
        outcome_text = str(value if value is not None else run.status)
        report_case = case.expected_behavior == _REPORT_BEHAVIOR
        expected_versions = (
            version_rows
            if case.required_artifact_type is None
            else [row for row in version_rows if row[1].artifact_type == case.required_artifact_type]
        )
        excel_openable = all(_is_openable_excel(version) for version, _artifact in expected_versions)
        hard_checks = {
            "no_secret": True,
            "lineage_complete": all(
                isinstance(version.lineage_snapshot_json, dict)
                and isinstance(version.validation_json, dict)
                and version.validation_json.get("valid") is True
                for version, _artifact in version_rows
            ),
            "expected_artifact": (not report_case) or bool(expected_versions),
            "excel_openable": (not report_case) or (bool(expected_versions) and excel_openable),
            "bi_payload_same_version": (not report_case)
            or all(isinstance(version.payload_json, dict) for version, _artifact in expected_versions),
        }
        return PocCaseResult(
            case_id=case.case_id,
            runtime="pi",
            run_id=run_id,
            status="completed",
            error_code=None,
            outcome=outcome_text,
            artifact_versions=versions,
            evidence_ids=evidence,
            metrics={
                "coverage": len(evidence),
                "mcp_parameter_validity": (call_settled or 0) / (call_total or 1),
                "artifact_completeness": int(bool(versions)) if report_case else 1,
                "human_readability": None,
                "datatap_tool_calls": int(call_total or 0),
                "points_settled": int(points_settled or 0),
                "points_reserved": int(points_reserved or 0),
            },
            diagnostic_path=str(output_root / case.case_id / "pi.json"),
            hard_checks=hard_checks,
        )


def _is_openable_excel(version: AgentArtifactVersion) -> bool:
    """只做 openpyxl 的结构验证，且只从该不可变 Version 渲染。"""
    try:
        workbook = load_workbook(BytesIO(export_artifact(version)), data_only=False)
    except (ArtifactExportUnsupported, OSError, ValueError, TypeError):
        return False
    return bool(workbook.sheetnames)


__all__ = [
    "HARD_CHECKS",
    "CaseExecutionStatus",
    "PiRuntimeCaseExecutor",
    "PocCase",
    "PocCaseFactory",
    "PocCaseResult",
    "RuntimeCaseExecutor",
    "RuntimeName",
    "Summary",
    "assess_gate_a",
    "begin_round",
    "build_real_pi_client_factory",
    "evaluate_case",
    "finalize_execution",
    "load_cases",
    "make_non_executed_result",
    "write_append_only_round",
    "write_case_result",
    "write_execution_manifest",
    "write_round_summary",
]
