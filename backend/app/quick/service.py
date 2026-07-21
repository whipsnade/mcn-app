"""快捷功能（2x2 按钮）的同步执行护栏与业务编排。

QuickCallService 串联：require_enabled → validate_input → transport.call_tool →
validate_output → 记账（reserve→成功 settle/失败 release，reference_type=
"quick_mcp_call"）→ 写 quick_mcp_calls 留痕 → 响应归一化（{result: str} 解析）。

爆贴/达人推荐/达人详情的工具选择与参数填充由 quick/agent.py 的模型小循环
决策；本模块只做：执行护栏、结果归一化清洗（模型 result 可能不完整）、
预算过滤/top50 截断（纯代码排序过滤）与评估上传。
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from datetime import UTC, datetime, timedelta
from typing import Any
from uuid import uuid4

import openpyxl
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.billing.models import WalletTransaction
from app.billing.service import WalletService
from app.db.session import SessionFactory
from app.identity.models import User
from app.mcp_gateway.accounting import MCP_COST
from app.mcp_gateway.registry import ToolNotEnabledError, ToolRegistryService
from app.mcp_gateway.transport import (
    JsonValue,
    McpCircuitOpen,
    McpConnectionError,
    McpConnectionTimeout,
    McpGatewayTimeout,
    McpProtocolError,
    McpQueueTimeout,
    McpTransport,
    McpUpstreamHttpError,
    PossiblySentTimeout,
)
from app.mcp_gateway.validation import McpValidationError, validate_input, validate_output
from app.model.contracts import ChatMessage, ModelAdapter, StructuredModelRequest
from app.model.prompts import EVALUATE_PROMPT
from app.orchestration.schemas import PlannerTool
from app.quick.agent import (
    DATASOURCE_BY_PLATFORM,
    KOL_SEARCH_TOOLS,
    KolDetailFeatureResult,
    QuickToolCaller,
    quick_feature_tool_names,
    run_quick_feature,
)
from app.quick.errors import QuickCallFailedError
from app.quick.models import QuickMcpCall
from app.quick.schemas import KolRecommendationItem, TopPostItem


__all__ = [
    "DATASOURCE_BY_PLATFORM",
    "KOL_SEARCH_TOOLS",
    "MAX_UPLOAD_BYTES",
    "QuickCallFailedError",
    "QuickCallService",
    "QuickService",
]


MAX_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_TABLE_CHARS = 8000
# 报价有效性下限：低于该值（或为 0/缺失）视为无报价。
MIN_VALID_PRICE = 500.0


def _now() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def parse_datatap_result(output: JsonValue) -> JsonValue:
    """DataTap {result: str} 包装 → JSON 解析；非 JSON 文本原样透传。"""
    if isinstance(output, dict) and isinstance(output.get("result"), str):
        text = output["result"]
        try:
            return json.loads(text)
        except ValueError:
            return text
    return output


class QuickCallService:
    """单次快捷 MCP 调用：校验、计费、留痕一体化。"""

    def __init__(
        self,
        db_session: AsyncSession,
        transport: McpTransport,
        *,
        registry: ToolRegistryService | None = None,
    ) -> None:
        self._db = db_session
        self._transport = transport
        self._registry = registry or ToolRegistryService(db_session, transport)

    async def call_tool(
        self,
        user_id: str,
        *,
        feature: str,
        internal_tool_name: str,
        arguments: dict[str, Any],
    ) -> JsonValue:
        try:
            approved = await self._registry.require_enabled(internal_tool_name)
        except ToolNotEnabledError as error:
            raise QuickCallFailedError("tool_not_enabled") from error
        try:
            validated = validate_input(arguments, approved.input_schema)
        except McpValidationError as error:
            raise QuickCallFailedError("input_validation_error") from error

        call_id = str(uuid4())
        wallet = WalletService(self._db)
        # 余额不足时 InsufficientPointsError 直接上抛（路由映射 409），不产生留痕。
        await wallet.reserve(
            user_id,
            MCP_COST,
            f"quick:{call_id}:reserve",
            call_id,
            reference_type="quick_mcp_call",
        )
        reserve_tx = await self._transaction(f"quick:{call_id}:reserve")
        row = QuickMcpCall(
            id=call_id,
            user_id=user_id,
            feature=feature,
            internal_tool_name=internal_tool_name,
            arguments_json=validated,
            status="running",
            points_cost=MCP_COST,
            reserve_transaction_id=reserve_tx.id if reserve_tx is not None else None,
            created_at=_now(),
        )
        self._db.add(row)
        # 预留与留痕先于网络调用持久化，崩溃后由恢复清扫释放。
        await self._db.commit()

        try:
            result = await self._transport.call_tool(
                approved.service, approved.remote_name, validated
            )
        except PossiblySentTimeout:
            await self._finish_failed(row, "possibly_sent_timeout")
        except McpConnectionTimeout:
            await self._finish_failed(row, "connection_timeout")
        except McpConnectionError:
            await self._finish_failed(row, "connection_error")
        except McpGatewayTimeout:
            await self._finish_failed(row, "upstream_timeout")
        except McpUpstreamHttpError:
            await self._finish_failed(row, "upstream_http_error")
        except McpQueueTimeout:
            await self._finish_failed(row, "mcp_queue_timeout")
        except McpCircuitOpen:
            await self._finish_failed(row, "mcp_service_unavailable")
        except McpProtocolError:
            await self._finish_failed(row, "protocol_error")
        except Exception:
            await self._finish_failed(row, "upstream_error")

        if result.is_error:
            await self._finish_failed(row, "upstream_tool_error")
        # 与 McpCallService 一致：is_error=False 的 null content 是合法空结果。
        if result.structured_content is None:
            validated_output: JsonValue = None
        else:
            try:
                validated_output = validate_output(result.structured_content, approved.output_schema)
            except McpValidationError:
                await self._finish_failed(row, "output_validation_error")

        try:
            await wallet.settle(
                user_id,
                MCP_COST,
                f"quick:{row.id}:settle",
                row.id,
                reference_type="quick_mcp_call",
            )
        except ValueError:
            # 预留已被恢复清扫释放：留痕失败，不再重复结算。
            await self._finish_failed(row, "reservation_lost", release=False)
        settle_tx = await self._transaction(f"quick:{row.id}:settle")
        row.status = "succeeded"
        row.error_type = None
        row.settlement_transaction_id = settle_tx.id if settle_tx is not None else None
        row.completed_at = _now()
        await self._db.commit()
        return parse_datatap_result(validated_output)

    async def _finish_failed(
        self, row: QuickMcpCall, error_type: str, *, release: bool = True
    ) -> None:
        if release:
            await WalletService(self._db).release(
                row.user_id,
                row.points_cost,
                f"quick:{row.id}:release",
                row.id,
                reference_type="quick_mcp_call",
            )
            release_tx = await self._transaction(f"quick:{row.id}:release")
            row.settlement_transaction_id = (
                release_tx.id if release_tx is not None else row.settlement_transaction_id
            )
        row.status = "failed"
        row.error_type = error_type
        row.completed_at = _now()
        await self._db.commit()
        raise QuickCallFailedError(error_type)

    async def _transaction(self, idempotency_key: str) -> WalletTransaction | None:
        return await self._db.scalar(
            select(WalletTransaction).where(WalletTransaction.idempotency_key == idempotency_key)
        )


async def release_stale_quick_calls(*, older_than_seconds: int = 300) -> int:
    """释放 5 分钟前悬挂（status=running）的快捷预留；与恢复循环同周期调用。"""
    async with SessionFactory.begin() as db:
        return await sweep_stale_quick_calls(db, older_than_seconds=older_than_seconds)


async def sweep_stale_quick_calls(
    db: AsyncSession, *, older_than_seconds: int = 300
) -> int:
    """清扫实现（调用方持有事务）；release 与服务内失败路径共用
    quick:{id}:release 幂等键，天然去重。"""
    cutoff = _now() - timedelta(seconds=older_than_seconds)
    rows = list(
        (
            await db.scalars(
                select(QuickMcpCall)
                .where(
                    QuickMcpCall.status == "running",
                    QuickMcpCall.created_at < cutoff,
                )
                .with_for_update()
            )
        ).all()
    )
    wallet = WalletService(db)
    for row in rows:
        await wallet.release(
            row.user_id,
            row.points_cost,
            f"quick:{row.id}:release",
            row.id,
            reference_type="quick_mcp_call",
        )
        release_tx = await db.scalar(
            select(WalletTransaction).where(
                WalletTransaction.idempotency_key == f"quick:{row.id}:release"
            )
        )
        row.status = "failed"
        row.error_type = row.error_type or "recovery_released"
        row.settlement_transaction_id = (
            release_tx.id if release_tx is not None else row.settlement_transaction_id
        )
        row.completed_at = _now()
    await db.flush()
    return len(rows)


def user_industries(user: User) -> list[str]:
    """行业属性（多值）；空值兜底 ["美食"]。"""
    values = [str(item).strip() for item in (user.industries or []) if str(item).strip()]
    return values or ["美食"]


def _number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(",", "").strip())
        except ValueError:
            return None
    return None


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _first(item: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = item.get(key)
        if value is not None and value != "":
            return value
    return None


def _find_rows(payload: JsonValue, preferred_keys: tuple[str, ...]) -> list[dict[str, Any]]:
    """从解析后的载荷中定位记录列表（上游中文字段名 + 防御性兜底）。"""
    if isinstance(payload, list):
        return [entry for entry in payload if isinstance(entry, dict)]
    if not isinstance(payload, dict):
        return []
    for key in preferred_keys:
        value = payload.get(key)
        if isinstance(value, list):
            return [entry for entry in value if isinstance(entry, dict)]
    for value in payload.values():
        if isinstance(value, list) and value and all(isinstance(entry, dict) for entry in value):
            return value
    return []


def extract_price(item: dict[str, Any]) -> float | None:
    """从搜索结果提取报价：官方/预估报价（标量或 键/值 数组）取最低有效价。

    有效价 = ≥500 的数值；为 0 或缺失视为无报价（返回 None）。
    """
    candidates: list[float] = []
    for key, value in item.items():
        if "报价" not in key:
            continue
        if isinstance(value, list):
            for entry in value:
                if isinstance(entry, dict):
                    number = _number(entry.get("值"))
                else:
                    number = _number(entry)
                if number is not None:
                    candidates.append(number)
        else:
            number = _number(value)
            if number is not None:
                candidates.append(number)
    valid = [value for value in candidates if value >= MIN_VALID_PRICE]
    return min(valid) if valid else None


@dataclass
class _KolCandidate:
    item: KolRecommendationItem
    interact: float = 0.0


def _normalize_kol(payload: JsonValue, platform: str) -> list[_KolCandidate]:
    """KOL 原始行 → 推荐条目；platform 为缺省值，行内 平台/platform 字段优先。"""
    candidates: list[_KolCandidate] = []
    for entry in _find_rows(payload, ("KOL 列表", "达人列表", "list", "items")):
        kw_uid = _text(_first(entry, "账号ID (kwUid)", "kwUid", "kw_uid"))
        if kw_uid is None:
            continue
        fans = _number(_first(entry, "粉丝数", "抖音粉丝数", "有效粉丝数"))
        engagement = _number(
            _first(entry, "互动率-日常作品", "互动率-图文笔记", "互动率-视频笔记", "互动率")
        )
        tags_raw = _first(
            entry, "Grow-博主类目标签", "Grow-达人类型标签", "星图-达人类型标签", "行业标签"
        )
        tags = (
            [str(tag) for tag in tags_raw if str(tag).strip()]
            if isinstance(tags_raw, list)
            else []
        )
        candidates.append(
            _KolCandidate(
                item=KolRecommendationItem(
                    platform=_text(_first(entry, "平台", "platform")) or platform,
                    kw_uid=kw_uid,
                    nickname=_text(_first(entry, "昵称", "nickname")),
                    fans=int(fans) if fans is not None else None,
                    price=extract_price(entry),
                    engagement_rate=engagement,
                    score=_number(_first(entry, "综合评分", "score")),
                    city=_text(_first(entry, "城市", "IP属地", "省份")),
                    tags=tags,
                ),
                interact=_number(_first(entry, "平均互动", "互动数")) or 0.0,
            )
        )
    return candidates


def _normalize_posts(payload: JsonValue, platform: str) -> list[TopPostItem]:
    posts: list[TopPostItem] = []
    for entry in _find_rows(payload, ("帖子列表", "数据列表", "posts", "list", "items")):
        posts.append(
            TopPostItem(
                title=_text(_first(entry, "标题", "帖子标题", "内容", "title")),
                nickname=_text(_first(entry, "用户昵称", "昵称", "作者", "nickname")),
                interact=_number(_first(entry, "互动数", "互动量", "interact")),
                like=_number(_first(entry, "点赞数", "点赞", "like")),
                comment=_number(_first(entry, "评论数", "评论", "comment")),
                collect=_number(_first(entry, "收藏数", "收藏", "collect")),
                publish_time=_text(_first(entry, "发布时间", "发布日期", "publish_time")),
                url=_text(_first(entry, "帖子链接", "链接", "url")),
                platform=platform,
            )
        )
    return posts


def _extract_detail(payload: JsonValue) -> dict[str, Any]:
    """达人详情提取：模型可能给包装载荷（{"详情列表": [...]}）或裸详情对象。

    裸详情对象本身也含 dict 列表字段（如价格趋势），防御性展开仅当载荷的
    所有值都是 dict 列表（整体呈"包装"形态）时才启用，避免误拆裸对象。
    """
    if isinstance(payload, list):
        return payload[0] if payload and isinstance(payload[0], dict) else {}
    if not isinstance(payload, dict):
        return {}
    for key in ("详情列表", "达人详情列表"):
        value = payload.get(key)
        if isinstance(value, list) and value and isinstance(value[0], dict):
            return value[0]
    if payload and all(
        isinstance(value, list)
        and value
        and all(isinstance(entry, dict) for entry in value)
        for value in payload.values()
    ):
        return next(iter(payload.values()))[0]
    return payload


def apply_budget_filter(
    candidates: list[_KolCandidate], budget: int
) -> list[KolRecommendationItem]:
    """预算过滤与排序（端点层纯代码）：超预算丢弃，无报价排最后，按互动量
    降序取 top50。不做任何工具决策。"""
    priced = sorted(
        (
            candidate
            for candidate in candidates
            if candidate.item.price is not None and candidate.item.price <= budget
        ),
        key=lambda candidate: candidate.interact,
        reverse=True,
    )
    unpriced = sorted(
        (candidate for candidate in candidates if candidate.item.price is None),
        key=lambda candidate: candidate.interact,
        reverse=True,
    )
    return [candidate.item for candidate in (priced + unpriced)[:50]]


class EvaluateDocument(BaseModel):
    """quick_evaluate 模型输出契约。"""

    title: str
    analysis_markdown: str


def render_upload_table(filename: str, content: bytes) -> str:
    """xlsx/csv → 紧凑文本表（截断 ~8k 字符）；不支持的类型与解析失败抛 ValueError。"""
    lowered = filename.lower()
    if lowered.endswith(".xlsx"):
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as error:
            raise ValueError("upload_parse_failed") from error
        lines: list[str] = []
        multiple = len(workbook.sheetnames) > 1
        for sheet in workbook.worksheets:
            if multiple:
                lines.append(f"# sheet: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if cell is None else str(cell) for cell in row]
                if not any(cell.strip() for cell in cells):
                    continue
                lines.append("\t".join(cells).rstrip())
        workbook.close()
    elif lowered.endswith(".csv"):
        text = content.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        lines = [
            "\t".join(row).rstrip()
            for row in reader
            if any(cell.strip() for cell in row)
        ]
    else:
        raise ValueError("unsupported_file_type")
    return "\n".join(lines)[:MAX_TABLE_CHARS]


@dataclass
class QuickService:
    """四个快捷端点的业务编排；points 跟踪本次请求实际计费调用数。"""

    db: AsyncSession
    transport: McpTransport | None = None
    model: ModelAdapter | None = None
    points: int = field(default=0, init=False)
    _calls_service: QuickCallService | None = field(default=None, init=False, repr=False)

    def _calls(self) -> QuickCallService:
        if self.transport is None:
            raise QuickCallFailedError("transport_unavailable")
        if self._calls_service is None:
            self._calls_service = QuickCallService(self.db, self.transport)
        return self._calls_service

    def _require_model(self) -> ModelAdapter:
        if self.model is None:
            raise QuickCallFailedError("model_unavailable")
        return self.model

    async def _call_tool(
        self,
        user_id: str,
        *,
        feature: str,
        internal_tool_name: str,
        arguments: dict[str, Any],
    ) -> JsonValue:
        payload = await self._calls().call_tool(
            user_id,
            feature=feature,
            internal_tool_name=internal_tool_name,
            arguments=arguments,
        )
        self.points += MCP_COST
        return payload

    def _caller(self, user_id: str, feature: str) -> QuickToolCaller:
        async def call(internal_tool_name: str, arguments: dict[str, Any]) -> JsonValue:
            return await self._call_tool(
                user_id,
                feature=feature,
                internal_tool_name=internal_tool_name,
                arguments=arguments,
            )

        return call

    async def _feature_tools(
        self, feature: str, platforms: tuple[str, ...]
    ) -> tuple[PlannerTool, ...]:
        if self.transport is None:
            raise QuickCallFailedError("transport_unavailable")
        registry = ToolRegistryService(self.db, self.transport)
        enabled = await registry.list_enabled()
        wanted = set(quick_feature_tool_names(feature, platforms))
        return tuple(
            PlannerTool.from_approved(item) for item in enabled if item.internal_name in wanted
        )

    async def _run_feature(
        self,
        user: User,
        *,
        feature: str,
        goal: str,
        scenario: dict[str, Any],
        platforms: tuple[str, ...],
        tags: list[str],
        period_days: int = 29,
    ) -> Any:
        industries = user_industries(user)
        tools = await self._feature_tools(feature, platforms)
        return await run_quick_feature(
            db=self.db,
            model=self._require_model(),
            call=self._caller(user.id, feature),
            tools=tools,
            user_id=user.id,
            feature=feature,
            goal=goal,
            scenario=scenario,
            industries=industries,
            tags=tags,
            period_days=period_days,
        )

    async def top_posts(self, user: User, *, platform: str) -> tuple[list[TopPostItem], int]:
        industries = user_industries(user)
        rows = await self._run_feature(
            user,
            feature="top_posts",
            goal=(
                f"获取最近7日（7日热榜）「{platform}」平台「{industries[0]}」行业"
                "互动数最高的10条帖子（按互动数倒序）"
            ),
            scenario={"platform": platform, "datasource": DATASOURCE_BY_PLATFORM[platform]},
            platforms=(platform,),
            tags=["quick:top_posts", f"industry:{industries[0]}", f"platform:{platform}"],
            period_days=6,
        )
        return _normalize_posts(rows, platform)[:10], self.points

    async def kol_recommendations(
        self, user: User, *, budget: int, platforms: list[str]
    ) -> tuple[list[KolRecommendationItem], int]:
        industries = user_industries(user)
        rows = await self._run_feature(
            user,
            feature="kol_recommend",
            goal=(
                f"检索「{'、'.join(platforms)}」平台「{industries[0]}」行业的"
                "达人候选（每平台至多50条，保留报价与平台字段）"
            ),
            scenario={"platforms": platforms, "budget": budget},
            platforms=tuple(platforms),
            tags=(
                ["quick:kol_recommend", f"industry:{industries[0]}"]
                + [f"platform:{platform}" for platform in platforms]
            ),
        )
        merged = _normalize_kol(rows, platforms[0] if platforms else "")
        return apply_budget_filter(merged, budget), self.points

    async def kol_detail(
        self, user: User, *, platform: str, kw_uid: str, nickname: str
    ) -> tuple[dict[str, Any], list[TopPostItem], bool, int]:
        industries = user_industries(user)
        result: KolDetailFeatureResult = await self._run_feature(
            user,
            feature="kol_detail",
            goal=(
                f"获取「{platform}」平台达人「{nickname}」（kw_uid={kw_uid}）的详情"
                "（fansAudience/postSummaryStatistics/priceTrend）"
                "与其近30天互动最高的10条帖子"
            ),
            scenario={
                "platform": platform,
                "kw_uid": kw_uid,
                "nickname": nickname,
                "datasource": DATASOURCE_BY_PLATFORM[platform],
            },
            platforms=(platform,),
            tags=["quick:kol_detail", f"industry:{industries[0]}", f"platform:{platform}"],
        )
        return (
            _extract_detail(result.detail),
            _normalize_posts(result.posts, platform)[:10],
            result.posts_degraded,
            self.points,
        )

    async def evaluate(
        self, user: User, *, filename: str, content: bytes
    ) -> EvaluateDocument:
        industries = user_industries(user)
        table_text = render_upload_table(filename, content)
        result = await self._require_model().complete_json(
            StructuredModelRequest(
                purpose="quick_evaluate",
                template_name=EVALUATE_PROMPT.name,
                messages=(
                    ChatMessage(role="system", content=EVALUATE_PROMPT.system),
                    ChatMessage(
                        role="user",
                        content=json.dumps(
                            {
                                "industries": industries,
                                "uploaded_table": table_text,
                            },
                            ensure_ascii=False,
                        ),
                    ),
                ),
                output_model=EvaluateDocument,
                max_tokens=4096,
                log_context={
                    "user_id": user.id,
                    "tags": ["quick:evaluate", f"industry:{industries[0]}"],
                },
            )
        )
        return result.value
