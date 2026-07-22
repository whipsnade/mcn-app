from __future__ import annotations

import asyncio
from copy import copy
from dataclasses import dataclass, field
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.selection.contract import EXPORT_FIELD_CONTRACT_VERSION
from app.selection.models import SessionKolSelection
from app.selection.scoring import rating, score_reason
from app.selection.service import KolSelectionService
from app.workspace.models import WorkspaceSession


TEMPLATE_PATH = Path(__file__).with_name("templates") / "KOL匹配度分析报告.xlsx"
CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(frozen=True)
class ExportCandidate:
    rank: int
    platform: str
    nickname: str
    followers: int | None
    city: str | None
    total_score: float | None
    rating: str
    stars: str
    profile_url: str | None = None
    dimension_scores: dict[str, float | None] = field(default_factory=dict)
    values: dict[str, Any] = field(default_factory=dict)
    score_reason: str = ""
    source_names: tuple[str, ...] = ()
    collected_at: str | None = None


@dataclass(frozen=True)
class ExportedWorkbook:
    content: bytes
    filename: str
    content_type: str = CONTENT_TYPE


async def export_session_selection(
    db: AsyncSession, user_id: str, session_id: str
) -> ExportedWorkbook:
    rows = await KolSelectionService(db).get_all_for_export(
        user_id=user_id, session_id=session_id
    )  # 归属校验在 service 内，无权限/无会话抛 LookupError("session_not_found")
    if not rows:
        raise LookupError("no_kol_selection")
    session = await db.scalar(
        select(WorkspaceSession).where(
            WorkspaceSession.id == session_id,
            WorkspaceSession.user_id == user_id,
            WorkspaceSession.deleted_at.is_(None),
        )
    )
    if session is None:
        raise LookupError("session_not_found")
    candidates = [
        _selection_candidate(row, rank) for rank, row in enumerate(rows, start=1)
    ]
    generated_at = datetime.now(UTC).replace(tzinfo=None)
    metadata = {
        "brand": session.brand,
        "category": session.category,
        "target_audience": session.target_audience,
        "locations": _export_locations(session),
        "generated_at": generated_at,
        "field_contract_version": EXPORT_FIELD_CONTRACT_VERSION,
    }
    filename = _safe_filename(
        f"{session.brand}_{session.category or '未分类'}_KOL匹配度分析_"
        f"{generated_at.strftime('%Y%m%d_%H%M')}.xlsx"
    )
    return ExportedWorkbook(
        # 模板渲染是 CPU 密集的同步 openpyxl 操作，放到线程避免阻塞事件循环；
        # 不触 db，ORM 属性已在上面的 _selection_candidate 阶段全部取出。
        content=await asyncio.to_thread(
            render_workbook, metadata=metadata, candidates=candidates
        ),
        filename=filename,
    )


def _export_locations(session: WorkspaceSession) -> list[str]:
    """地区标签数据源：brainstorm 画像 region 优先，回退 target_fan_locations。"""
    filters = session.filters_snapshot or {}
    profile = filters.get("brainstorm_profile") or {}
    if isinstance(profile, dict):
        region = str(profile.get("region") or "").strip()
        if region:
            return [region]
    locations = filters.get("target_fan_locations") or []
    if isinstance(locations, list):
        return [str(value) for value in locations if value]
    return []


def _target_region_rate_label(metadata: dict[str, Any]) -> str:
    """达人详细画像的地区占比展示标签：{地区}粉丝占比，缺省「目标地区粉丝占比」。"""
    locations = metadata.get("locations") or []
    if locations:
        first = str(locations[0]).strip()
        if first:
            return f"{first}粉丝占比"
    return "目标地区粉丝占比"


def _selection_candidate(row: SessionKolSelection, rank: int) -> ExportCandidate:
    fields = row.fields_json or {}
    score = row.score_json or {}
    dimensions = score.get("dimensions") or {}
    # 模板 8 个评分列由 6 个评分维度的 raw_score 映射而来（沿用旧 pipeline 口径）。
    scores = {
        "industry_interest": _raw_score(dimensions, "content"),
        "target_region": _raw_score(dimensions, "audience"),
        "target_age": _raw_score(dimensions, "audience"),
        "engagement": _raw_score(dimensions, "engagement"),
        "active_follower": _raw_score(dimensions, "growth"),
        "content": _raw_score(dimensions, "content"),
        "followers": _raw_score(dimensions, "audience"),
        "engagement_follower_ratio": _raw_score(dimensions, "engagement"),
    }
    raw_values = fields.get("export_fields", {})
    values = dict(raw_values) if isinstance(raw_values, dict) else {}
    values.setdefault("engagement_rate", _percentage_text(fields.get("engagement_rate")))
    values.setdefault("content_tags", "数据缺失")
    total = score.get("total")
    total_score = float(total) if total is not None else None
    rating_text = score.get("rating")
    stars = score.get("stars")
    if not rating_text or not stars:
        rating_text, stars = rating(total_score or 0.0)
    return ExportCandidate(
        rank=rank,
        platform=row.platform,
        nickname=row.nickname or "未命名达人",
        profile_url=row.profile_url,
        followers=row.followers,
        city=row.city or values.get("city"),
        total_score=total_score,
        rating=rating_text,
        stars=stars,
        dimension_scores=scores,
        values=values,
        score_reason=score_reason(fields),
        source_names=("已授权数据服务",),
        collected_at=row.updated_at.isoformat() if row.updated_at else None,
    )


def _raw_score(dimensions: dict[str, Any], name: str) -> float | None:
    value = dimensions.get(name)
    return value.get("raw_score") if isinstance(value, dict) else None


def _percentage_text(value: Any) -> str:
    if value is None:
        return "数据缺失"
    return f"{float(value):.2f}%"


def _safe_filename(filename: str) -> str:
    return "".join("_" if char in '\\/:*?"<>|' else char for char in filename)


def render_workbook(*, metadata: dict[str, Any], candidates: Sequence[ExportCandidate]) -> bytes:
    workbook = load_workbook(TEMPLATE_PATH)
    summary = workbook["小红书KOL匹配度筛选"]
    summary.title = "KOL匹配度筛选"
    _render_summary(summary, metadata, candidates)
    _render_detail_standard_blocks(workbook["达人详细画像"], metadata, candidates)
    _render_fan_profile(workbook["粉丝画像详情"], metadata, candidates)
    _render_methodology_preserving_template(workbook["评分方法论与数据来源"], metadata, candidates)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _render_summary(sheet: Any, metadata: dict[str, Any], candidates: Sequence[ExportCandidate]) -> None:
    _ensure_summary_merge(sheet, "A1:Q1")
    _ensure_summary_merge(sheet, "A2:Q2")
    # Candidate rows are unbounded. Move the rating block below them so that a
    # large latest pool never writes into the template's A20:Q20 merge.
    rating_title_row = max(20, 5 + len(candidates) + 2)
    for merged in list(sheet.merged_cells.ranges):
        if str(merged) in {"A20:P20", "A20:Q20"}:
            sheet.unmerge_cells(str(merged))
    sheet.merge_cells(f"A{rating_title_row}:Q{rating_title_row}")
    sheet["A1"] = _cell_value(
        f"{metadata.get('brand') or 'KOL'}{metadata.get('category') or ''}达人 — KOL匹配度筛选分析报告"
    )
    locations = "、".join(str(item) for item in metadata.get("locations", [])) or "未指定"
    platforms = list(dict.fromkeys(_platform_label(item.platform) for item in candidates))
    sheet["A2"] = (
        f"客户: {locations} | 报告日期: {str(metadata.get('generated_at') or '')[:10]} | "
        f"达人总数: {len(candidates)} | 平台: {'、'.join(platforms) or '未指定'}"
    )
    headers = _summary_headers(metadata)
    clear_end = max(sheet.max_row, rating_title_row + 7, 40)
    _clear_rows(sheet, 4, clear_end, 19)
    _write_styled_row(sheet, 4, headers, source_row=4)
    for row_index, candidate in enumerate(candidates, start=5):
        values = _summary_values(candidate)
        _write_styled_row(sheet, row_index, values, source_row=5)
        _apply_score_fill(sheet.cell(row_index, 15), candidate.total_score)
        _apply_rating_fill(sheet.cell(row_index, 16), candidate.rating)
        sheet.cell(row_index, 17).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[row_index].height = 42

    sheet.cell(rating_title_row, 1).value = "评级分布汇总"
    _render_rating_summary(sheet, candidates, start_row=rating_title_row + 1)
    _move_summary_charts(sheet, rating_title_row + 7)
    _set_widths(sheet, {"A": 7, "B": 12, "C": 18, "D": 11, "E": 12, "F": 12, "G": 14, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14, "O": 12, "P": 13, "Q": 48})


def _summary_headers(metadata: dict[str, Any]) -> list[str]:
    category = metadata.get("category") or "行业"
    region = "、".join(str(item) for item in metadata.get("locations", [])) or "目标地区"
    age = metadata.get("target_audience") or "目标年龄段"
    return [
        "序号", "平台", "昵称", "评级(★)", "粉丝数", "城市",
        f"{category}兴趣分\n(满分20)", f"{region}粉丝分\n(满分15)",
        f"{age}分\n(满分15)", "互动率分\n(满分15)", "活跃粉丝分\n(满分10)",
        "内容标签分\n(满分10)", "粉丝规模分\n(满分10)", "互动沉淀与粉丝比分\n(满分5)",
        "综合评分", "匹配评估", "评分理由",
    ]


def _summary_values(candidate: ExportCandidate) -> list[Any]:
    scores = candidate.dimension_scores
    present = _present
    return [
        candidate.rank,
        _platform_label(candidate.platform),
        candidate.nickname,
        candidate.stars,
        present(candidate.followers),
        present(candidate.city),
        present(scores.get("industry_interest")),
        present(scores.get("target_region")),
        present(scores.get("target_age")),
        present(scores.get("engagement")),
        present(scores.get("active_follower")),
        present(scores.get("content")),
        present(scores.get("followers")),
        present(scores.get("engagement_follower_ratio")),
        present(candidate.total_score),
        present(candidate.rating),
        present(candidate.score_reason),
    ]


def _render_rating_summary(sheet: Any, candidates: Sequence[ExportCandidate], *, start_row: int = 21) -> None:
    _write_styled_row(sheet, start_row, ["评级", "星级", "分数区间", "达人数量", "占比"], source_row=21)
    buckets = (("重点推荐", "★★★★★", "≥78", 78, 101), ("推荐", "★★★★", "62-77", 62, 78), ("可考虑", "★★★", "48-61", 48, 62), ("观察", "★★", "<48", -1, 48))
    total = len(candidates)
    for bucket_index, (label, stars, interval, lower, upper) in enumerate(buckets):
        row = start_row + 1 + bucket_index
        count = sum(1 for item in candidates if item.total_score is not None and lower <= item.total_score < upper)
        ratio = count / total if total else 0
        # Rows 22:26 are the template's rating styles. Reuse them even
        # when the rating table is moved below an oversized candidate pool.
        _write_styled_row(sheet, row, [label, stars, interval, count, ratio], source_row=22 + bucket_index)
        sheet.cell(row, 5).number_format = "0.0%"
        _apply_rating_fill(sheet.cell(row, 1), label)
    sheet.cell(start_row, 18).value = "评级"
    sheet.cell(start_row, 19).value = "数量"
    for row, (label, _stars, _interval, _lower, _upper) in enumerate(buckets, start=start_row):
        sheet.cell(row, 18).value = label
        sheet.cell(row, 19).value = sum(1 for item in candidates if item.rating == label)
    for chart in getattr(sheet, "_charts", ()):
        for series in getattr(chart, "ser", ()):
            if getattr(getattr(series, "cat", None), "numRef", None) is not None:
                series.cat.numRef.f = f"'{sheet.title}'!$R${start_row}:$R${start_row + 3}"
            if getattr(getattr(series, "val", None), "numRef", None) is not None:
                series.val.numRef.f = f"'{sheet.title}'!$S${start_row}:$S${start_row + 3}"


def _render_detail_standard_blocks(sheet: Any, metadata: dict[str, Any], candidates: Sequence[ExportCandidate]) -> None:
    """Render every detail page as an independent copy of the first 31-row block.

    The attachment contains legacy blocks whose starting rows are not uniform.
    Keeping those merges while appending fixed-size blocks lets a merge from a
    later legacy block swallow cells in the current block.  We therefore take
    the first block's styles as the canonical template, remove every existing
    merge on this sheet, and then rebuild the same merge topology for each
    candidate (including dynamically appended candidates).
    """
    block_size = 31
    base_styles = {
        (offset, column): copy(sheet.cell(1 + offset, column)._style)
        for offset in range(block_size)
        for column in range(1, 7)
    }
    base_heights = {offset: sheet.row_dimensions[1 + offset].height for offset in range(block_size)}

    for merged in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged))
    for row in sheet.iter_rows():
        for cell in row:
            cell.value = None
    known_merges: set[str] = set()

    _set_widths(sheet, {"A": 24, "B": 36, "C": 22, "D": 22, "E": 22, "F": 22})
    for index, candidate in enumerate(candidates):
        start = 1 + index * block_size
        for offset in range(block_size):
            row = start + offset
            # Assign even ``None`` so historical explicit heights on the
            # attachment's later blocks cannot leak into the standard block.
            sheet.row_dimensions[row].height = base_heights[offset]
            for column in range(1, 7):
                sheet.cell(row, column)._style = copy(base_styles[(offset, column)])

        title = (
            f"#{candidate.rank} {_platform_label(candidate.platform)} {candidate.nickname} — "
            f"{candidate.rating} {candidate.stars} (综合评分: {_display(candidate.total_score)}；"
            f"公开主页：{candidate.profile_url or '数据缺失'})"
        )
        _merge_range_if_missing(sheet, known_merges, start, 1, 6)
        _write_detail_header(sheet, start, title)
        sections = [
            (1, "【达人概况】", (("城市", candidate.city), ("粉丝数", candidate.followers), ("总赞藏", candidate.values.get("total_likes")), ("赞藏/粉丝比", candidate.values.get("likes_followers_ratio")), ("内容标签", candidate.values.get("content_tags")), ("性别", candidate.values.get("gender")))),
            (8, "【帖子表现】", (("平均阅读", candidate.values.get("average_reads")), ("平均互动", candidate.values.get("average_interactions")), ("互动率", candidate.values.get("engagement_rate")))),
            (12, "【粉丝画像】", (("<18岁", candidate.values.get("age_under_18")), ("18-24岁", candidate.values.get("age_18_24")), ("25-34岁", candidate.values.get("age_25_34")), ("35-44岁", candidate.values.get("age_35_44")), (">44岁", candidate.values.get("age_over_44")), (_target_region_rate_label(metadata), candidate.values.get("target_region_rate")), ("活跃粉丝率", candidate.values.get("active_follower_rate")), ("兴趣Top标签", candidate.values.get("content_tags")), ("省份", candidate.values.get("province")))),
            (22, "【综合评估】", (("综合评分", candidate.total_score), ("星级", candidate.stars), ("评级", candidate.rating), ("评分明细", _score_detail(candidate.dimension_scores)), ("评估理由", candidate.score_reason))),
            (28, "【综合概述】", (("报告摘要", candidate.values.get("summary")),)),
        ]
        for section_offset, section, entries in sections:
            section_row = start + section_offset
            _merge_range_if_missing(sheet, known_merges, section_row, 1, 6)
            section_cell = sheet.cell(section_row, 1)
            section_cell.value = section
            section_cell.font = Font(name="微软雅黑", bold=True, color="1F4E79", size=12)
            for entry_offset, (label, value) in enumerate(entries, start=section_offset + 1):
                row = start + entry_offset
                rendered = _cell_value(_present(value))
                if section == "【综合概述】":
                    _merge_range_if_missing(sheet, known_merges, row, 1, 6)
                    sheet.cell(row, 1).value = f"{label}：{rendered}"
                else:
                    _merge_range_if_missing(sheet, known_merges, row, 2, 6)
                    sheet.cell(row, 1).value = label
                    sheet.cell(row, 2).value = rendered
                    sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
                sheet.cell(row, 1).font = Font(name="微软雅黑", bold=True, size=10)
        # Offset 30 is intentionally left blank between independent blocks.


def _render_fan_profile(sheet: Any, metadata: dict[str, Any], candidates: Sequence[ExportCandidate]) -> None:
    _clear_sheet(sheet)
    headers = ["序号", "平台", "昵称", "粉丝数", "<18岁%", "18-24岁%", "25-34岁%", "35-44岁%", ">44岁%", "目标地区粉丝%", "活跃粉丝%", f"{metadata.get('category') or '行业'}兴趣%", "综合评分"]
    _write_styled_row(sheet, 1, headers, source_row=1)
    for row, candidate in enumerate(candidates, start=2):
        values = [candidate.rank, _platform_label(candidate.platform), candidate.nickname, _present(candidate.followers)]
        values.extend(_present(candidate.values.get(key, "数据缺失")) for key in ("age_under_18", "age_18_24", "age_25_34", "age_35_44", "age_over_44", "target_region_rate", "active_follower_rate", "industry_interest_rate"))
        values.append(_present(candidate.total_score))
        _write_styled_row(sheet, row, values, source_row=2)
        _apply_score_fill(sheet.cell(row, 13), candidate.total_score)
    _set_widths(sheet, {"A": 7, "B": 12, "C": 18, "D": 12, "E": 11, "F": 11, "G": 11, "H": 11, "I": 11, "J": 14, "K": 14, "L": 14, "M": 12})


def _render_methodology_preserving_template(
    sheet: Any, metadata: dict[str, Any], candidates: Sequence[ExportCandidate]
) -> None:
    """Populate the original methodology layout without destroying its merges."""
    _clear_sheet(sheet)
    known_merges = {str(merged) for merged in sheet.merged_cells.ranges}
    _merge_range_if_missing(sheet, known_merges, 1, 1, 4)
    sheet["A1"] = "评分方法论与数据来源"
    sheet["A1"].font = Font(name="微软雅黑", bold=True, size=16, color="1F4E79")
    dimensions = [
        (f"{metadata.get('category') or '行业'}兴趣占比", 20, "按 MCP 返回的行业兴趣占比评分，上限20", "行业兴趣是本轮投放匹配的核心指标"),
        ("目标地区粉丝占比", 15, "按用户指定地区占比评分，上限15", "地区由本轮会话筛选条件确定"),
        ("目标年龄段占比", 15, "按 MCP 返回的年龄分布评分，上限15", "年龄分桶按 MCP 实际返回口径聚合"),
        ("互动率", 15, "按平台规范化互动率评分，上限15", "平台口径在采集结果中保留"),
        ("活跃粉丝率", 10, "按活跃粉丝比例评分，上限10", "缺失时标记数据缺失"),
        ("内容标签匹配", 10, "按行业和提问中的内容要求评估，上限10", "不得编造未返回标签"),
        ("粉丝规模", 10, "按粉丝规模评分，上限10", "粉丝数为规范化数值"),
        ("互动沉淀与粉丝比", 5, "按各平台可获得的赞、藏、评等指标计算，上限5", "抖音和小红书采用各自平台口径"),
    ]
    _merge_range_if_missing(sheet, known_merges, 3, 1, 4)
    sheet["A3"] = "一、评分维度与计算方式"
    _write_styled_row(sheet, 4, ["维度", "满分", "计算方式", "说明"], source_row=4)
    for row, values in enumerate(dimensions, start=5):
        _write_styled_row(sheet, row, values, source_row=row)

    _merge_range_if_missing(sheet, known_merges, 15, 1, 4)
    sheet["A15"] = "二、评级映射"
    _write_styled_row(sheet, 16, ["评级", "星级", "分数区间", "建议"], source_row=16)
    ratings = [
        ("重点推荐", "★★★★★", "≥78", "优先合作，匹配度极高"),
        ("推荐", "★★★★", "62-77", "建议合作，匹配度良好"),
        ("可考虑", "★★★", "48-61", "可考虑合作，需关注短板"),
        ("观察", "★★", "<48", "匹配度偏低，保持观察"),
    ]
    for row, values in enumerate(ratings, start=17):
        _write_styled_row(sheet, row, values, source_row=row)
        _apply_rating_fill(sheet.cell(row, 1), values[0])

    _merge_range_if_missing(sheet, known_merges, 24, 1, 4)
    sheet["A24"] = "三、数据来源说明"
    source_rows = [
        f"1. 本轮候选总数：{len(candidates)}。",
        "2. 来源服务：{}。".format("、".join(sorted({name for item in candidates for name in item.source_names})) or "数据来源未标注"),
        "3. 缺失字段统一显示“数据缺失”，评分理由会标注按规则处理的字段。",
        "4. 本工作簿不包含内部 ID、MCP 调用 ID、密钥、接口地址或原始响应。",
        "5. 多平台候选按最新任务候选池的综合评分排序稳定导出。",
        "6. 主页链接仅保留达人公开主页，字段缺失时显示“数据缺失”。",
        f"7. 报告生成日期：{str(metadata.get('generated_at') or '')[:10]}。",
    ]
    for row, text in enumerate(source_rows, start=25):
        _merge_range_if_missing(sheet, known_merges, row, 1, 4)
        sheet.cell(row, 1).value = text
        sheet.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    _set_widths(sheet, {"A": 24, "B": 18, "C": 48, "D": 48})


def _write_detail_header(sheet: Any, row: int, title: str) -> None:
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = sheet.cell(row, 1, title)
    cell.font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center")


def _write_styled_row(sheet: Any, row: int, values: Sequence[Any], *, source_row: int) -> None:
    # sheet.max_column 在 openpyxl 中是 O(全表单元格数)，热循环外取一次。
    max_column = max(1, sheet.max_column)
    for index, value in enumerate(values, start=1):
        target = sheet.cell(row, index)
        # openpyxl represents every non-anchor cell of a merged range as a
        # read-only MergedCell. Dynamic rows can overlap template merges, so
        # leave those cells untouched rather than assigning to them.
        if isinstance(target, MergedCell):
            continue
        source = sheet.cell(source_row, min(index, max_column))
        if source.has_style:
            target._style = copy(source._style)
        target.value = _cell_value(value)
        target.alignment = copy(source.alignment) if source.has_style else Alignment(vertical="top")
        if isinstance(value, float) and (0 <= value <= 1):
            target.number_format = "0.0%"
        elif isinstance(value, float):
            target.number_format = "0.0"


def _clear_rows(sheet: Any, start: int, end: int, columns: int) -> None:
    for row in range(start, end + 1):
        for column in range(1, columns + 1):
            cell = sheet.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _clear_sheet(sheet: Any) -> None:
    """Clear values while preserving the template's merged-cell topology."""
    for row in sheet.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.value = None


def _merge_range_if_missing(
    sheet: Any, known: set[str], row: int, start_column: int, end_column: int
) -> None:
    """按本地集合判重后合并单元格，避免每次线性扫描 merged_cells.ranges。"""
    reference = f"{get_column_letter(start_column)}{row}:{get_column_letter(end_column)}{row}"
    if reference not in known:
        sheet.merge_cells(reference)
        known.add(reference)


def _ensure_summary_merge(sheet: Any, reference: str) -> None:
    for merged in list(sheet.merged_cells.ranges):
        if str(merged).startswith(reference[:2]):
            sheet.unmerge_cells(str(merged))
    sheet.merge_cells(reference)


def _move_summary_charts(sheet: Any, row: int) -> None:
    """Keep the template's rating chart below dynamically-sized candidate rows."""
    for chart in getattr(sheet, "_charts", ()):
        anchor = getattr(chart, "anchor", None)
        source = getattr(anchor, "_from", None)
        if source is not None:
            source.row = max(0, row - 1)


def _set_widths(sheet: Any, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def _apply_score_fill(cell: Any, score: float | None) -> None:
    if score is None:
        return
    color = "63BE7B" if score >= 62 else "FFEB84" if score >= 48 else "F4B183" if score >= 35 else "F8696B"
    cell.fill = PatternFill("solid", fgColor=color)


def _apply_rating_fill(cell: Any, rating: str) -> None:
    colors = {"重点推荐": "318B32", "推荐": "63BE7B", "可考虑": "FFDD35", "观察": "FFEB84"}
    cell.fill = PatternFill("solid", fgColor=colors.get(rating, "FFFFFF"))


def _score_detail(scores: dict[str, float | None]) -> str:
    labels = {
        "industry_interest": "行业兴趣",
        "target_region": "目标地区",
        "target_age": "目标年龄",
        "engagement": "互动表现",
        "active_follower": "活跃粉丝",
        "content": "内容匹配",
        "followers": "粉丝规模",
        "engagement_follower_ratio": "互动沉淀与粉丝比",
    }
    return " + ".join(f"{labels.get(key, key)}:{_display(value)}" for key, value in scores.items()) or "数据缺失"


def _display(value: Any) -> str:
    if value is None:
        return "数据缺失"
    if isinstance(value, float):
        return f"{value:.1f}"
    if isinstance(value, (list, tuple, set)):
        return "、".join(_display(item) for item in value) if value else "数据缺失"
    if isinstance(value, dict):
        return "；".join(f"{key}: {_display(item)}" for key, item in value.items()) or "数据缺失"
    return f"{value:,}" if isinstance(value, int) else str(value)


_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _cell_value(value: Any) -> Any:
    """将结构化 MCP 字段转换为 Excel 可写的标量，同时保留数字类型。

    第三方可控文本（昵称、标签等）以 ``=``、``+``、``-``、``@`` 开头时
    openpyxl/Excel 会按公式解析，前缀 ``'`` 转义防止公式注入。
    """
    rendered = _display(value) if isinstance(value, (list, tuple, set, dict)) else value
    if isinstance(rendered, str) and rendered.startswith(_FORMULA_PREFIXES):
        return f"'{rendered}"
    return rendered


def _present(value: Any) -> Any:
    """将缺失业务字段显式标记，避免 Excel 中出现难以区分的空白。"""
    if value is None or (isinstance(value, str) and not value.strip()):
        return "数据缺失"
    return value


def _platform_label(platform: str) -> str:
    return {"xiaohongshu": "小红书", "douyin": "抖音", "bilibili": "哔哩哔哩", "weibo": "微博", "wechat": "微信"}.get(platform, platform)


__all__ = [
    "CONTENT_TYPE",
    "ExportCandidate",
    "ExportedWorkbook",
    "export_session_selection",
    "render_workbook",
]
