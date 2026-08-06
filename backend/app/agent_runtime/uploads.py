"""安全上传与解析服务（Gate B Task 3）。

用户 CSV/XLSX 上传：扩展名白名单（.csv/.xlsx，拒绝 .xlsm 等宏格式）、
大小上限（默认 20 MiB）、数据行上限（默认 50,000）；SHA-256 作为不可变
文件名的一部分，路径只由服务生成（绝不拼接用户文件名）。CSV 用
``utf-8-sig``，XLSX 用 ``openpyxl.read_only + data_only + keep_links=False``
（不执行宏/公式）。解析在线程执行；结构化 rows 作为 upload Evidence
（``tool_call_id`` 为 NULL、``upload_id`` 有值，XOR 由 DB 约束保证）。

原始上传不可变：重新解析产生新 Evidence，不覆盖来源（Evidence 层
append-only 语义）。
"""

from __future__ import annotations

import asyncio
import csv
import hashlib
import io
import logging
import re
import xml.etree.ElementTree as ElementTree
import zipfile
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.agent_runtime.evidence import EvidenceWriter
from app.agent_runtime.models import AgentUpload
from app.core.config import get_settings

logger = logging.getLogger(__name__)

_ALLOWED_EXTENSIONS = (".csv", ".xlsx")
_MIME_BY_EXTENSION = {
    ".csv": "text/csv",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

# upload Evidence preview 行上限（raw payload 不受影响，仍完整落库）。
_PREVIEW_ROW_CAP = 5000

# XLSX ZIP 安全检查阈值（Gate B P1：集中常量，不散落魔法数字）。
_MAX_ZIP_ENTRIES = 2000
_MAX_TOTAL_UNCOMPRESSED = 100 * 1024 * 1024  # 100 MiB
_MAX_SINGLE_ENTRY = 50 * 1024 * 1024  # 50 MiB
_MAX_COMPRESSION_RATIO = 100
# 工作表物理维度上限（Gate B P1-4）：恶意 XLSX 可声明
# <dimension ref="A1:XFD1048576"/> 或把下一条 row 的 r 设为极大值，让 openpyxl
# 为中间空行持续迭代。列数上限 1000；行数上限由 max_rows+1（含表头）决定。
_MAX_XLSX_COLUMNS = 1000


class UploadRejectedError(Exception):
    """上传被拒绝（大小/格式/行数）；携带 HTTP 状态码与结构化错误码。"""

    def __init__(self, *, status_code: int, error_code: str, message: str) -> None:
        super().__init__(message)
        self.status_code = status_code
        self.error_code = error_code


def _now() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)


def _parse_csv(content: bytes, *, max_rows: int) -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict[str, Any]] = []
    for row in reader:
        if not any(v not in (None, "") for v in row.values()):
            continue
        rows.append({k: v for k, v in row.items() if k is not None})
        if len(rows) > max_rows:
            break
    return (reader.fieldnames or []), rows


@dataclass(frozen=True)
class WorksheetBounds:
    """活动 worksheet 的真实物理边界（来自 row/cell 坐标预检，dimension 不可信）。"""

    physical_rows: int
    max_row: int
    max_column: int


# cell 引用严格格式：列字母 + 行号（行号 >= 1、无前导零）。
_CELL_REF_RE = re.compile(r"([A-Za-z]+)([1-9]\d*)\Z")
_ROW_R_RE = re.compile(r"[1-9]\d*\Z")


def _column_index_from_ref(ref: str) -> tuple[int, int]:
    """把单元格引用（如 B2/XFD1048576）解析为 (列号, 行号)；A=1、行号 >= 1。

    非法引用（A0/A01/B2X/缺列或缺行号）抛受控错误；行号只做格式与正数校验，
    上限由调用方按 max_rows 检查。
    """
    match = _CELL_REF_RE.fullmatch(ref)
    if match is None:
        raise UploadRejectedError(
            status_code=400, error_code="invalid_worksheet_coordinate",
            message=f"invalid cell reference: {ref!r}",
        )
    letters, row_str = match.groups()
    column = 0
    for ch in letters:
        column = column * 26 + (ord(ch.upper()) - ord("A") + 1)
    return column, int(row_str)


def _resolve_active_worksheet_path(zf: zipfile.ZipFile, workbook: Any) -> str:
    """返回活动 worksheet 在 ZIP 内的路径；路径缺失/越界返回受控错误。

    使用 openpyxl read_only 提供的 worksheet path（``_worksheet_path``），并校验
    它存在于包内且属于 ``xl/worksheets/``——不让外部路径进入 ZIP 读取。
    """
    sheet = workbook.active
    raw = getattr(sheet, "_worksheet_path", None)
    if not isinstance(raw, str) or not raw:
        raise UploadRejectedError(
            status_code=400, error_code="invalid_worksheet_coordinate",
            message="cannot resolve active worksheet path",
        )
    path = raw.lstrip("/")
    if not path.casefold().startswith("xl/worksheets/"):
        raise UploadRejectedError(
            status_code=400, error_code="invalid_worksheet_coordinate",
            message="active worksheet path outside xl/worksheets/",
        )
    names = {name.casefold() for name in zf.namelist()}
    if path.casefold() not in names:
        raise UploadRejectedError(
            status_code=400, error_code="invalid_worksheet_coordinate",
            message="active worksheet path missing in package",
        )
    return path


def _scan_worksheet_coordinates(zf: zipfile.ZipFile, path: str, *, max_rows: int) -> WorksheetBounds:
    """流式扫描活动 worksheet XML 的真实坐标（禁止 zf.read 整包加载）。

    start 事件中统计 row 元素数量、最大行坐标（row r 与 cell 引用中的行号）与
    最大列号；省略 ``c@r`` 的合法单元格按当前 row 内顺序推导隐式列号。XML 解析
    错误与非法坐标一律映射为受控 UploadRejectedError；物理行数超过上限时提前
    中止，避免恶意文件触发整表扫描。
    """
    physical_rows = 0
    max_row_coordinate = 0
    actual_max_column = 0
    cell_index = 0
    try:
        with zf.open(path) as stream:
            for _event, element in ElementTree.iterparse(stream, events=("start",)):
                local = element.tag.rsplit("}", 1)[-1]
                if local == "row":
                    cell_index = 0
                    physical_rows += 1
                    if physical_rows > max_rows + 1:
                        raise UploadRejectedError(
                            status_code=400, error_code="worksheet_dimensions_exceeded",
                            message=f"worksheet has more than {max_rows + 1} rows",
                        )
                    raw_r = element.attrib.get("r")
                    if raw_r is not None:
                        if _ROW_R_RE.fullmatch(raw_r) is None:
                            raise UploadRejectedError(
                                status_code=400, error_code="invalid_worksheet_coordinate",
                                message=f"invalid row coordinate: {raw_r!r}",
                            )
                        row_no = int(raw_r)
                        if row_no > max_row_coordinate:
                            max_row_coordinate = row_no
                elif local == "c":
                    cell_index += 1
                    ref = element.attrib.get("r")
                    if ref:
                        column, cell_row = _column_index_from_ref(ref)
                        if cell_row > max_row_coordinate:
                            max_row_coordinate = cell_row
                    else:
                        # 合法省略 c@r：按当前 row 内单元格顺序推导隐式列号。
                        column = cell_index
                    if column > actual_max_column:
                        actual_max_column = column
                element.clear()
    except UploadRejectedError:
        raise
    except ElementTree.ParseError:
        raise UploadRejectedError(
            status_code=400, error_code="invalid_worksheet_coordinate",
            message="worksheet XML is not well-formed",
        )
    return WorksheetBounds(
        physical_rows=physical_rows,
        max_row=max_row_coordinate,
        max_column=actual_max_column,
    )


def _validate_worksheet_bounds(bounds: WorksheetBounds, *, max_rows: int) -> None:
    """按真实坐标校验物理边界：行数/行坐标/列坐标任一超限立即拒绝。"""
    if bounds.physical_rows > max_rows + 1:
        raise UploadRejectedError(
            status_code=400, error_code="worksheet_dimensions_exceeded",
            message=f"worksheet has more than {max_rows + 1} rows",
        )
    if bounds.max_row > max_rows + 1:
        raise UploadRejectedError(
            status_code=400, error_code="worksheet_dimensions_exceeded",
            message=f"worksheet max row {bounds.max_row} exceeds {max_rows + 1}",
        )
    if bounds.max_column > _MAX_XLSX_COLUMNS:
        raise UploadRejectedError(
            status_code=400, error_code="worksheet_columns_exceeded",
            message=f"worksheet max column {bounds.max_column} exceeds {_MAX_XLSX_COLUMNS}",
        )


def _parse_xlsx(content: bytes, *, max_rows: int) -> tuple[list[str], list[dict[str, Any]]]:
    from openpyxl import load_workbook

    # 拒绝含宏部件或外部链接的 xlsx（即使扩展名不是 .xlsm），并做 ZIP bomb 防护。
    with zipfile.ZipFile(io.BytesIO(content)) as zf:
        infos = zf.infolist()
        if len(infos) > _MAX_ZIP_ENTRIES:
            raise UploadRejectedError(
                status_code=400, error_code="zip_too_many_entries",
                message=f"zip exceeds {_MAX_ZIP_ENTRIES} entries",
            )
        total_uncompressed = 0
        for info in infos:
            name = info.filename
            # ZIP 部件名比较一律 casefold（宏部件/外部链接目录名大小写不敏感）。
            folded = name.casefold()
            if "vbaproject" in folded:
                raise UploadRejectedError(
                    status_code=415, error_code="macro_detected", message="file contains macro parts"
                )
            if folded.startswith("xl/externallinks/"):
                raise UploadRejectedError(
                    status_code=415, error_code="external_links_detected",
                    message="file contains external links",
                )
            if info.file_size > _MAX_SINGLE_ENTRY:
                raise UploadRejectedError(
                    status_code=400, error_code="zip_entry_too_large",
                    message=f"single zip entry exceeds {_MAX_SINGLE_ENTRY} bytes",
                )
            if info.compress_size == 0 and info.file_size > 0:
                raise UploadRejectedError(
                    status_code=400, error_code="zip_suspicious_ratio",
                    message="zip entry has zero compressed size with non-empty content",
                )
            if info.file_size > 0 and info.compress_size > 0:
                ratio = info.file_size / info.compress_size
                if ratio > _MAX_COMPRESSION_RATIO:
                    raise UploadRejectedError(
                        status_code=400, error_code="zip_compression_ratio",
                        message=f"zip compression ratio exceeds {_MAX_COMPRESSION_RATIO}",
                    )
            total_uncompressed += info.file_size
            if total_uncompressed > _MAX_TOTAL_UNCOMPRESSED:
                raise UploadRejectedError(
                    status_code=400, error_code="zip_total_too_large",
                    message=f"total uncompressed size exceeds {_MAX_TOTAL_UNCOMPRESSED} bytes",
                )
        # 所有 .rels 用 XML 解析，仅当 Relationship 的 TargetMode == "external"
        # 才拒绝。xl/worksheets/_rels/ 下的 worksheet→drawing/chart 等是合法
        # 内部关系，必须接受；字符串搜索（如 'TargetMode="External"'）会被单
        # 引号、属性顺序与空白绕过，故用 XML 解析遍历 Relationship 节点。
        for info in infos:
            if not info.filename.casefold().endswith(".rels"):
                continue
            try:
                rels_bytes = zf.read(info.filename)
            except KeyError:
                continue
            try:
                root = ElementTree.fromstring(rels_bytes)
            except ElementTree.ParseError:
                raise UploadRejectedError(
                    status_code=415, error_code="invalid_relationships_xml",
                    message="file contains invalid relationships XML",
                )
            for relationship in root.iter():
                # Relationship 元素带默认命名空间（tag 形如 {ns}Relationship），
                # 按本地名匹配，避免命名空间差异导致漏检。
                if relationship.tag.rsplit("}", 1)[-1] != "Relationship":
                    continue
                target_mode = relationship.attrib.get("TargetMode", "").casefold()
                if target_mode == "external":
                    raise UploadRejectedError(
                        status_code=415, error_code="external_links_detected",
                        message="file contains external relationships",
                    )
    workbook = load_workbook(
        io.BytesIO(content),
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    try:
        sheet = workbook.active
        # 物理维度快速拒绝（保留）：声明的 dimension 超限立即拒绝，但 dimension
        # 不再作为实际解析宽度的唯一来源（可缺失/可伪造）。
        if sheet.max_row is not None and sheet.max_row > max_rows + 1:
            raise UploadRejectedError(
                status_code=400, error_code="worksheet_dimensions_exceeded",
                message=f"worksheet declares more than {max_rows + 1} rows",
            )
        if sheet.max_column is not None and sheet.max_column > _MAX_XLSX_COLUMNS:
            raise UploadRejectedError(
                status_code=400, error_code="worksheet_columns_exceeded",
                message=f"worksheet exceeds {_MAX_XLSX_COLUMNS} columns",
            )
        # 坐标预检：真实 row/cell 坐标确定物理边界与实际宽度（流式 iterparse，
        # 不整包加载 worksheet XML）。
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            worksheet_path = _resolve_active_worksheet_path(zf, workbook)
            bounds = _scan_worksheet_coordinates(zf, worksheet_path, max_rows=max_rows)
        _validate_worksheet_bounds(bounds, max_rows=max_rows)
        # 解析宽度 = 真实 cell 坐标得到的最大列号（dimension 缺失/比真实小/比真实
        # 大都以实际坐标为准），并确保不超上限。
        effective_max_col = min(max(bounds.max_column, 1), _MAX_XLSX_COLUMNS)
        rows: list[dict[str, Any]] = []
        columns: list[str] = []
        # 显式传递迭代边界，不依赖工作簿自身声明的范围；空白行也计入迭代次数
        # （max_row 上限），不能只统计非空数据行。
        for index, row in enumerate(
            sheet.iter_rows(
                values_only=True,
                max_row=max_rows + 1,
                max_col=effective_max_col,
            )
        ):
            if row is None or all(cell is None for cell in row):
                continue
            if index == 0:
                columns = [str(cell) if cell is not None else "" for cell in row]
                continue
            rows.append(
                {
                    columns[cell_index]: cell
                    for cell_index, cell in enumerate(row)
                    if cell_index < len(columns)
                }
            )
            if len(rows) >= max_rows:
                break
        return columns, rows
    finally:
        workbook.close()


class UploadService:
    """上传落盘 + 解析 + upload Evidence 写入（单事务）。"""

    def __init__(
        self,
        db_session: AsyncSession,
        *,
        storage_dir: str | None = None,
        max_bytes: int | None = None,
        max_rows: int | None = None,
    ) -> None:
        settings = get_settings()
        self._db = db_session
        self._storage_dir = Path(storage_dir or settings.agent_upload_storage_dir)
        self._max_bytes = max_bytes or settings.agent_upload_max_bytes
        self._max_rows = max_rows or settings.agent_upload_max_rows

    @property
    def max_bytes(self) -> int:
        return self._max_bytes

    @property
    def max_rows(self) -> int:
        return self._max_rows

    async def create_and_parse(
        self,
        *,
        user_id: str,
        session_id: str,
        filename: str,
        mime_type: str,
        content: bytes,
    ) -> AgentUpload:
        extension = Path(filename).suffix.lower()
        if extension not in _ALLOWED_EXTENSIONS:
            raise UploadRejectedError(
                status_code=415,
                error_code="unsupported_media_type",
                message=f"only {', '.join(_ALLOWED_EXTENSIONS)} uploads are allowed",
            )
        if len(content) > self._max_bytes:
            raise UploadRejectedError(
                status_code=413,
                error_code="upload_too_large",
                message=f"file exceeds {self._max_bytes} bytes",
            )

        upload_id = str(uuid4())
        sha256 = hashlib.sha256(content).hexdigest()
        # 路径只由服务生成：sha256 片段 + 服务生成的 upload_id，绝不拼接用户文件名。
        storage_key = f"{user_id}/{upload_id[:8]}-{sha256[:16]}{extension}"
        await asyncio.to_thread(self._write_file, storage_key, content)

        upload = AgentUpload(
            id=upload_id,
            user_id=user_id,
            session_id=session_id,
            run_id=None,
            original_filename=filename,
            mime_type=mime_type or _MIME_BY_EXTENSION[extension],
            size_bytes=len(content),
            sha256=sha256,
            storage_key=storage_key,
            status="uploaded",
            created_at=_now(),
        )
        self._db.add(upload)
        await self._db.flush()

        try:
            columns, rows = await asyncio.to_thread(
                self._parse, extension, content, self._max_rows
            )
        except UploadRejectedError as exc:
            await asyncio.to_thread(self._delete_file, storage_key)
            upload.status = "failed"
            upload.error_code = exc.error_code
            upload.completed_at = _now()
            await self._db.flush()
            raise
        except Exception:
            await asyncio.to_thread(self._delete_file, storage_key)
            upload.status = "failed"
            upload.error_code = "parse_failed"
            upload.completed_at = _now()
            await self._db.flush()
            raise UploadRejectedError(
                status_code=400,
                error_code="parse_failed",
                message="file could not be parsed",
            ) from None

        if len(rows) > self._max_rows:
            await asyncio.to_thread(self._delete_file, storage_key)
            upload.status = "failed"
            upload.error_code = "rows_exceeded"
            upload.completed_at = _now()
            await self._db.flush()
            raise UploadRejectedError(
                status_code=400,
                error_code="rows_exceeded",
                message=f"file exceeds {self._max_rows} data rows",
            )

        # 上传列映射诊断：所有列都是用户自定义 schema，不映射 MCP 规范键，
        # 全部列为 unmapped_fields 供模型参考（不误报失败）。
        from app.agent_runtime.normalization import NormalizationResult

        upload_normalization = NormalizationResult(
            version="upload_v1",
            status="not_applicable",
            preview={"columns": columns, "row_count": len(rows)},
            field_mapping={},
            unmapped_fields=tuple(columns),
            truncated=len(rows) > _PREVIEW_ROW_CAP,
        )
        await EvidenceWriter(self._db).write(
            session_id=session_id,
            run_id=None,
            tool_call_id=None,
            upload_id=upload_id,
            source_type="user_upload",
            source_name="user_upload",
            scope_json=None,
            period_json=None,
            raw_payload={
                "columns": columns,
                "rows": rows,
                "truncated": len(rows) > _PREVIEW_ROW_CAP,
            },
            normalization=upload_normalization,
        )
        upload.status = "parsed"
        upload.completed_at = _now()
        await self._db.flush()
        return upload

    def _write_file(self, storage_key: str, content: bytes) -> None:
        path = self._storage_dir / storage_key
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_bytes(content)

    def _delete_file(self, storage_key: str) -> None:
        path = self._storage_dir / storage_key
        try:
            path.unlink(missing_ok=True)
        except OSError:
            logger.warning("failed to delete upload file %s", storage_key, exc_info=True)

    def _parse(self, extension: str, content: bytes, max_rows: int) -> tuple[list[str], list[dict[str, Any]]]:
        if extension == ".csv":
            return _parse_csv(content, max_rows=max_rows)
        return _parse_xlsx(content, max_rows=max_rows)

    async def get_owned(self, *, user_id: str, upload_id: str) -> AgentUpload | None:
        return await self._db.scalar(
            select(AgentUpload).where(
                AgentUpload.id == upload_id,
                AgentUpload.user_id == user_id,
            )
        )


__all__ = [
    "UploadRejectedError",
    "UploadService",
]
