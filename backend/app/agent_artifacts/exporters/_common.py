"""agent_artifacts Excel 导出共享工具（Task 18 / 设计 §10.1 表现层）。

导出器只做展示：从已发布 Version 的 payload 渲染受控模板，绝不调用模型/MCP，
也绝不参与任何业务计算。本模块承载两个导出器共用的取值/格式/受限披露逻辑。
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Sequence

from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

MISSING = "未采集"
PCT_FORMAT = "0.00%"
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_PLATFORM_LABELS = {
    "xiaohongshu": "小红书",
    "douyin": "抖音",
    "bilibili": "哔哩哔哩",
    "weibo": "微博",
    "wechat": "微信",
}

@dataclass(frozen=True)
class TableLayout:
    """write_table 写出的小节版式：供图表引用真实数据行与后续小节级联定位。"""

    title_row: int | None
    header_row: int
    data_start: int
    data_end: int
    next_row: int

    @property
    def has_data(self) -> bool:
        return self.data_end >= self.data_start


_DATA_FONT = Font(name="微软雅黑", size=10)
_DATA_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_ALTERNATE_FILL = PatternFill("solid", fgColor="D6E4F0")
_DATA_ALIGNMENT = Alignment(vertical="center", wrap_text=False)
_DATA_ROW_HEIGHT = 20


def present(value: Any) -> Any:
    """None 或空白文本显式标记，避免 Excel 中难区分的空白。"""
    if value is None or (isinstance(value, str) and not value.strip()):
        return MISSING
    return value


def cell_value(value: Any) -> Any:
    """第三方可控文本以 = + - @ 开头时前缀 ' 转义，防公式注入。"""
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def write_value(cell: Any, value: Any) -> None:
    cell.value = cell_value(present(value))


def write_pct(cell: Any, value: float | None) -> None:
    """占比/变化率直接写 payload 已算数值（小数单位），用 Excel 百分比格式。"""
    if value is None:
        cell.value = MISSING
        cell.number_format = "General"
        return
    cell.value = value
    cell.number_format = PCT_FORMAT


def platform_label(name: str) -> str:
    return _PLATFORM_LABELS.get(name, name)


def clear_rows(sheet: Any, start: int, end: int, columns: int) -> None:
    for row in range(start, end + 1):
        for column in range(1, columns + 1):
            cell = sheet.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None


def clear_rows_unmerged(sheet: Any, start: int, end: int, columns: int) -> None:
    """解除与数据区重叠的合并区域后再清值（模板数据区可能残留合并）。"""
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row >= start and merged.min_row <= end:
            sheet.unmerge_cells(str(merged))
    clear_rows(sheet, start, end, columns)


def write_table(
    sheet: Any,
    start_row: int,
    title: str | None,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    columns: int,
    pct_columns: Sequence[int] = (),
    note: str | None = None,
    alternate_fill: bool = True,
    row_height: int | None = _DATA_ROW_HEIGHT,
) -> TableLayout:
    """写一个 [可选标题 + 表头 + 数据行] 小节，返回 :class:`TableLayout`。

    空数据行保留标题与表头，并在数据区写受限/未采集说明；绝不伪造数据。
    Gate C 审核：动态数据行统一应用 font/border/alignment/交替填充与行高，
    保证 Top20 等长名单样式一致。返回的版式供图表引用真实表头/数据行，
    以及后续小节按 ``next_row`` 级联定位（绝不手算行号造成重叠）。
    """
    row = start_row
    title_row: int | None = None
    if title is not None:
        title_row = row
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
        title_cell = sheet.cell(row, 1)
        title_cell.value = title
        title_cell.font = Font(name="微软雅黑", bold=True, size=11, color="1F4E79")
        row += 1
    header_row = row
    for column, header in enumerate(headers, start=1):
        header_cell = sheet.cell(row, column)
        header_cell.value = header
        header_cell.font = Font(name="微软雅黑", bold=True, size=10)
    row += 1
    data_start = row
    if not rows:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
        sheet.cell(row, 1).value = note or MISSING
        return TableLayout(title_row, header_row, data_start, data_start - 1, row + 1)
    for index, values in enumerate(rows):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column)
            if column in pct_columns:
                if isinstance(value, (int, float)) and not isinstance(value, bool):
                    write_pct(cell, value)
                else:
                    write_pct(cell, None)
            else:
                write_value(cell, value)
            # Gate C 审核：数据行统一样式（字体/边框/对齐/交替填充/行高）。
            cell.font = _DATA_FONT
            cell.border = _DATA_BORDER
            cell.alignment = _DATA_ALIGNMENT
            if alternate_fill and index % 2 == 1:
                cell.fill = _ALTERNATE_FILL
        if row_height is not None:
            sheet.row_dimensions[row].height = row_height
        row += 1
    return TableLayout(title_row, header_row, data_start, row - 1, row)


def _limitation_covers(limitation: Any, section: str) -> bool:
    """一条 limitation 覆盖某章节：affected_paths 为空（通用）或首段等于章节名。"""
    if not limitation.affected_paths:
        return True
    for path in limitation.affected_paths:
        parts = path.split(".")
        if parts and parts[0] == "data":
            parts = parts[1:]
        if parts and parts[0] == section:
            return True
    return False


def section_note(payload: Any, section: str) -> str | None:
    """受限章节 → 「数据受限：...」披露；完整/未受限 → None。"""
    availability = payload.availability.get(section)
    if not availability or availability.status == "complete":
        return None
    reasons: list[str] = [
        limitation.message
        for limitation in payload.limitations
        if _limitation_covers(limitation, section)
    ]
    if not reasons:
        reasons.append(f"该章节数据{availability.status}")
    if availability.reason_codes:
        reasons.append("原因码：" + "、".join(availability.reason_codes))
    return "数据受限：" + "；".join(dict.fromkeys(reasons))


def empty_note(payload: Any, section: str) -> str:
    """空章节说明：受限章节给受限披露，否则「未采集」。"""
    return section_note(payload, section) or MISSING
