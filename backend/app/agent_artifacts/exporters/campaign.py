"""campaign_report_v2 Excel 导出渲染器（Gate C Task 5 / 设计 §8.3）。

按受控模板 ``templates/campaign_report_v2.xlsx`` 渲染 9 个基础 Sheet；具备可靠
ROI 数据时在末尾动态插入第 10 个「ROI与转化」Sheet，否则不生成、不暗示 ROI。

视觉规范：标题 #1F4E79、交替行 #D6E4F0、微软雅黑、统一打印区（模板承载）；
无数据章节保留表头并写受限/未采集说明；只在有可靠数据时创建图表。
导出只读已发布 Version 的 payload，不调用模型/MCP。
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import PatternFill

from app.agent_artifacts.exporters._common import (
    MISSING,
    TableLayout,
    cell_value,
    clear_rows_unmerged,
    empty_note,
    platform_label,
    write_table,
)
from app.agent_artifacts.payloads.campaign import CampaignReportV2

TEMPLATE_PATH = (
    Path(__file__).resolve().parents[1] / "templates" / "campaign_report_v2.xlsx"
)
SHEET_ORDER = (
    "活动综合概览",
    "周期对比与趋势",
    "平台表现",
    "情感与内容分析",
    "热门帖子TOP",
    "达人投放表现",
    "自然传播与受众",
    "洞察与建议",
    "方法论",
)
ROI_SHEET = "ROI与转化"
ALTERNATE_FILL = PatternFill("solid", fgColor="D6E4F0")


def _bar_chart(
    sheet,
    layout: TableLayout,
    *,
    data_cols: tuple[int, ...],
    cat_col: int,
    title: str,
    width: int = 15,
    height: int = 8,
) -> BarChart:
    """按 write_table 返回的真实版式建图：系列名取表头单元格（中文），
    数值/类别区间严格等于数据行（data_start..data_end），绝不手算行号。"""
    chart = BarChart()
    chart.title = title
    for col in data_cols:
        chart.add_data(
            Reference(sheet, min_col=col, min_row=layout.header_row, max_row=layout.data_end),
            titles_from_data=True,
        )
    chart.set_categories(
        Reference(sheet, min_col=cat_col, min_row=layout.data_start, max_row=layout.data_end)
    )
    chart.height = height
    chart.width = width
    return chart


def render_campaign_workbook(payload: dict) -> bytes:
    """把已发布 campaign_report_v2 payload 渲染为 .xlsx bytes（同步 CPU 密集）。"""
    report = CampaignReportV2.model_validate(payload)
    workbook = load_workbook(TEMPLATE_PATH)
    _write_title(workbook["活动综合概览"], f"{report.scope.brand}「{report.scope.campaign}」活动分析报告")
    _write_title(workbook["周期对比与趋势"], "周期对比与趋势")
    _write_title(workbook["平台表现"], "平台表现")
    _write_title(workbook["情感与内容分析"], "情感与内容分析")
    _write_title(workbook["热门帖子TOP"], "热门帖子 TOP")
    _write_title(workbook["达人投放表现"], "达人投放表现")
    _write_title(workbook["自然传播与受众"], "自然传播与受众")
    _write_title(workbook["洞察与建议"], "洞察与建议")
    _write_title(workbook["方法论"], "数据说明与方法论")
    _render_overview(workbook["活动综合概览"], report)
    _render_comparisons(workbook["周期对比与趋势"], report)
    _render_platforms(workbook["平台表现"], report)
    _render_sentiment_content(workbook["情感与内容分析"], report)
    _render_top_posts(workbook["热门帖子TOP"], report)
    _render_kols(workbook["达人投放表现"], report)
    _render_organic_audience(workbook["自然传播与受众"], report)
    _render_insights(workbook["洞察与建议"], report)
    _render_methodology(workbook["方法论"], report)
    # ROI Sheet 来自受控模板（带视觉结构）；无 ROI 时移除，不暗示。
    roi_sheet = workbook[ROI_SHEET]
    if report.data.roi is None:
        workbook.remove(roi_sheet)
    else:
        _render_roi(roi_sheet, report)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_title(sheet, text: str) -> None:
    sheet["A1"] = cell_value(text)


def _sheet_headers(sheet, headers: list[str], row: int = 3, columns: int = 8) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row, column)
        cell.value = header
        cell.fill = PatternFill("solid", fgColor="D9E2F3")


def _render_overview(sheet, report) -> None:
    scope = report.scope
    sheet["A1"] = cell_value(f"{scope.brand}「{scope.campaign}」活动分析报告")
    clear_rows_unmerged(sheet, 2, 60, 8)
    overview = report.data.overview
    rows = [
        ["总声量", overview.total_volume],
        ["总互动", overview.total_engagement],
        ["总帖数", overview.total_posts],
        ["达人数量", overview.total_creators],
        ["情感分", overview.sentiment_score],
        ["平台", "、".join(platform_label(p) for p in scope.platforms) or MISSING],
        ["周期", f"{scope.period.start} 至 {scope.period.end}"],
        ["关键词", "、".join(scope.keywords) or MISSING],
        ["对比模式", scope.comparison_mode or MISSING],
    ]
    write_table(
        sheet,
        2,
        None,
        ["指标", "值"],
        rows,
        columns=2,
    )


def _render_comparisons(sheet, report) -> None:
    clear_rows_unmerged(sheet, 2, 100, 6)
    sheet._charts = []
    comparisons = report.data.comparisons
    cursor = 4
    chart_row = 4
    for title, series in (
        ("活动期 vs 活动前", comparisons.current_baseline),
        ("活动期 vs 活动后观察期", comparisons.current_post),
    ):
        rows = [[s.metric, s.current, s.baseline, s.delta, s.rate] for s in series]
        layout = write_table(
            sheet,
            cursor,
            title,
            ["指标", "当前", "对比期", "差值", "变化率"],
            rows,
            columns=6,
            pct_columns=(5,),
            note=empty_note(report, "comparisons"),
        )
        if layout.has_data:
            sheet.add_chart(
                _bar_chart(sheet, layout, data_cols=(2, 3), cat_col=1, title=title),
                f"H{chart_row}",
            )
            chart_row += 16
        cursor = layout.next_row + 1
    timeline = report.data.timeline
    timeline_rows = [
        [item.date, platform_label(item.platform), item.volume, item.engagement, item.posts]
        for item in timeline
    ]
    timeline_layout = write_table(
        sheet,
        cursor,
        "时间线趋势",
        ["日期", "平台", "声量", "互动", "发帖"],
        timeline_rows,
        columns=5,
        note=empty_note(report, "timeline"),
    )
    if timeline_layout.has_data:
        sheet.add_chart(
            _bar_chart(
                sheet, timeline_layout, data_cols=(3,), cat_col=1, title="时间线趋势"
            ),
            f"H{chart_row}",
        )


def _render_platforms(sheet, report) -> None:
    clear_rows_unmerged(sheet, 4, 100, 6)
    sheet._charts = []
    rows = [
        [
            platform_label(item.platform),
            item.volume,
            item.engagement,
            item.posts,
            item.creators,
            item.share,
        ]
        for item in report.data.platform_contributions
    ]
    layout = write_table(
        sheet,
        4,
        None,
        ["平台", "声量", "互动", "发帖", "达人", "占比"],
        rows,
        columns=6,
        pct_columns=(6,),
        note=empty_note(report, "platform_contributions"),
    )
    if layout.has_data:
        sheet.add_chart(
            _bar_chart(sheet, layout, data_cols=(2,), cat_col=1, title="平台声量分布"),
            f"A{layout.next_row + 1}",
        )


def _render_sentiment_content(sheet, report) -> None:
    clear_rows_unmerged(sheet, 4, 100, 6)
    sentiment = report.data.sentiment
    summary = sentiment.summary
    rows = [
        ["正面", summary.positive.count, summary.positive.share],
        ["中性", summary.neutral.count, summary.neutral.share],
        ["负面", summary.negative.count, summary.negative.share],
    ]
    sheet._charts = []
    layout = write_table(
        sheet,
        4,
        "情感分布",
        ["情感", "计数", "占比"],
        rows,
        columns=3,
        pct_columns=(3,),
        note=empty_note(report, "sentiment"),
    )
    if layout.has_data and (
        summary.positive.count is not None or summary.negative.count is not None
    ):
        sheet.add_chart(
            _bar_chart(
                sheet, layout, data_cols=(2,), cat_col=1, title="情感分布", width=14
            ),
            f"E{layout.title_row or layout.header_row}",
        )
    write_table(
        sheet,
        layout.next_row + 1,
        "内容类型",
        ["内容类型", "发帖", "声量", "互动"],
        [
            [item.type, item.posts, item.volume, item.engagement]
            for item in report.data.content_types
        ],
        columns=4,
        note=empty_note(report, "content_types"),
    )


def _render_top_posts(sheet, report) -> None:
    posts = report.data.top_posts
    clear_rows_unmerged(sheet, 2, 100, 8)
    for column, header in enumerate(
        ("排名", "平台", "标题", "作者", "互动数", "点赞", "评论", "分享"), start=1
    ):
        sheet.cell(3, column).value = header
    if not posts:
        sheet.merge_cells("A4:H4")
        sheet["A4"] = empty_note(report, "top_posts")
        return
    for index, post in enumerate(posts, start=1):
        row = 3 + index
        sheet.cell(row, 1).value = index
        sheet.cell(row, 2).value = cell_value(platform_label(post.platform))
        title_cell = sheet.cell(row, 3)
        title_cell.value = cell_value(post.title)
        if post.url is not None:
            title_cell.hyperlink = post.url
        sheet.cell(row, 4).value = cell_value(post.author)
        sheet.cell(row, 5).value = post.engagement
        sheet.cell(row, 6).value = post.likes
        sheet.cell(row, 7).value = post.comments
        sheet.cell(row, 8).value = post.shares


def _render_kols(sheet, report) -> None:
    clear_rows_unmerged(sheet, 4, 100, 6)
    rows = [
        [
            platform_label(item.platform),
            item.nickname,
            item.posts,
            item.volume,
            item.engagement,
            item.contribution_share,
        ]
        for item in report.data.kol_contributions
    ]
    write_table(
        sheet,
        4,
        None,
        ["平台", "达人", "发帖", "声量", "互动", "贡献占比"],
        rows,
        columns=6,
        pct_columns=(6,),
        note=empty_note(report, "kol_contributions"),
    )


def _render_organic_audience(sheet, report) -> None:
    clear_rows_unmerged(sheet, 2, 100, 6)
    sheet._charts = []
    organic = report.data.organic_summary
    rows = []
    if organic is not None:
        rows = [
            ["自然声量", organic.volume],
            ["自然互动", organic.engagement],
            ["自然帖数", organic.posts],
            ["自然声量占比", organic.share_of_volume],
        ]
    organic_layout = write_table(
        sheet,
        4,
        "自然传播",
        ["指标", "值"],
        rows,
        columns=2,
        note=empty_note(report, "organic_summary"),
    )
    audience_layout = write_table(
        sheet,
        organic_layout.next_row + 1,
        "受众地域",
        ["地区", "声量", "占比"],
        [
            [item.region, item.volume, item.share]
            for item in report.data.audience_regions
        ],
        columns=3,
        pct_columns=(3,),
        note=empty_note(report, "audience_regions"),
    )
    attribution = report.data.attribution
    if attribution is None or not any(
        v is not None for v in (attribution.paid_confirmed, attribution.organic, attribution.unknown)
    ):
        return
    attr_rows = [
        ["付费确认", attribution.paid_confirmed],
        ["自然", attribution.organic],
        ["未知", attribution.unknown],
    ]
    attr_layout = write_table(
        sheet,
        audience_layout.next_row + 1,
        "内容归属",
        ["归属", "帖数"],
        attr_rows,
        columns=2,
        note=empty_note(report, "attribution"),
    )
    if attr_layout.has_data:
        sheet.add_chart(
            _bar_chart(
                sheet, attr_layout, data_cols=(2,), cat_col=1, title="内容归属", width=14
            ),
            f"D{attr_layout.title_row or attr_layout.header_row}",
        )


def _render_insights(sheet, report) -> None:
    clear_rows_unmerged(sheet, 4, 100, 4)
    write_table(
        sheet,
        4,
        "核心发现",
        ["发现", "详情"],
        [[item.title, item.detail] for item in report.narrative.findings],
        columns=4,
        note=empty_note(report, "narrative"),
    )


def _render_methodology(sheet, report) -> None:
    clear_rows_unmerged(sheet, 4, 60, 4)
    methodology = report.methodology
    rows = [
        ["数据来源", "、".join(methodology.source_names) or MISSING],
        ["数据截至", f"{methodology.data_as_of:%Y-%m-%d %H:%M}"],
        ["对比模式", report.scope.comparison_mode or MISSING],
        ["归属规则", "、".join(report.scope.attribution_rules) or MISSING],
        ["排除规则", "、".join(report.scope.exclusions) or MISSING],
        ["官方账号", "、".join(report.scope.official_accounts) or MISSING],
    ]
    for note in methodology.notes:
        rows.append(["说明", note])
    write_table(
        sheet,
        4,
        None,
        ["项目", "说明"],
        rows,
        columns=2,
    )


def _render_roi(sheet, report) -> None:
    roi = report.data.roi
    sheet["A1"] = cell_value("ROI 与转化")
    clear_rows_unmerged(sheet, 2, 40, 6)
    rows = [
        ["投放金额", roi.spend],
        ["销售额", roi.revenue],
        ["转化数", roi.conversions],
        ["归因窗口", roi.attribution_window],
        ["ROI", roi.roi],
        ["ROAS", roi.roas],
    ]
    write_table(
        sheet,
        2,
        None,
        ["指标", "值"],
        rows,
        columns=2,
    )


__all__ = [
    "ROI_SHEET",
    "SHEET_ORDER",
    "TEMPLATE_PATH",
    "render_campaign_workbook",
]
