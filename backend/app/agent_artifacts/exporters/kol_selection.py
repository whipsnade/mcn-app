"""kol_selection_v3 Excel 导出渲染器（Gate C Task 3 / 设计 §7.3）。

按受控模板 ``templates/kol_selection_v3.xlsx`` 渲染 4 个 Sheet：
达人圈选总表 / 达人详细画像（Top20 全详情块）/ 粉丝画像详情 /
评分方法论与数据来源。模板由 ``scripts/build_agent_artifact_templates.py``
从用户来源模板清洗生成（样例数据与图表清除）。

- v3 名单展示效果分/价格效率分/价值总分/报价/评级/数据完整度 + 八维原始分；
- 历史 ``kol_score_v2`` 快照仍可导出（total/rating/stars/完整度 + v2 八维）；
- 导出只读已发布 Version 的 payload，不调用模型/MCP；
- URL 仅 http/https（payload 已校验）；所有第三方可控文本经 cell_value 防
  公式注入。
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

from app.agent_artifacts.exporters._common import (
    cell_value,
    clear_rows_unmerged,
    platform_label,
    present,
    write_table,
    write_value,
)
from app.agent_artifacts.payloads.kol_selection import (
    V3_DIMENSIONS,
    KolSelectionV3,
)
from app.agent_artifacts.validation import validate_kol_candidates

TEMPLATE_PATH = (
    Path(__file__).resolve().parents[1] / "templates" / "kol_selection_v3.xlsx"
)
SUMMARY_SHEET = "达人圈选总表"
DETAIL_SHEET = "达人详细画像"
FAN_SHEET = "粉丝画像详情"
METHOD_SHEET = "评分方法论与数据来源"
MAX_DETAIL_BLOCKS = 20

V3_DIMENSION_LABELS = {
    "average_interactions": "平均互动表现",
    "active_follower": "活跃/有效粉丝",
    "engagement_follower_ratio": "互动粉丝比",
    "content_match": "内容与行业匹配",
    "followers": "粉丝规模",
    "industry_interest": "目标行业兴趣",
    "target_region": "目标地区匹配",
    "target_age": "目标年龄匹配",
}

V2_DIMENSION_LABELS = {
    "industry_interest": "行业兴趣",
    "target_region": "目标地区",
    "target_age": "目标年龄",
    "engagement": "互动表现",
    "active_follower": "活跃粉丝",
    "content": "内容质量",
    "followers": "粉丝规模",
    "engagement_follower_ratio": "互动粉丝比",
}

# 总表列：序号/昵称/平台/粉丝数/评级 + 版本相关列 + 八维原始分。
_SUMMARY_HEADER_PREFIX = ("序号", "昵称", "平台", "粉丝数", "评级")
_V3_SCORE_HEADERS = ("效果分", "价格效率分", "价值总分", "报价", "价格样本数", "数据完整度")
_V2_SCORE_HEADERS = ("综合分", "星级", "数据完整度")


def render_kol_selection_workbook(payload: dict) -> bytes:
    """把已发布 kol_selection_v3 payload 渲染为 .xlsx bytes（同步 CPU 密集）。"""
    selection = KolSelectionV3.model_validate(payload)
    issues = validate_kol_candidates(selection)
    if issues:
        raise ValueError("kol selection failed publication validity")
    workbook = load_workbook(TEMPLATE_PATH)
    _render_summary(workbook[SUMMARY_SHEET], selection)
    _render_detail_blocks(workbook[DETAIL_SHEET], selection)
    _render_fan_profiles(workbook[FAN_SHEET], selection)
    _render_methodology(workbook[METHOD_SHEET], selection)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _is_v3(snapshot) -> bool:
    return getattr(snapshot, "version", "") == "kol_value_score_v3"


def _summary_values(item) -> tuple[list[str], list[object]]:
    """按快照版本返回 (额外列头, 行值)。"""
    snapshot = item.score_snapshot
    if _is_v3(snapshot):
        headers = list(_V3_SCORE_HEADERS)
        values: list[object] = [
            snapshot.effect_score,
            snapshot.price_efficiency_score,
            snapshot.value_score,
            snapshot.quoted_price,
            snapshot.price_sample_size,
            snapshot.data_completeness,
        ]
        dims = list(V3_DIMENSIONS)
        labels = V3_DIMENSION_LABELS
    else:
        headers = list(_V2_SCORE_HEADERS)
        values = [snapshot.total, snapshot.stars, snapshot.data_completeness]
        dims = sorted(snapshot.dimensions)
        labels = V2_DIMENSION_LABELS
    for dim in dims:
        headers.append(labels.get(dim, dim))
        values.append(snapshot.dimensions[dim].raw_score)
    return headers, values


def _render_summary(sheet, selection) -> None:
    scope = selection.scope
    clear_rows_unmerged(sheet, 1, 200, 19)
    brand = scope.brand or "KOL"
    sheet["A1"] = cell_value(f"{brand} KOL 圈选名单")
    platforms = "、".join(platform_label(name) for name in scope.platforms) or "未指定"
    summary = selection.data.summary
    sheet["A2"] = cell_value(
        f"平台: {platforms} | 候选: {present(summary.candidate_count)} | "
        f"圈选: {present(summary.selected_count)} | 评分: {selection.data.scoring.version}"
    )
    items = list(selection.data.items)
    extra_headers, _ = _summary_values(items[0]) if items else ([], [])
    headers = list(_SUMMARY_HEADER_PREFIX) + extra_headers
    rows = []
    for item in items:
        _headers, values = _summary_values(item)
        rows.append(
            [
                item.rank,
                item.nickname,
                platform_label(item.platform),
                item.followers,
                item.score_snapshot.rating,
                *values,
            ]
        )
    write_table(
        sheet,
        4,
        None,
        headers,
        rows,
        columns=19,
        note="候选 KOL 数据不足，无法形成完整名单",
    )


def _render_detail_blocks(sheet, selection) -> None:
    clear_rows_unmerged(sheet, 1, 1000, 6)
    row = 1
    for item in selection.data.items[:MAX_DETAIL_BLOCKS]:
        row = _write_detail_block(sheet, row, item)


def _write_detail_block(sheet, start_row: int, item) -> int:
    snapshot = item.score_snapshot
    if _is_v3(snapshot):
        title = (
            f"#{item.rank} {item.nickname} — {snapshot.rating} "
            f"(价值 {snapshot.value_score} 分 / 效果 {snapshot.effect_score} / "
            f"价格效率 {snapshot.price_efficiency_score})"
        )
    else:
        title = (
            f"#{item.rank} {item.nickname} — {snapshot.rating} "
            f"(综合 {snapshot.total} 分)"
        )
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    title_cell = sheet.cell(start_row, 1)
    title_cell.value = cell_value(title)
    title_cell.font = Font(name="微软雅黑", bold=True, size=11, color="1F4E79")
    row = start_row + 1

    def _section(section_row: int, text: str) -> int:
        sheet.merge_cells(
            start_row=section_row, start_column=1, end_row=section_row, end_column=6
        )
        section_cell = sheet.cell(section_row, 1)
        section_cell.value = text
        section_cell.font = Font(name="微软雅黑", bold=True, size=10)
        return section_row + 1

    row = _section(row, "【达人概况】")
    overview_rows: list[list[object]] = [
        ["平台", platform_label(item.platform)],
        ["粉丝数", item.followers],
        ["主页", item.homepage_url],
        ["评级", snapshot.rating],
    ]
    if _is_v3(snapshot):
        overview_rows.extend(
            [
                ["效果分", snapshot.effect_score],
                ["价格效率分", snapshot.price_efficiency_score],
                ["价值总分", snapshot.value_score],
                ["报价", snapshot.quoted_price],
                ["价格样本数", snapshot.price_sample_size],
                ["数据完整度", snapshot.data_completeness],
            ]
        )
    else:
        overview_rows.extend(
            [
                ["综合分", snapshot.total],
                ["星级", snapshot.stars],
                ["数据完整度", snapshot.data_completeness],
            ]
        )
    _DATA_FONT = Font(name="微软雅黑", size=10)
    _DATA_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    _DATA_ALIGNMENT = Alignment(vertical="center")
    for label, value in overview_rows:
        label_cell = sheet.cell(row, 1)
        label_cell.value = label
        label_cell.font = _DATA_FONT
        label_cell.border = _DATA_BORDER
        label_cell.alignment = _DATA_ALIGNMENT
        value_cell = sheet.cell(row, 2)
        write_value(value_cell, value)
        value_cell.font = _DATA_FONT
        value_cell.border = _DATA_BORDER
        value_cell.alignment = _DATA_ALIGNMENT
        sheet.row_dimensions[row].height = 20
        row += 1

    row = _section(row + 1, "【评分维度】")
    labels = V3_DIMENSION_LABELS if _is_v3(snapshot) else V2_DIMENSION_LABELS
    for dim in snapshot.dimensions:
        entry = snapshot.dimensions[dim]
        sheet.cell(row, 1).value = labels.get(dim, dim)
        sheet.cell(row, 2).value = entry.raw_score
        sheet.cell(row, 3).value = entry.weighted_score
        write_value(sheet.cell(row, 4), "缺失" if entry.missing_reason else "完整")
        write_value(sheet.cell(row, 5), present(entry.missing_reason))
        for column in range(1, 6):
            cell = sheet.cell(row, column)
            cell.font = _DATA_FONT
            cell.border = _DATA_BORDER
            cell.alignment = _DATA_ALIGNMENT
        sheet.row_dimensions[row].height = 20
        row += 1
    return row + 1  # 块后空一行


def _render_fan_profiles(sheet, selection) -> None:
    clear_rows_unmerged(sheet, 1, 200, 12)
    rows = []
    for item in selection.data.items:
        audience = item.audience
        rows.append(
            [
                item.rank,
                item.nickname,
                item.followers,
                "、".join(audience.regions) or None,
                "、".join(audience.age_ranges) or None,
                "、".join(audience.interests) or None,
            ]
        )
    write_table(
        sheet,
        1,
        None,
        ["序号", "昵称", "粉丝数", "目标地区", "目标年龄", "目标兴趣"],
        rows,
        columns=12,
        note="粉丝画像分布数据未采集",
    )


def _render_methodology(sheet, selection) -> None:
    clear_rows_unmerged(sheet, 5, 100, 4)
    scoring = selection.data.scoring
    weights = scoring.weights
    if scoring.version == "kol_value_score_v3":
        labels = V3_DIMENSION_LABELS
        rows = [
            [
                labels.get(dim, dim),
                weights[dim],
                (
                    "同平台 mid-rank percentile（绝对量）/ 0–100 clamp（比例匹配）；"
                    "缺失计 0 不重分配"
                ),
                "效果与匹配度合计 70 分",
            ]
            for dim in weights
        ]
        rows.append(
            [
                "价格效率",
                30,
                "effect/quote 全候选 mid-rank 分位 × 30",
                "有效报价 ≥3 才计算；样本不足计 0",
            ]
        )
        rows.append(
            [
                "价值总分",
                100,
                "效果 70 + 价格效率 30",
                "评级：≥78 重点推荐 / 62–77 推荐 / 48–61 可考虑 / <48 或效果<35 观察",
            ]
        )
    else:
        rows = [
            [dim, weights[dim], "八维加权求和（raw×weight/100）", "缺失计 0 不重分配"]
            for dim in weights
        ]
    write_table(
        sheet,
        5,
        None,
        ["维度", "权重", "计算方式", "说明"],
        rows,
        columns=4,
    )
    sheet["A3"] = cell_value(f"评分版本：{scoring.version}（{scoring.method}）")


__all__ = [
    "DETAIL_SHEET",
    "FAN_SHEET",
    "METHOD_SHEET",
    "SUMMARY_SHEET",
    "TEMPLATE_PATH",
    "render_kol_selection_workbook",
]
