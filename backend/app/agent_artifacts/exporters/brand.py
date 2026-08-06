"""brand_report_v3 Excel 导出渲染器（Gate C Task 3 / 设计 §12.1）。

渲染器只读已发布不可变 Version 的 payload（强类型 ``BrandReportV3``），按受控
模板 ``templates/brand_report_v3.xlsx`` 填充 8 个 Sheet（综合概览/情感分析/
日趋势/内容类型与达人/地域分布/热门帖子TOP/舆情洞察/方法论）。模板由
``scripts/build_agent_artifact_templates.py`` 从用户来源模板清洗生成：只保留
版式，样例数据与图表被清除，图表由本渲染器现场重建。

空章节保留表头并写「数据受限/未采集」说明，绝不因空数据画误导性图表；
受限章节经 ``availability``/``limitations`` 披露。URL 仅 http/https（payload
已校验），所有第三方可控文本经 ``cell_value`` 防公式注入。
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference

from app.agent_artifacts.exporters._common import (
    MISSING,
    cell_value,
    clear_rows_unmerged,
    empty_note,
    platform_label,
    write_value,
    write_table,
)
from app.agent_artifacts.payloads.brand import BrandReportV3

TEMPLATE_PATH = (
    Path(__file__).resolve().parents[1] / "templates" / "brand_report_v3.xlsx"
)
SHEET_ORDER = (
    "综合概览",
    "情感分析",
    "日趋势",
    "内容类型与达人",
    "地域分布",
    "热门帖子TOP",
    "舆情洞察",
    "方法论",
)


def _write_title(sheet, text: str, columns: int = 8) -> None:
    sheet["A1"] = cell_value(text)


def render_brand_workbook(payload: dict) -> bytes:
    """把已发布 brand_report_v3 payload 渲染为 .xlsx bytes（同步 CPU 密集）。"""
    report = BrandReportV3.model_validate(payload)
    workbook = load_workbook(TEMPLATE_PATH)
    _write_title(workbook["综合概览"], f"{report.scope.brand} 品牌社交媒体表现分析报告")
    _write_title(workbook["情感分析"], "内容情感分布")
    _write_title(workbook["日趋势"], "每日声量与互动趋势")
    _write_title(workbook["内容类型与达人"], "内容类型与达人分布")
    _write_title(workbook["地域分布"], "发帖用户地域分布")
    _write_title(workbook["热门帖子TOP"], "热门帖子 TOP")
    _write_title(workbook["舆情洞察"], "舆情洞察与内容主题分析")
    _write_title(workbook["方法论"], "数据说明与方法论")
    _render_overview(workbook["综合概览"], report)
    _render_sentiment(workbook["情感分析"], report)
    _render_daily_trend(workbook["日趋势"], report)
    _render_content_creators(workbook["内容类型与达人"], report)
    _render_regions(workbook["地域分布"], report)
    _render_top_posts(workbook["热门帖子TOP"], report)
    _render_insights(workbook["舆情洞察"], report)
    _render_methodology(workbook["方法论"], report)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


# ---------- 各 Sheet 渲染 ----------


def _render_overview(sheet, report) -> None:
    scope = report.scope
    sheet["B2"] = cell_value(
        f"{scope.period.start} 至 {scope.period.end}（数据截至 {report.methodology.data_as_of:%Y-%m-%d}）"
    )
    platforms = "、".join(platform_label(name) for name in scope.platforms) or MISSING
    sheet["B3"] = cell_value(f"平台：{platforms}")
    sheet["B4"] = cell_value(f"关键词：{'、'.join(scope.keywords) or MISSING}")

    # 指标表固定标签：A6=指标（行名列）、E6=总计（合计列）；平台列头写 B/C/D…，
    # 列数不足用「未采集」占位。
    sheet.cell(6, 1).value = "指标"
    sheet.cell(6, 5).value = "总计"
    overview = report.data.overview
    platform_entries = list(overview.platforms)
    for index, entry in enumerate(platform_entries[:3]):
        sheet.cell(6, 2 + index).value = cell_value(platform_label(entry.platform))
    for index in range(len(platform_entries[:3]), 3):
        sheet.cell(6, 2 + index).value = MISSING

    metric_rows = [
        ("声量(帖数)", [entry.volume for entry in platform_entries], overview.total_volume),
        ("互动数", [entry.engagement for entry in platform_entries], overview.total_engagement),
        ("发帖数", [entry.posts for entry in platform_entries], overview.total_posts),
        ("情感分", [None for _ in platform_entries], overview.sentiment_score),
    ]
    for offset, (label, per_platform, total) in enumerate(metric_rows):
        row = 7 + offset
        sheet.cell(row, 1).value = label
        for index, value in enumerate(per_platform[:3]):
            write_value(sheet.cell(row, 2 + index), value)
        write_value(sheet.cell(row, 5), total)


def _render_sentiment(sheet, report) -> None:
    clear_rows_unmerged(sheet, 4, 60, 5)
    rows = []
    for item in report.data.sentiment.by_platform:
        for bucket_label, bucket in (
            ("正面", item.positive),
            ("中性", item.neutral),
            ("负面", item.negative),
        ):
            rows.append(
                [
                    platform_label(item.platform),
                    bucket_label,
                    bucket.count,
                    None,  # 互动数无对应字段
                    bucket.share,
                ]
            )
    write_table(
        sheet,
        4,
        None,
        ["平台", "情感", "声量", "互动数", "占比"],
        rows,
        columns=5,
        pct_columns=(5,),
        note=empty_note(report, "sentiment"),
    )


def _render_daily_trend(sheet, report) -> None:
    points = report.data.daily_trend
    clear_rows_unmerged(sheet, 2, 100, 4)
    sheet._charts = []
    # 表头由 exporter 写（系列名来自表头，避免 Series 1/Series 2）。
    for column, header in enumerate(("日期", "声量", "互动数", "备注"), start=1):
        sheet.cell(3, column).value = header
    if not points:
        sheet.merge_cells("A4:D4")
        sheet["A4"] = empty_note(report, "daily_trend")
        return
    for index, point in enumerate(points):
        row = 4 + index
        # 日期写文本（ISO），避免 Excel 日期序列号。
        sheet.cell(row, 1).value = point.date.isoformat()
        sheet.cell(row, 2).value = point.volume
        sheet.cell(row, 3).value = point.engagement
        sheet.cell(row, 4).value = cell_value(platform_label(point.platform))
    last_data_row = 3 + len(points)
    chart = LineChart()
    chart.title = "每日声量与互动趋势"
    chart.add_data(
        Reference(sheet, min_col=2, min_row=3, max_row=last_data_row), titles_from_data=True
    )
    chart.add_data(
        Reference(sheet, min_col=3, min_row=3, max_row=last_data_row), titles_from_data=True
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=4, max_row=last_data_row))
    chart.height = 9
    chart.width = 17
    sheet.add_chart(chart, f"F{last_data_row + 2}")


def _render_content_creators(sheet, report) -> None:
    clear_rows_unmerged(sheet, 4, 60, 6)
    rows = [
        [
            item.type,
            item.volume,
            item.engagement,
            item.posts,
        ]
        for item in report.data.content_types
    ]
    content_layout = write_table(
        sheet,
        4,
        None,
        ["内容类型", "声量", "互动数", "发帖数"],
        rows,
        columns=4,
        note=empty_note(report, "content_types"),
    )
    # 达人分层 6 列（平台/层级/达人数量/发帖数/声量/互动数），接在内容类型表后。
    write_table(
        sheet,
        content_layout.next_row + 1,
        "达人分层",
        ["平台", "层级", "达人数量", "发帖数", "声量", "互动数"],
        [
            [
                platform_label(item.platform),
                item.tier,
                item.creator_count,
                item.posts,
                item.volume,
                item.engagement,
            ]
            for item in report.data.creator_tiers
        ],
        columns=6,
        note=empty_note(report, "creator_tiers"),
    )


def _render_regions(sheet, report) -> None:
    regions = report.data.regions
    clear_rows_unmerged(sheet, 4, 60, 4)
    sheet._charts = []
    if not regions:
        sheet.merge_cells("A4:D4")
        sheet["A4"] = empty_note(report, "regions")
        return
    rows = [
        [item.region, item.volume, item.share] for item in regions
    ]
    write_table(
        sheet,
        4,
        None,
        ["省份", "声量", "互动数", "声量占比"],
        [[row[0], row[1], None, row[2]] for row in rows],
        columns=4,
        pct_columns=(4,),
    )
    last_data_row = 4 + len(rows)
    chart = BarChart()
    chart.title = "发帖用户地域分布"
    chart.add_data(
        Reference(sheet, min_col=2, min_row=5, max_row=last_data_row), titles_from_data=False
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=5, max_row=last_data_row))
    chart.height = 8
    chart.width = 16
    sheet.add_chart(chart, f"F{last_data_row + 2}")


def _render_top_posts(sheet, report) -> None:
    posts = report.data.top_posts
    clear_rows_unmerged(sheet, 2, 100, 8)
    for column, header in enumerate(
        ("排名", "平台", "标题", "用户昵称", "互动数", "阅读数", "点赞", "评论"), start=1
    ):
        sheet.cell(3, column).value = header
    if not posts:
        sheet.merge_cells("A4:H4")
        sheet["A4"] = empty_note(report, "top_posts")
        return
    for index, post in enumerate(posts, start=1):
        row = 3 + index
        sheet.cell(row, 1).value = index
        sheet.cell(row, 2).value = cell_value(platform_label(post.platform))
        title_cell = sheet.cell(row, 3)
        title_cell.value = cell_value(post.title)
        if post.url is not None:
            title_cell.hyperlink = post.url  # 仅 http/https（payload 已校验）
        sheet.cell(row, 4).value = cell_value(post.author)
        sheet.cell(row, 5).value = post.engagement
        sheet.cell(row, 6).value = None  # 阅读数无对应字段
        sheet.cell(row, 7).value = post.likes
        sheet.cell(row, 8).value = post.comments


def _render_insights(sheet, report) -> None:
    clear_rows_unmerged(sheet, 4, 60, 4)
    findings = [
        [finding.title, finding.detail, None]
        for finding in report.narrative.findings
    ]
    write_table(
        sheet,
        4,
        None,
        ["洞察维度", "具体表现", "代表特征"],
        findings,
        columns=4,
        note=empty_note(report, "narrative"),
    )
    topics = report.data.topics
    row = 4 + len(findings)
    write_table(
        sheet,
        row + 1,
        "内容主题",
        ["主题", "声量", "互动数", "情感分"],
        [
            [item.topic, item.volume, item.engagement, item.sentiment_score]
            for item in topics
        ],
        columns=4,
        note=empty_note(report, "topics"),
    )


def _render_methodology(sheet, report) -> None:
    clear_rows_unmerged(sheet, 4, 60, 2)
    methodology = report.methodology
    rows = [
        ["数据来源", "、".join(methodology.source_names) or MISSING],
        ["数据截至", f"{methodology.data_as_of:%Y-%m-%d %H:%M}"],
        [
            "分析周期",
            f"{report.scope.period.start} 至 {report.scope.period.end}",
        ],
        ["平台", "、".join(platform_label(p) for p in report.scope.platforms) or MISSING],
        ["关键词", "、".join(report.scope.keywords) or MISSING],
    ]
    for note in methodology.notes:
        rows.append(["说明", note])
    write_table(
        sheet,
        4,
        None,
        ["项目", "说明"],
        rows,
        columns=2,
    )


__all__ = [
    "SHEET_ORDER",
    "TEMPLATE_PATH",
    "render_brand_workbook",
]
