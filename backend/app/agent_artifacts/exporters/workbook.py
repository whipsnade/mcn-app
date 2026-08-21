"""`workbook_v1` 的安全、确定性 Excel 渲染基础设施。

Workbook 只是同一 Report Version 的表现层投影。这里不接受二进制、公式、宏或
脚本输入；所有单元格、链接、分页和技术上限都由可信代码控制。
"""

from __future__ import annotations

import hashlib
import json
import re
import zipfile
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from typing import Any
from urllib.parse import urlsplit

from openpyxl import Workbook
from openpyxl.chart import AreaChart, BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from app.agent_artifacts.exporters.errors import ArtifactExportUnsupported
from app.agent_artifacts.payloads.analysis_report import (
    AnalysisReportBlock,
    AnalysisReportChartBlock,
    AnalysisReportLinkListBlock,
    AnalysisReportMetricCardsBlock,
    AnalysisReportMethodologyLimitationsBlock,
    AnalysisReportNarrativeBlock,
    AnalysisReportTimeSeriesBlock,
    AnalysisReportTypedTableBlock,
    AnalysisReportV1,
    AnalysisReportWorkbookLayout,
    AnalysisReportWorkbookSheet,
)

from ._common import MISSING, cell_value

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_FIXED_ZIP_TIME = (1980, 1, 1, 0, 0, 0)
_HEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
_HEADER_FONT = Font(name="微软雅黑", bold=True, color="1F1F1F")
_LINK_FONT = Font(name="微软雅黑", color="0563C1", underline="single")
_DATA_FONT = Font(name="微软雅黑", size=10)
_DATA_ALIGNMENT = Alignment(vertical="top", wrap_text=True)


@dataclass(frozen=True)
class WorkbookLimits:
    """Workbook 表现层技术上限；不表达任何业务数量上限。"""

    max_sheets: int = 32
    max_rows_per_sheet: int = 100_000
    max_columns: int = 256
    max_cell_chars: int = 32_767
    max_bytes: int = 20 * 1024 * 1024

    def __post_init__(self) -> None:
        for name in (
            "max_sheets",
            "max_rows_per_sheet",
            "max_columns",
            "max_cell_chars",
            "max_bytes",
        ):
            if getattr(self, name) <= 0:
                raise ValueError(f"{name} must be positive")

    @classmethod
    def from_settings(cls, settings: Any) -> WorkbookLimits:
        return cls(
            max_sheets=settings.workbook_max_sheets,
            max_rows_per_sheet=settings.workbook_max_rows_per_sheet,
            max_columns=settings.workbook_max_columns,
            max_cell_chars=settings.analysis_report_max_cell_chars,
            max_bytes=settings.workbook_max_bytes,
        )


class WorkbookTechnicalLimitExceeded(ArtifactExportUnsupported):
    """Workbook 超过技术上限；不得静默截断业务数据。"""

    code = "workbook_technical_limit_exceeded"

    def __init__(self, reason: str) -> None:
        super().__init__("workbook_v1", reason=reason)


@dataclass(frozen=True)
class _Section:
    headers: tuple[str, ...]
    keys: tuple[str, ...]
    types: tuple[str, ...]
    rows: tuple[tuple[Any, ...], ...]
    chart_type: str | None = None


def _jsonable(value: Any) -> Any:
    if isinstance(value, BaseModel):
        return value.model_dump(mode="json")
    return value


def workbook_layout_digest(layout: AnalysisReportWorkbookLayout | None) -> str:
    """对规范化布局投影计算稳定 SHA-256。"""
    normalized = _jsonable(layout) if layout is not None else {
        "schema_version": "workbook_v1",
        "sheets": [],
    }
    encoded = json.dumps(
        normalized,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _safe_sheet_title(title: str, used: set[str]) -> str:
    base = _INVALID_SHEET_CHARS.sub("_", title).strip() or "Report"
    base = base[:31]
    candidate = base
    suffix = 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{base[:31 - len(tail)]}{tail}"
        suffix += 1
    used.add(candidate)
    return candidate


def _safe_url(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    return value if urlsplit(value).scheme in {"http", "https"} else None


def _display(value: Any, *, max_cell_chars: int) -> Any:
    if value is None:
        return MISSING
    if isinstance(value, str) and len(value) > max_cell_chars:
        raise WorkbookTechnicalLimitExceeded("cell value exceeds max_cell_chars")
    return cell_value(value)


def _sortable(value: Any) -> tuple[int, str]:
    if value is None:
        return (0, "")
    return (1, str(value))


def _table_section(
    block: AnalysisReportTypedTableBlock,
    layout_sheet: AnalysisReportWorkbookSheet,
) -> _Section:
    by_key = {column.key: column for column in block.columns}
    selected = list(layout_sheet.columns) if layout_sheet.columns else list(block.columns)
    for column in selected:
        if column.key not in by_key:
            raise ValueError(f"workbook column {column.key!r} is not in block {block.id!r}")
    keys = tuple(column.key for column in selected)
    types = tuple(by_key[key].type for key in keys)
    headers = tuple(column.label for column in selected)
    indexes = [next(index for index, source in enumerate(block.columns) if source.key == key) for key in keys]
    rows = [tuple(row[index] for index in indexes) for row in block.rows]
    if layout_sheet.sort_by:
        if not set(layout_sheet.sort_by) <= set(keys):
            raise ValueError("workbook sort_by references an unselected column")
        sort_indexes = [keys.index(key) for key in layout_sheet.sort_by]
        rows.sort(key=lambda row: tuple(_sortable(row[index]) for index in sort_indexes))
    return _Section(headers=headers, keys=keys, types=types, rows=tuple(rows))


def _block_section(block: AnalysisReportBlock, layout_sheet: AnalysisReportWorkbookSheet) -> _Section:
    if isinstance(block, AnalysisReportTypedTableBlock):
        return _table_section(block, layout_sheet)
    if isinstance(block, AnalysisReportMetricCardsBlock):
        return _Section(
            headers=("指标", "数值", "单位"),
            keys=("label", "value", "unit"),
            types=("string", "number", "string"),
            rows=tuple((card.label, card.value, card.unit) for card in block.cards),
        )
    if isinstance(block, AnalysisReportTimeSeriesBlock):
        value_keys = sorted({key for point in block.points for key in point.values})
        return _Section(
            headers=("时间", *value_keys),
            keys=("timestamp", *value_keys),
            types=("datetime", *("number" for _ in value_keys)),
            rows=tuple(
                (point.timestamp, *(point.values.get(key) for key in value_keys))
                for point in block.points
            ),
            chart_type="line",
        )
    if isinstance(block, AnalysisReportLinkListBlock):
        return _Section(
            headers=("名称", "说明", "链接"),
            keys=("label", "description", "url"),
            types=("string", "string", "url"),
            rows=tuple((item.label, item.description, item.url) for item in block.items),
        )
    if isinstance(block, AnalysisReportChartBlock):
        series = list(block.series)
        return _Section(
            headers=("分类", *(item.label for item in series)),
            keys=("category", *(item.key for item in series)),
            types=("string", *("number" for _ in series)),
            rows=tuple(
                (
                    category,
                    *(series_index.values[index] if index < len(series_index.values) else None for series_index in series),
                )
                for index, category in enumerate(block.categories)
            ),
            chart_type=block.chart_type,
        )
    if isinstance(block, AnalysisReportNarrativeBlock):
        return _Section(
            headers=("标题", "内容"),
            keys=("title", "content"),
            types=("string", "string"),
            rows=((block.title, block.content),),
        )
    if isinstance(block, AnalysisReportMethodologyLimitationsBlock):
        rows = [("方法论", block.methodology)]
        rows.extend(("限制", limitation) for limitation in block.limitations)
        return _Section(
            headers=("项目", "说明"),
            keys=("kind", "detail"),
            types=("string", "string"),
            rows=tuple(rows),
        )
    raise TypeError(f"unsupported report block: {type(block).__name__}")


def _report_metadata_section(report: AnalysisReportV1) -> _Section:
    methodology = report.methodology
    rows: list[tuple[Any, Any]] = [
        ("数据状态", report.data_status),
        ("数据截至", methodology.data_as_of.isoformat(sep=" ")),
        ("数据来源", "、".join(methodology.source_names) or MISSING),
    ]
    rows.extend(("方法说明", note) for note in methodology.notes)
    rows.extend(
        ("限制", f"{limitation.code}：{limitation.message}")
        for limitation in report.limitations
    )
    rows.extend(
        (
            "结果完整性",
            f"{item.key}: {item.actual_count}/{item.requested_min} · {item.status} · {item.reason}",
        )
        for item in report.fulfillment
    )
    return _Section(
        headers=("项目", "说明"),
        keys=("kind", "detail"),
        types=("string", "string"),
        rows=tuple(rows),
    )


def _effective_layout(report: AnalysisReportV1) -> AnalysisReportWorkbookLayout:
    if report.workbook is not None and report.workbook.sheets:
        return report.workbook
    return AnalysisReportWorkbookLayout(
        sheets=(
            {
                "key": "report",
                "title": report.title[:31],
                "block_ids": tuple(block.id for block in report.blocks),
            },
        )
    )


def _write_cell(
    cell: Any,
    value: Any,
    value_type: str,
    *,
    limits: WorkbookLimits,
    number_format: str | None = None,
) -> None:
    cell.value = _display(value, max_cell_chars=limits.max_cell_chars)
    cell.font = _DATA_FONT
    cell.alignment = _DATA_ALIGNMENT
    if number_format:
        cell.number_format = number_format
    elif value_type == "percent":
        cell.number_format = "0.00%"
    elif value_type == "date":
        cell.number_format = "yyyy-mm-dd"
    elif value_type == "datetime":
        cell.number_format = "yyyy-mm-dd hh:mm:ss"
    if value_type == "url":
        url = _safe_url(value)
        if url:
            cell.hyperlink = url
            cell.font = _LINK_FONT


def _write_section(
    worksheet: Any,
    section: _Section,
    start_row: int,
    *,
    layout_sheet: AnalysisReportWorkbookSheet,
    limits: WorkbookLimits,
) -> tuple[int, int, int]:
    if len(section.headers) > limits.max_columns:
        raise WorkbookTechnicalLimitExceeded("columns exceed max_columns")
    header_row = start_row
    for index, header in enumerate(section.headers, start=1):
        cell = worksheet.cell(header_row, index, _display(header, max_cell_chars=limits.max_cell_chars))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _DATA_ALIGNMENT
    for row_index, values in enumerate(section.rows, start=header_row + 1):
        for column_index, value in enumerate(values, start=1):
            _write_cell(
                worksheet.cell(row_index, column_index),
                value,
                section.types[column_index - 1],
                limits=limits,
                number_format=(
                    layout_sheet.columns[column_index - 1].number_format
                    if layout_sheet.columns and column_index <= len(layout_sheet.columns)
                    else None
                ),
            )
    end_row = header_row + len(section.rows)
    if layout_sheet.columns:
        for index, column in enumerate(layout_sheet.columns, start=1):
            width = column.width or max(10, min(60, len(column.label) * 1.5))
            worksheet.column_dimensions[get_column_letter(index)].width = width
    else:
        for index, header in enumerate(section.headers, start=1):
            worksheet.column_dimensions[get_column_letter(index)].width = max(
                10, min(60, len(header) * 1.5)
            )
    return header_row, end_row, end_row + 1


def _add_chart(worksheet: Any, section: _Section, header_row: int, end_row: int) -> None:
    if section.chart_type is None or end_row <= header_row:
        return
    chart_cls = {
        "bar": BarChart,
        "line": LineChart,
        "area": AreaChart,
        "pie": PieChart,
    }.get(section.chart_type)
    if chart_cls is None or len(section.headers) < 2:
        return
    chart = chart_cls()
    chart.title = section.headers[0]
    if section.chart_type == "pie":
        chart.add_data(Reference(worksheet, min_col=2, min_row=header_row, max_row=end_row), titles_from_data=True)
        chart.set_categories(Reference(worksheet, min_col=1, min_row=header_row + 1, max_row=end_row))
    else:
        chart.add_data(
            Reference(worksheet, min_col=2, max_col=len(section.headers), min_row=header_row, max_row=end_row),
            titles_from_data=True,
        )
        chart.set_categories(Reference(worksheet, min_col=1, min_row=header_row + 1, max_row=end_row))
    chart.height = 7
    chart.width = 14
    worksheet.add_chart(chart, f"{get_column_letter(len(section.headers) + 2)}{header_row}")


def _new_worksheet(workbook: Workbook, title: str, used_names: set[str], limits: WorkbookLimits) -> Any:
    if len(workbook.worksheets) >= limits.max_sheets:
        raise WorkbookTechnicalLimitExceeded("sheets exceed max_sheets")
    return workbook.create_sheet(_safe_sheet_title(title, used_names))


def _render_layout_sheet(
    workbook: Workbook,
    report: AnalysisReportV1,
    layout_sheet: AnalysisReportWorkbookSheet,
    blocks: list[AnalysisReportBlock],
    *,
    is_first_sheet: bool,
    used_names: set[str],
    limits: WorkbookLimits,
) -> None:
    sections: list[_Section] = []
    if is_first_sheet:
        sections.append(_report_metadata_section(report))
    sections.extend(_block_section(block, layout_sheet) for block in blocks)
    worksheet = _new_worksheet(workbook, layout_sheet.title, used_names, limits)
    row = 1
    if is_first_sheet:
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, min(limits.max_columns, 4)))
        worksheet.cell(1, 1).value = _display(report.title, max_cell_chars=limits.max_cell_chars)
        worksheet.cell(1, 1).font = Font(name="微软雅黑", bold=True, size=14)
        row = 2
    page_limit = min(layout_sheet.page_size or limits.max_rows_per_sheet, limits.max_rows_per_sheet)
    if page_limit < 2:
        raise WorkbookTechnicalLimitExceeded("max_rows_per_sheet must allow a header row")
    for section in sections:
        if len(section.headers) > limits.max_columns:
            raise WorkbookTechnicalLimitExceeded("columns exceed max_columns")
        data_cursor = 0
        if not section.rows:
            if row + 1 > page_limit:
                worksheet = _new_worksheet(workbook, f"{layout_sheet.title}_{len(workbook.worksheets) + 1}", used_names, limits)
                row = 1
            header_row, end_row, row = _write_section(
                worksheet,
                section,
                row,
                layout_sheet=layout_sheet,
                limits=limits,
            )
            _add_chart(worksheet, section, header_row, end_row)
            continue
        while data_cursor < len(section.rows):
            if row >= page_limit:
                worksheet = _new_worksheet(workbook, f"{layout_sheet.title}_{len(workbook.worksheets) + 1}", used_names, limits)
                row = 1
            capacity = page_limit - row
            if capacity <= 0:
                worksheet = _new_worksheet(workbook, f"{layout_sheet.title}_{len(workbook.worksheets) + 1}", used_names, limits)
                row = 1
                capacity = page_limit - row
            take = min(capacity, len(section.rows) - data_cursor)
            chunk = _Section(
                headers=section.headers,
                keys=section.keys,
                types=section.types,
                rows=section.rows[data_cursor:data_cursor + take],
                chart_type=section.chart_type,
            )
            header_row, end_row, row = _write_section(
                worksheet,
                chunk,
                row,
                layout_sheet=layout_sheet,
                limits=limits,
            )
            _add_chart(worksheet, chunk, header_row, end_row)
            data_cursor += take
            if data_cursor < len(section.rows):
                worksheet = _new_worksheet(workbook, f"{layout_sheet.title}_{len(workbook.worksheets) + 1}", used_names, limits)
                row = 1
    for worksheet in workbook.worksheets:
        if worksheet.title == layout_sheet.title or worksheet.title.startswith(f"{layout_sheet.title}_"):
            if layout_sheet.freeze_rows:
                worksheet.freeze_panes = f"A{layout_sheet.freeze_rows + 1}"
            if layout_sheet.auto_filter and worksheet.max_row >= 1 and worksheet.max_column >= 1:
                worksheet.auto_filter.ref = f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}"


def _deterministic_zip(raw: bytes) -> bytes:
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(raw), "r") as source, zipfile.ZipFile(
        output, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as target:
        for name in sorted(source.namelist()):
            original = source.getinfo(name)
            info = zipfile.ZipInfo(name, date_time=_FIXED_ZIP_TIME)
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 0
            info.external_attr = original.external_attr
            target.writestr(info, source.read(name))
    return output.getvalue()


def render_workbook_v1(
    report: AnalysisReportV1 | dict[str, Any],
    *,
    exporter_version: str,
    limits: WorkbookLimits,
) -> bytes:
    """从同一 Report Version 的 payload 确定性生成安全 xlsx bytes。"""
    validated = AnalysisReportV1.model_validate(report)
    layout = _effective_layout(validated)
    by_id = {block.id: block for block in validated.blocks}
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "KOL Insight AI"
    workbook.properties.subject = f"workbook_v1:{exporter_version}"
    fixed_time = datetime(2000, 1, 1)
    workbook.properties.created = fixed_time
    workbook.properties.modified = fixed_time
    used_names: set[str] = set()
    for sheet_index, layout_sheet in enumerate(layout.sheets):
        selected_blocks: list[AnalysisReportBlock] = []
        for block_id in layout_sheet.block_ids:
            block = by_id.get(block_id)
            if block is None:
                raise ValueError(f"workbook references unknown block {block_id!r}")
            selected_blocks.append(block)
        _render_layout_sheet(
            workbook,
            validated,
            layout_sheet,
            selected_blocks,
            is_first_sheet=sheet_index == 0,
            used_names=used_names,
            limits=limits,
        )
    if not workbook.worksheets:
        raise WorkbookTechnicalLimitExceeded("workbook must contain at least one sheet")
    raw = BytesIO()
    workbook.save(raw)
    content = _deterministic_zip(raw.getvalue())
    if len(content) > limits.max_bytes:
        raise WorkbookTechnicalLimitExceeded("workbook bytes exceed max_bytes")
    return content


__all__ = [
    "WorkbookLimits",
    "WorkbookTechnicalLimitExceeded",
    "render_workbook_v1",
    "workbook_layout_digest",
]
