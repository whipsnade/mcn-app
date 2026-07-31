"""brand_report_v2 Excel 导出渲染器（Task 7）。

模板 `templates/brand_report_v2.xlsx` 由 `scripts/build_brand_report_template.py`
一次性生成（样例数据行已清除，保留表头/合并拓扑/列宽/数字格式）。渲染器把
`BrandReportPayload` 按固定表头行写入各 Sheet：占比直接写 payload 已算数值
（不写跨行公式），缺失单元格写「未提供」，日趋势/地域图表仅在数据存在时新建，
空章节保留 Sheet + 列头 + availability.reason 受限说明。任何异常向上抛，
由端点映射为明确错误，绝不输出半截文件。
"""

from __future__ import annotations

import asyncio
import re
from io import BytesIO
from pathlib import Path
from typing import Any, Sequence
from urllib.parse import urlparse

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.reporting.brand_payload import BrandReportPayload
from app.reporting.models import AnalysisReport
from app.selection.exporter import ExportedWorkbook
from app.workspace.models import WorkspaceSession

TEMPLATE_PATH = Path(__file__).with_name("templates") / "brand_report_v2.xlsx"
CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SHEET_ORDER = (
    "综合概览", "情感分析", "日趋势", "内容类型与达人",
    "地域分布", "热门帖子TOP", "舆情洞察", "方法论",
)

MISSING = "未提供"
PCT_FORMAT = '0.00"%"'

TOP_POST_HEADERS = (
    "排名", "平台", "标题", "用户昵称", "互动数", "阅读数",
    "点赞", "评论", "收藏", "转发", "情感", "达人层级", "链接",
)

_CHAPTER_LABELS = {
    "overview": "综合概览",
    "sentiment": "情感分析",
    "daily_trend": "日趋势",
    "content_creators": "内容类型与达人",
    "regions": "地域分布",
    "top_posts": "热门帖子",
    "insights": "舆情洞察",
    "methodology": "方法论",
}

_INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_WHITESPACE_RUN = re.compile(r"\s+")
_UNDERSCORE_RUN = re.compile(r"_+")
_FORMULA_PREFIXES = ("=", "+", "-", "@")
_PLATFORM_LABELS = {
    "xiaohongshu": "小红书",
    "douyin": "抖音",
    "bilibili": "哔哩哔哩",
    "weibo": "微博",
    "wechat": "微信",
}
# 热门帖子 Sheet 平台段顺序：小红书段 + 抖音段在前，其余平台按名称排后。
_PLATFORM_ORDER = ("xiaohongshu", "douyin")


def sanitize_report_filename(brand: str, start: str, end: str, version: int) -> str:
    """剔除控制字符与 < > : " / \\ | ? *；合并连续空白/下划线；去尾部点和空格。

    品牌片段截 80 字符；清洗后为空用「未命名品牌」。
    """
    cleaned = _INVALID_FILENAME_CHARS.sub("", brand or "")
    cleaned = _WHITESPACE_RUN.sub(" ", cleaned)
    cleaned = _UNDERSCORE_RUN.sub("_", cleaned).strip()
    cleaned = cleaned[:80].rstrip(". ")
    if not cleaned:
        cleaned = "未命名品牌"
    return f"{cleaned}_品牌社媒分析报告_{start}-{end}_v{version}.xlsx"


async def export_brand_report(
    db: AsyncSession, user_id: str, session_id: str, report_id: str
) -> ExportedWorkbook:
    """归属与类型校验后渲染品牌报告 Excel；全部校验失败统一 LookupError。

    不调用模型/MCP/积分系统。文件名取 payload.scope 的品牌与周期 + report.version。
    """
    session = await db.scalar(
        select(WorkspaceSession).where(
            WorkspaceSession.id == session_id,
            WorkspaceSession.user_id == user_id,
            WorkspaceSession.deleted_at.is_(None),
        )
    )
    if session is None:
        raise LookupError("report_not_found")
    report = await db.get(AnalysisReport, report_id)
    if (
        report is None
        or report.session_id != session_id
        or report.report_type != "brand_analysis"
        or report.template_version != "brand_report_v2"
        or not isinstance(report.payload_json, dict)
    ):
        raise LookupError("report_not_found")
    try:
        payload = BrandReportPayload.model_validate(report.payload_json)
    except ValidationError as error:
        raise LookupError("report_not_found") from error
    scope = payload.scope
    filename = sanitize_report_filename(
        scope.brand,
        scope.period_start or "未指定",
        scope.period_end or "未指定",
        report.version,
    )
    # openpyxl 渲染是 CPU 密集的同步操作，放线程避免阻塞事件循环。
    content = await asyncio.to_thread(render_brand_workbook, payload)
    return ExportedWorkbook(content=content, filename=filename)


def render_brand_workbook(payload: BrandReportPayload) -> bytes:
    """同步 openpyxl 渲染（调用方 asyncio.to_thread）。"""
    workbook = load_workbook(TEMPLATE_PATH)
    _render_overview(workbook["综合概览"], payload)
    _render_sentiment(workbook["情感分析"], payload)
    _render_daily_trend(workbook["日趋势"], payload)
    _render_content_creators(workbook["内容类型与达人"], payload)
    _render_regions(workbook["地域分布"], payload)
    _render_top_posts(workbook["热门帖子TOP"], payload)
    _render_insights(workbook["舆情洞察"], payload)
    _render_methodology(workbook["方法论"], payload)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


# ---------- 各 Sheet 渲染 ----------


def _render_overview(sheet: Any, payload: BrandReportPayload) -> None:
    scope = payload.scope
    overview = payload.data.overview
    sheet["A1"] = _cell_value(f"{scope.brand or '未命名品牌'} 品牌社交媒体表现分析报告")
    period = f"{scope.period_start or '未指定'} 至 {scope.period_end or '未指定'}"
    if scope.data_as_of:
        period += f"（数据截至 {scope.data_as_of}）"
    sheet["B2"] = period
    sheet["B3"] = f"DataTap 聆媒洞察 MCP（{_platforms_label(scope.platforms)}）"
    sheet["B4"] = _cell_value(_query_label(payload))
    platforms = overview.platforms
    headers = ["指标"] + [_platform_label(item.platform) for item in platforms] + ["合计"]
    for column in range(1, 7):
        sheet.cell(6, column).value = headers[column - 1] if column <= len(headers) else None
    metrics = (("声量(帖数)", "mentions"), ("互动数", "interactions"), ("阅读/播放数", "exposure"))
    for offset, (label, field) in enumerate(metrics):
        row = 7 + offset
        sheet.cell(row, 1).value = label
        values: list[float | None] = []
        for index, item in enumerate(platforms):
            value = getattr(item, field)
            values.append(value)
            sheet.cell(row, 2 + index).value = _num(value)
        sheet.cell(row, 2 + len(platforms)).value = _total(values)
        for column in range(3 + len(platforms), 7):
            sheet.cell(row, column).value = None
    # 模板行 10-15 是样例独有的细分指标（用户数/点赞等），payload 无此口径，整行清除。
    _clear_rows(sheet, 10, 15, 6)
    reason = _chapter_reason(payload, "overview")
    sheet["A16"] = reason
    comparisons = (
        ("声量", overview.total_mentions),
        ("互动数", overview.total_interactions),
        ("阅读/播放数", overview.total_exposure),
    )
    for offset, (label, comparison) in enumerate(comparisons):
        row = 19 + offset
        sheet.cell(row, 1).value = label
        sheet.cell(row, 2).value = _num(comparison.current)
        _write_pct(sheet.cell(row, 3), comparison.mom_change_pct)
        _write_pct(sheet.cell(row, 4), comparison.yoy_change_pct)
    _clear_rows(sheet, 22, 22, 6)
    split = overview.sentiment_split
    parts = (("正面声量", split.positive), ("中性声量", split.neutral), ("负面声量", split.negative))
    known = [value for _, value in parts if value is not None]
    total = sum(known) if len(known) == len(parts) else None
    for offset, (label, value) in enumerate(parts):
        row = 26 + offset
        sheet.cell(row, 1).value = label
        sheet.cell(row, 2).value = _num(value)
        share = round(value / total * 100, 2) if value is not None and total else None
        _write_pct(sheet.cell(row, 3), share)


def _render_sentiment(sheet: Any, payload: BrandReportPayload) -> None:
    _unmerge_below_row(sheet, 1)
    _clear_rows(sheet, 4, max(sheet.max_row, 20), 5)
    rows = payload.data.sentiment.rows
    for index, item in enumerate(rows):
        row = 4 + index
        sheet.cell(row, 1).value = _platform_label(item.platform)
        sheet.cell(row, 2).value = _cell_value(item.sentiment)
        sheet.cell(row, 3).value = _num(item.mentions)
        sheet.cell(row, 4).value = _num(item.interactions)
        _write_pct(sheet.cell(row, 5), item.share_pct)
    next_row = 4 + len(rows) + 1
    reason = _chapter_reason(payload, "sentiment")
    if reason:
        sheet.cell(next_row, 1).value = reason
        next_row += 1
    sheet.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=5)
    title_cell = sheet.cell(next_row, 1)
    title_cell.value = "关键发现"
    title_cell.font = Font(name="微软雅黑", bold=True, size=11)
    findings = payload.narrative.key_findings if payload.narrative else []
    if not findings:
        sheet.cell(next_row + 1, 1).value = MISSING
    for index, finding in enumerate(findings, start=1):
        sheet.cell(next_row + index, 1).value = _cell_value(f"{index}. {finding}")


def _render_daily_trend(sheet: Any, payload: BrandReportPayload) -> None:
    scope = payload.scope
    sheet["A1"] = (
        f"每日声量与互动趋势 ({scope.period_start or '未指定'} 至 {scope.period_end or '未指定'})"
    )
    _clear_rows(sheet, 4, max(sheet.max_row, 40), 4)
    sheet._charts = []
    trend = payload.data.daily_trend
    if not trend.points:
        # 无数据不建图，Sheet 内写受限说明。
        sheet["A4"] = _chapter_reason(payload, "daily_trend") or MISSING
        return
    for index, point in enumerate(trend.points):
        row = 4 + index
        sheet.cell(row, 1).value = point.date
        sheet.cell(row, 2).value = _num(point.mentions)
        sheet.cell(row, 3).value = _num(point.interactions)
        note = "声量峰值" if trend.peak_date and point.date == trend.peak_date else None
        sheet.cell(row, 4).value = note
    total_row = 4 + len(trend.points)
    sheet.cell(total_row, 1).value = "合计"
    sheet.cell(total_row, 2).value = _total([point.mentions for point in trend.points])
    sheet.cell(total_row, 3).value = _total([point.interactions for point in trend.points])
    last_data_row = 3 + len(trend.points)
    mentions_chart = _line_chart(sheet, "每日声量趋势", column=2, last_row=last_data_row)
    sheet.add_chart(mentions_chart, f"A{total_row + 2}")
    interactions_chart = _line_chart(sheet, "每日互动数趋势", column=3, last_row=last_data_row)
    sheet.add_chart(interactions_chart, f"A{total_row + 17}")


def _render_content_creators(sheet: Any, payload: BrandReportPayload) -> None:
    _unmerge_below_row(sheet, 1)
    _clear_rows(sheet, 3, max(sheet.max_row, 30), 4)
    data = payload.data
    next_row = _write_table(
        sheet, 3, None, ("内容类型", "声量", "占比"),
        [[item.content_type, _num(item.mentions), item.share_pct] for item in data.content_types],
        pct_columns=(3,),
    )
    next_row = _write_table(
        sheet, next_row + 1, "达人层级分布 (品牌相关内容)", ("达人层级", "声量", "占比"),
        [[item.tier, _num(item.mentions), item.share_pct] for item in data.creator_tiers],
        pct_columns=(3,),
    )
    organic = data.organic_vs_paid
    _write_table(
        sheet, next_row + 1, "商单 vs 自然内容", ("类型", "声量", "占比"),
        [
            ["自然内容(非商单)", _num(organic.organic_mentions), organic.organic_share_pct],
            ["商单内容", _num(organic.paid_mentions), organic.paid_share_pct],
        ],
        pct_columns=(3,),
    )
    reason = _chapter_reason(payload, "content_creators")
    if reason:
        sheet.cell(2, 1).value = reason


def _render_regions(sheet: Any, payload: BrandReportPayload) -> None:
    _clear_rows(sheet, 4, max(sheet.max_row, 30), 4)
    sheet._charts = []
    regions = payload.data.regions
    if not regions:
        sheet["A4"] = _chapter_reason(payload, "regions") or MISSING
        return
    for index, item in enumerate(regions):
        row = 4 + index
        sheet.cell(row, 1).value = _cell_value(item.region)
        sheet.cell(row, 2).value = _num(item.mentions)
        sheet.cell(row, 3).value = _num(item.interactions)
        _write_pct(sheet.cell(row, 4), item.share_pct)
    last_data_row = 3 + len(regions)
    chart = BarChart()
    chart.type = "col"
    chart.title = "发帖用户地域声量分布"
    chart.add_data(
        Reference(sheet, min_col=2, min_row=3, max_row=last_data_row), titles_from_data=True
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=4, max_row=last_data_row))
    chart.height = 9
    chart.width = 16
    sheet.add_chart(chart, f"A{last_data_row + 2}")


def _render_top_posts(sheet: Any, payload: BrandReportPayload) -> None:
    """双段布局：小红书段 + 抖音段顺序排列，按平台动态重写（unmerge/clear/rewrite）。"""
    for merged in list(sheet.merged_cells.ranges):
        sheet.unmerge_cells(str(merged))
    _clear_rows(sheet, 1, max(sheet.max_row, 45), len(TOP_POST_HEADERS))
    sheet.column_dimensions["M"].width = 40
    posts = payload.data.top_posts
    if not posts:
        sheet.merge_cells("A1:M1")
        sheet["A1"] = "热门帖子 TOP（品牌相关，按互动数排序）"
        sheet["A3"] = _chapter_reason(payload, "top_posts") or MISSING
        return
    ordered_platforms = sorted(
        {post.platform for post in posts},
        key=lambda name: (
            _PLATFORM_ORDER.index(name) if name in _PLATFORM_ORDER else len(_PLATFORM_ORDER),
            name,
        ),
    )
    row = 1
    for platform in ordered_platforms:
        group = [post for post in posts if post.platform == platform]
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
        title_cell = sheet.cell(row, 1)
        title_cell.value = (
            f"{_platform_label(platform)}热门帖子 TOP {len(group)}（品牌相关，按互动数排序）"
        )
        title_cell.font = Font(name="微软雅黑", bold=True, size=12)
        header_row = row + 1
        for column, header in enumerate(TOP_POST_HEADERS, start=1):
            cell = sheet.cell(header_row, column)
            cell.value = header
            cell.font = Font(name="微软雅黑", bold=True, size=10)
        for rank, post in enumerate(group, start=1):
            data_row = header_row + rank
            values = (
                rank,
                _platform_label(post.platform),
                _present(post.title),
                _present(post.author),
                _num(post.interactions),
                _num(post.exposure_count),
                _num(post.like_count),
                _num(post.comment_count),
                _num(post.collect_count),
                _num(post.share_count),
                _present(post.sentiment),
                _present(post.creator_tier),
            )
            for column, value in enumerate(values, start=1):
                sheet.cell(data_row, column).value = _cell_value(value)
            url_cell = sheet.cell(data_row, 13)
            if post.url and _is_valid_url(post.url):
                url_cell.value = post.url
                url_cell.hyperlink = post.url
                url_cell.font = Font(color="0563C1", underline="single")
            else:
                # url 列只写合法 URL，否则「未提供」，绝不拼接猜测。
                url_cell.value = MISSING
        row = header_row + len(group) + 2


def _render_insights(sheet: Any, payload: BrandReportPayload) -> None:
    _unmerge_below_row(sheet, 1)
    _clear_rows(sheet, 2, max(sheet.max_row, 40), 4)
    narrative = payload.narrative
    next_row = 2
    reason = _chapter_reason(payload, "insights")
    if reason:
        next_row = _write_paragraph(sheet, next_row, reason) + 1
    sections = (
        ("用户好评集中点", narrative.praise_points if narrative else []),
        ("用户槽点与负面反馈", narrative.complaint_points if narrative else []),
        ("品牌扩张信号", narrative.expansion_signals if narrative else []),
    )
    for title, items in sections:
        next_row = _write_list_section(sheet, next_row, title, items)
    impact = narrative.impact_level if narrative else None
    next_row = _write_list_section(
        sheet, next_row, "负面影响程度", [f"负面影响程度：{impact}"] if impact else [],
        numbered=False,
    )
    noise = narrative.noise_notes if narrative else None
    _write_list_section(sheet, next_row, "噪音说明", [noise] if noise else [], numbered=False)


def _render_methodology(sheet: Any, payload: BrandReportPayload) -> None:
    """由 payload.scope/query_spec/sources + comparison_mode 口径生成。"""
    _clear_rows(sheet, 4, max(sheet.max_row, 20), 2)
    scope = payload.scope
    query = payload.query_spec
    period = f"{scope.period_start or '未指定'} 至 {scope.period_end or '未指定'}"
    if scope.data_as_of:
        period += f"（数据截至 {scope.data_as_of}）"
    if query.matched_tag:
        match_label = f'品牌标签匹配："{query.matched_tag}"'
    else:
        match_label = f'关键词搜索："{query.fallback_keyword or query.original_term or MISSING}"'
    comparison = query.comparison_definition or (
        "环比 + 同比（与上一同等时长周期及去年同期对比）"
        if scope.comparison_mode == "mom_yoy"
        else "环比（与上一同等时长周期对比）"
    )
    tools = "、".join(dict.fromkeys(source.tool for source in payload.sources))
    rows = (
        ("数据来源", f"DataTap 聆媒洞察 MCP（{_platforms_label(scope.platforms)}）"),
        ("分析品牌", query.original_term or scope.brand or MISSING),
        ("匹配方式", match_label),
        ("时间范围", period),
        ("对比口径", comparison),
        ("数据工具", tools or MISSING),
        ("章节可用性", _availability_summary(payload)),
        (
            "局限性",
            "1. 关键词/标签搜索可能遗漏未直接提及品牌的内容\n"
            "2. 情感标注为 AI 自动分类，存在少量误判\n"
            "3. 占比与变化百分比由服务端按已采集数值计算，受限周期不参与计算",
        ),
        ("报告生成", "KOL Insight AI 品牌报告导出（模板 brand_report_v2）"),
    )
    for index, (label, text) in enumerate(rows):
        row = 4 + index
        sheet.cell(row, 1).value = label
        value_cell = sheet.cell(row, 2)
        value_cell.value = _cell_value(text)
        value_cell.alignment = Alignment(wrap_text=True, vertical="top")


# ---------- 写入辅助 ----------


def _write_table(
    sheet: Any,
    start_row: int,
    title: str | None,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    *,
    pct_columns: Sequence[int] = (),
    columns: int = 4,
) -> int:
    """写一个 [可选标题 + 表头 + 数据行] 小节，返回下一空行行号。"""
    row = start_row
    if title is not None:
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
        title_cell = sheet.cell(row, 1)
        title_cell.value = title
        title_cell.font = Font(name="微软雅黑", bold=True, size=11)
        row += 1
    for column, header in enumerate(headers, start=1):
        sheet.cell(row, column).value = header
    row += 1
    if not rows:
        sheet.cell(row, 1).value = MISSING
        return row + 1
    for values in rows:
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row, column)
            if column in pct_columns:
                _write_pct(cell, value if isinstance(value, int | float) else None)
            else:
                cell.value = _cell_value(value)
        row += 1
    return row


def _write_list_section(
    sheet: Any, start_row: int, title: str, items: Sequence[str], *, numbered: bool = True
) -> int:
    """写一个 [标题 + 逐条列表] 小节（4 列宽合并行），返回下一空行行号。"""
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)
    title_cell = sheet.cell(start_row, 1)
    title_cell.value = title
    title_cell.font = Font(name="微软雅黑", bold=True, size=11)
    row = start_row + 1
    if not items:
        sheet.cell(row, 1).value = MISSING
        return row + 2
    for index, item in enumerate(items, start=1):
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = sheet.cell(row, 1)
        cell.value = _cell_value(f"{index}. {item}" if numbered else item)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    return row + 1


def _write_paragraph(sheet: Any, row: int, text: str) -> int:
    sheet.cell(row, 1).value = _cell_value(text)
    sheet.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
    return row + 1


def _line_chart(sheet: Any, title: str, *, column: int, last_row: int) -> LineChart:
    chart = LineChart()
    chart.title = title
    chart.add_data(
        Reference(sheet, min_col=column, min_row=3, max_row=last_row), titles_from_data=True
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=4, max_row=last_row))
    chart.height = 7
    chart.width = 15
    return chart


# ---------- 取值与格式 ----------


def _platform_label(platform: str) -> str:
    return _PLATFORM_LABELS.get(platform, platform)


def _platforms_label(platforms: Sequence[str]) -> str:
    return "、".join(_platform_label(name) for name in platforms) or MISSING


def _query_label(payload: BrandReportPayload) -> str:
    query = payload.query_spec
    if query.matched_tag:
        return f'品牌标签匹配: "{query.matched_tag}"'
    return f'关键词搜索: "{query.fallback_keyword or query.original_term or MISSING}"'


def _chapter_reason(payload: BrandReportPayload, chapter: str) -> str | None:
    availability = payload.availability.get(chapter)
    if availability and availability.status != "complete" and availability.reason:
        return f"数据受限：{availability.reason}"
    return None


def _availability_summary(payload: BrandReportPayload) -> str:
    limited = []
    for chapter, availability in payload.availability.items():
        if availability.status == "complete":
            continue
        label = _CHAPTER_LABELS.get(chapter, chapter)
        detail = f"（{availability.reason}）" if availability.reason else ""
        limited.append(f"{label}：{availability.status}{detail}")
    return "；".join(limited) if limited else "全部章节数据完整"


def _num(value: float | int | None) -> Any:
    return value if value is not None else MISSING


def _total(values: Sequence[float | None]) -> Any:
    known = [value for value in values if value is not None]
    return sum(known) if known else MISSING


def _write_pct(cell: Any, value: float | None) -> None:
    """占比/变化百分比直接写 payload 已算数值（百分数单位），不写跨行公式。"""
    if value is None:
        cell.value = MISSING
        cell.number_format = "General"
        return
    cell.value = value
    cell.number_format = PCT_FORMAT


def _present(value: Any) -> Any:
    if value is None or (isinstance(value, str) and not value.strip()):
        return MISSING
    return value


def _cell_value(value: Any) -> Any:
    """第三方可控文本以 = + - @ 开头时前缀 ' 转义，防公式注入。"""
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _is_valid_url(url: str) -> bool:
    parsed = urlparse(url)
    return parsed.scheme in ("http", "https") and bool(parsed.netloc)


def _clear_rows(sheet: Any, start: int, end: int, columns: int) -> None:
    for row in range(start, end + 1):
        for column in range(1, columns + 1):
            cell = sheet.cell(row, column)
            if not isinstance(cell, MergedCell):
                cell.value = None


def _unmerge_below_row(sheet: Any, keep_row: int) -> None:
    """保留 keep_row 行（含）以上的合并拓扑，解除其下的合并供动态重写。"""
    for merged in list(sheet.merged_cells.ranges):
        if merged.min_row > keep_row:
            sheet.unmerge_cells(str(merged))


__all__ = [
    "CONTENT_TYPE",
    "SHEET_ORDER",
    "TEMPLATE_PATH",
    "export_brand_report",
    "render_brand_workbook",
    "sanitize_report_filename",
]
